"""Ozon FBS cargo-place modal: warehouse_date formatting + expand details."""

from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch

from review_processor import ozon_fbs as oz
from review_processor import ozon_fbs_containers as ct


class FormatWarehouseDateTests(unittest.TestCase):
    def test_date_only(self) -> None:
        self.assertEqual(oz.format_warehouse_date("2026-08-30"), "30.08.2026")

    def test_naive_iso_keeps_local_clock(self) -> None:
        # Must NOT shift warehouse-local 15:30 as if it were UTC.
        self.assertEqual(oz.format_warehouse_date("2026-03-20T15:30:00"), "20.03.2026 15:30")

    def test_space_separated_local(self) -> None:
        self.assertEqual(oz.format_warehouse_date("2026-03-20 09:05:00"), "20.03.2026 09:05")

    def test_empty(self) -> None:
        self.assertEqual(oz.format_warehouse_date(""), "")
        self.assertEqual(oz.format_warehouse_date(None), "")


class ContainerModalDetailsTests(unittest.TestCase):
    def test_timeline_includes_local_move_and_warehouse_date(self) -> None:
        repo = MagicMock()
        with patch.object(
            ct,
            "get_supply_moved_to_delivering_at",
            return_value="2026-03-21T10:00:00+00:00",
        ), patch.object(
            ct,
            "_list_local_container_postings",
            return_value=[
                {
                    "posting_number": "12345-0001",
                    "offer_id": "ART",
                    "product_name": "Товар",
                    "quantity": 1,
                    "has_kiz": True,
                    "kiz_count": 1,
                    "pick_verified": False,
                    "container_barcode": "99",
                    "container_synced": True,
                    "tab": "delivering",
                    "status": "delivering",
                }
            ],
        ):
            out = ct.build_container_modal_details(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                container={
                    "container_id": 99,
                    "container_number": 3,
                    "status": "acceptance_in_progress",
                    "status_label": "Принято на СЦ",
                    "created_at": "2026-03-19T08:00:00Z",
                    "warehouse_date": "2026-03-20T15:30:00",
                },
            )
        keys = [x["key"] for x in out["timeline"]]
        self.assertEqual(
            keys, ["created", "moved_to_delivering", "warehouse_date", "status"]
        )
        self.assertEqual(out["postings_count"], 1)
        self.assertEqual(out["warehouse_date_display"], "20.03.2026 15:30")
        self.assertTrue(out["moved_to_delivering_at_display"])

    def test_enrich_list_attaches_display_fields(self) -> None:
        repo = MagicMock()
        listed = {
            "ok": True,
            "items": [
                {
                    "container_id": 1,
                    "status": "new",
                    "warehouse_date": "2026-03-20T15:30:00",
                    "created_at": "2026-03-19T08:00:00Z",
                }
            ],
        }
        with patch.object(
            ct, "get_supply_moved_to_delivering_at", return_value=""
        ):
            out = ct.enrich_containers_for_supply_modal(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                listed=listed,
            )
        row = out["items"][0]
        self.assertEqual(row["warehouse_date_display"], "20.03.2026 15:30")
        self.assertTrue(row["created_at_display"])
        self.assertEqual(row["moved_to_delivering_at"], "")

    def test_get_details_is_local_only(self) -> None:
        repo = MagicMock()
        with patch.object(ct, "build_container_modal_details") as build:
            build.return_value = {"ok": True, "container_id": 7}
            out = ct.get_container_modal_details(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                container_id=7,
                container={"container_id": 7, "status": "new"},
            )
        self.assertTrue(out["ok"])
        build.assert_called_once()
        kwargs = build.call_args.kwargs
        self.assertEqual(kwargs["container"]["container_id"], 7)
        self.assertIsNone(kwargs.get("client"))

    def test_build_details_prefers_richer_warehouse_date_from_get(self) -> None:
        repo = MagicMock()
        client = MagicMock()
        with patch.object(
            ct,
            "get_supply_moved_to_delivering_at",
            return_value="",
        ), patch.object(
            ct,
            "_list_local_container_postings",
            return_value=[],
        ), patch.object(
            ct,
            "_fetch_container_postings",
            return_value=(
                {
                    "container_id": 99,
                    "warehouse_date": "2026-09-01T14:25:00",
                    "created_at": "2026-09-01T11:00:00Z",
                    "status": "new",
                    "status_label": "Новое",
                },
                [],
                True,
            ),
        ):
            out = ct.build_container_modal_details(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                container={
                    "container_id": 99,
                    "status": "new",
                    "warehouse_date": "2026-09-01",
                    "created_at": "2026-09-01T11:00:00Z",
                },
                client=client,
            )
        self.assertEqual(out["warehouse_date"], "2026-09-01T14:25:00")
        self.assertEqual(out["warehouse_date_display"], "01.09.2026 14:25")
        wh_ev = next(x for x in out["timeline"] if x["key"] == "warehouse_date")
        self.assertEqual(wh_ev["at_display"], "01.09.2026 14:25")
        self.assertFalse(wh_ev.get("hint"))

    def test_build_details_merges_ozon_membership(self) -> None:
        repo = MagicMock()
        client = MagicMock()
        with patch.object(
            ct,
            "get_supply_moved_to_delivering_at",
            return_value="",
        ), patch.object(
            ct,
            "_list_local_container_postings",
            return_value=[],
        ), patch.object(
            ct,
            "_fetch_container_postings",
            return_value=(
                {"container_id": 99},
                ["A-1", "A-2"],
                True,
            ),
        ), patch.object(
            ct,
            "_merge_ozon_container_composition",
            return_value=[
                {"posting_number": "A-1", "product_name": "One"},
                {"posting_number": "A-2", "product_name": "Two"},
            ],
        ) as merge:
            out = ct.build_container_modal_details(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                container={"container_id": 99, "status": "new", "status_label": "Новое"},
                client=client,
            )
        self.assertEqual(out["postings_count"], 2)
        self.assertEqual(out["composition_source"], "ozon")
        self.assertTrue(out["ozon_fetch_ok"])
        merge.assert_called_once()
        self.assertEqual(
            merge.call_args.kwargs["ozon_posting_numbers"], ["A-1", "A-2"]
        )

    def test_merge_ozon_composition_soft_binds_unbound_supply_postings(self) -> None:
        repo = MagicMock()
        with patch.object(
            ct,
            "_load_posting_items_by_numbers",
            return_value={
                "A-1": {
                    "posting_number": "A-1",
                    "product_name": "One",
                    "offer_id": "",
                    "quantity": 1,
                    "has_kiz": False,
                    "kiz_count": 0,
                    "pick_verified": False,
                    "container_barcode": "",
                    "container_synced": False,
                    "tab": "",
                    "status": "",
                }
            },
        ), patch.object(
            ct.oz_sup,
            "get_supply",
            return_value={"posting_numbers": ["A-1", "A-2"]},
        ), patch.object(
            ct,
            "load_container_bind_map",
            return_value={"A-1": {"container_id": None}, "A-2": {"container_id": 5}},
        ), patch.object(ct, "_set_local_container_bind") as set_bind:
            out = ct._merge_ozon_container_composition(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                container_id=99,
                ozon_posting_numbers=["A-1", "A-2", "X-9"],
                local_postings=[],
            )
        self.assertEqual([p["posting_number"] for p in out], ["A-1", "A-2", "X-9"])
        self.assertTrue(out[2].get("outside_supply"))
        set_bind.assert_called_once()
        self.assertEqual(set_bind.call_args.kwargs["posting_number"], "A-1")
        self.assertEqual(set_bind.call_args.kwargs["container_id"], 99)

    def test_composition_xlsx_columns(self) -> None:
        repo = MagicMock()
        with patch.object(
            ct,
            "get_container_modal_details",
            return_value={
                "container_id": 42,
                "container_number": 3,
                "warehouse_date_display": "01.09.2026",
                "warehouse_date": "2026-09-01",
                "postings": [
                    {"posting_number": "016-1"},
                    {"posting_number": "016-3"},
                    {"posting_number": ""},
                ],
            },
        ):
            payload, fname = ct.build_container_composition_xlsx(
                repo,
                user_id=1,
                source_id=2,
                supply_id="sup-1",
                container_id=42,
            )
        self.assertTrue(fname.startswith("GM-42-N3-sostav"))
        self.assertTrue(fname.endswith(".xlsx"))
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(payload))
        ws = wb.active
        rows = [
            tuple(c.value for c in row)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            if any(c.value not in (None, "") for c in row)
        ]
        self.assertEqual(rows[0], ("Отправление", "Дата склада (Ozon)"))
        self.assertEqual(rows[1], ("016-1", "01.09.2026"))
        self.assertEqual(rows[2], ("016-3", "01.09.2026"))
        self.assertEqual(len(rows), 3)


if __name__ == "__main__":
    unittest.main()
