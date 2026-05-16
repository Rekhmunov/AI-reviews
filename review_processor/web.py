from __future__ import annotations

from collections.abc import Mapping
from datetime import UTC, datetime, timedelta
from html import escape
import json
import ipaddress
import io
import csv
import logging
from pathlib import Path
import re
import secrets
import threading
import time
from urllib.parse import urlparse
import urllib.error
import urllib.request

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .auth import create_session_token, hash_password, verify_password

_log = logging.getLogger(__name__)
from .config import AppConfig, load_app_config
from .repository import ReviewRepository
from .service import MarketplaceSyncError, ReviewAutomationService, _normalize_timestamp, _parse_ozon_message_text, _wb_image_url
from .models import ReviewInput
from .stock_service import StockScheduler, sync_stock_source

try:  # pragma: no cover - optional in sqlite-only environments
    import psycopg  # type: ignore
except Exception:  # pragma: no cover
    psycopg = None

CATEGORIES = [
    "negative_delivery",
    "negative_product",
    "negative_other",
    "positive_quality",
    "positive_product",
    "neutral_other",
]

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "web_templates"
STATIC_DIR = BASE_DIR / "web_static"

TEMPLATE_GROUPS: list[dict[str, object]] = [
    {
        "id": "positive",
        "title": "Позитив",
        "subgroups": [
            "Общий",
            "Вкус",
            "Материал",
            "Общий позитив",
            "Позитив доставка",
            "Позитив запах",
            "Позитив конструкция",
            "Позитив упаковка",
            "Позитив цвет",
            "Эффект",
        ],
    },
    {
        "id": "product_dissatisfaction",
        "title": "Недовольство товаром",
        "subgroups": [
            "Общий",
            "Брак и Б/У",
            "Высокая цена",
            "Качество",
            "Негатив запах",
            "Негатив конструкция",
            "Негатив цвет",
            "Не подошел лично мне",
            "Не соответствует фото",
            "Не устраивает эффект",
            "Общий негатив",
            "Побочные эффекты",
            "Подделка",
            "Срок годности",
            "Текстура, консистенция, материал",
        ],
    },
    {
        "id": "delivery_problems",
        "title": "Проблемы при доставке",
        "subgroups": [
            "Общий",
            "Долгая доставка",
            "Испорченная упаковка",
            "Наклейка",
            "Недостающая упаковка / грязное / поврежденное и сломанное",
            "Некомплект",
            "Не тот товар",
            "Общие доставка",
        ],
    },
    {
        "id": "wrong_size",
        "title": "Неправильный размер",
        "subgroups": [
            "Общий",
            "Альтернативные измерения",
            "Большемерит/маломерит",
            "Не подошел размер",
        ],
    },
    {
        "id": "textless_ratings",
        "title": "Оценки без текста",
        "subgroups": [
            "1-3 звезды",
            "4-5 звезд",
        ],
    },
]

TEXTLESS_RATINGS_GROUP_ID = "textless_ratings"
TEXTLESS_LOCKED_SUBGROUPS: tuple[str, ...] = ("1-3 звезды", "4-5 звезд")
GENERAL_LOCKED_SUBGROUP = "Общий"
GENERAL_LOCKED_GROUP_IDS: tuple[str, ...] = (
    "positive",
    "product_dissatisfaction",
    "delivery_problems",
    "wrong_size",
)

DEFAULT_TEMPLATE_CONTENT: dict[str, list[str]] = {
    "Общий": ["Спасибо за ваш отзыв! Мы ценим обратную связь и уже работаем над улучшениями."],
    "Вкус": [
        "%USER%, добрый день. Мы рады, что вы довольны покупкой. Попробуйте еще %%RECO%% — вам точно понравится!",
        "Добрый день, %USER%! Благодарим за доверие и внимание к вкусу нашего продукта.",
        "Здравствуйте, %USER%! Спасибо за высокую оценку. Будем ждать вас снова!",
    ],
    "Материал": ["Спасибо за отзыв! Рады, что материал вам понравился."],
    "Общий позитив": [
        "Приветствуем, %USER%! Благодарим вас за высокую оценку нашей продукции. Ваше мнение очень важно для нас. С уважением, %BRAND%",
        "Здравствуйте, %USER%! Мы искренне благодарны за ваше время и положительный отзыв о нашем продукте. Спасибо, что выбираете нас. Желаем вам прекрасного дня!",
        "Здравствуйте! Спасибо за ваше доверие и положительный отзыв о нашем продукте. Хорошего вам дня! С уважением, команда бренда %BRAND%",
        "%USER%, добрый день! Благодарим за отзыв! Мы всегда рады помочь вам. Прекрасного вам дня!",
        "%USER%, добрый день! Мы рады, что вы остались довольны нашим брендом. Спасибо за вашу поддержку! С надеждой на ваши будущие покупки, %BRAND%",
        "%USER%, добрый день. Мы стараемся предлагать только качественные товары. Прекрасного дня!",
        "%USER%, добрый день! Большое спасибо за приятные слова о нашем бренде %BRAND%. Это важно для нас! С надеждой на ваши будущие покупки, %BRAND%",
        "%USER%, добрый день! Мы рады, что вы довольны нашей продукцией. Благодарим за высокую оценку! С надеждой на ваши будущие покупки, %BRAND%",
        "Благодарим вас за добрые слова и оценку нашего продукта.",
        "Добрый день, спасибо за обратную связь! Рады, что наш продукт оправдал ваши ожидания. Мы дорожим мнением каждого покупателя!",
        "Добрый день! Спасибо за вашу поддержку! Надеемся на долгосрочное сотрудничество. Хорошего настроения!",
        "Добрый день, спасибо за обратную связь! Спасибо за отличный отзыв и оценку нашего продукта.",
        "Добрый день! Благодарим вас за хороший отзыв о нашем продукте. Надеемся, что он приносит вам удовольствие! Мы дорожим мнением каждого покупателя!",
        "Добрый день! Благодарим за ваш отзыв. Мы рекомендуем попробовать %RECO% — это отличный вариант для новых открытий!",
        "Добрый день! Ваше мнение важно для нас и помогает нам развиваться. С уважением, ваш %BRAND%!",
        "Добрый день! Мы рады, что наш товар приносит вам удовольствие. Спасибо за отзыв. Прекрасного дня!",
        "Добрый день! Мы рады, что наша продукция соответствует вашим ожиданиям. Спасибо за отзыв!",
        "Добрый день! Мы ценим ваш отзыв и поддержку. Рекомендуем попробовать и %RECO% — это может принести вам новые впечатления.",
        "Добрый день! Мы ценим ваше доверие к нашему бренду. Попробуйте и %RECO% — вы останетесь довольны результатом.",
        "Добрый день! Мы ценим ваше мнение о нашей продукции. Спасибо за положительный отзыв.",
        "Добрый день! Мы ценим вашу поддержку. Рекомендуем вам попробовать %RECO% — это наш бестселлер!",
        "Добрый день. Мы рады, что вы оценили качество нашего товара. Спасибо за высокую оценку. С уважением, ваш %BRAND%!",
        "Добрый день. Спасибо за ваше доброе отношение к нашему бренду. Ваши слова — лучшая награда для нас. Хорошего дня!",
        "Здравствуйте, %USER%! Очень приятно получить от вас такой отзыв. Спасибо за высокую оценку! С надеждой на ваши будущие покупки, %BRAND%",
        "Здравствуйте, спасибо за обратную связь! Ваше доверие — наше главное признание. Прекрасного вам дня!",
        "Здравствуйте, спасибо за обратную связь! Ваш отзыв вдохновляет нас на новые достижения. С уважением, ваш %BRAND%!",
        "Здравствуйте, спасибо за обратную связь! Мы рады, что вы довольны качеством нашей продукции. Спасибо за отличный отзыв.",
        "Здравствуйте, спасибо за обратную связь! Мы рады, что вы оценили качество нашего товара. Спасибо за высокую оценку. Прекрасного дня!",
        "Здравствуйте, спасибо за обратную связь! Огромное спасибо за положительную оценку и поддержку нашего бренда %BRAND%!",
        "Здравствуйте! Благодарим за вашу поддержку и хороший отзыв о продукции бренда!",
        "Здравствуйте! Благодарим за положительный отзыв о нашем бренде %BRAND%! Мы дорожим мнением каждого покупателя!",
        "Здравствуйте! Большое спасибо за вашу поддержку и высокую оценку нашего бренда %BRAND%!",
        "Здравствуйте! Благодарим за отзыв!",
        "Здравствуйте! Мы рады, что вы остались довольны нашим товаром. Спасибо за отзыв! Мы дорожим мнением каждого покупателя!",
        "Здравствуйте! Мы ценим ваш отзыв и благодарим за высокую оценку нашего продукта. Хорошего дня!",
        "Здравствуйте! Мы ценим ваш отзыв. Попробуйте еще %RECO% — это один из самых популярных товаров бренда.",
        "Здравствуйте! Мы ценим ваше мнение о нашей продукции. Спасибо за отзыв!",
        "Здравствуйте! Рады, что вы остались довольны нашей продукцией. Спасибо за отзыв!",
        "Мы рады, что вы выбрали именно наш бренд. Спасибо за вашу поддержку!",
        "Мы всегда стараемся делать качественные товары. Спасибо за ваш отзыв!",
        "Приветствуем! Благодарим за вашу поддержку и добрые слова о нашем бренде!",
        "Приветствуем! Большое спасибо за высокую оценку нашего бренда!",
        "Приветствуем! Мы рады, что вам нравится наш продукт. Не упустите шанс попробовать и %RECO% — это один из лучших товаров бренда.",
        "Приветствуем! Мы рады, что вы оценили нашу продукцию. Если вы хотите разнообразить выбор, попробуйте %RECO%.",
        "Приветствуем! Мы ценим ваше мнение и благодарим за высокую оценку. Не забудьте попробовать %RECO% — это хит продаж!",
        "Здравствуйте! Очень приятно получить такой отзыв. Спасибо за добрые слова о нашей продукции!",
        "Здравствуйте! Спасибо за вашу поддержку! Мы всегда рады видеть вас в числе наших клиентов. Прекрасного вам дня!",
        "Здравствуйте! Спасибо за добрые слова и поддержку. Прекрасного дня!",
        "Добрый день! Рады, что вы оценили качество нашего товара. Спасибо за вашу поддержку.",
        "Добрый день! Спасибо за высокую оценку нашего продукта! Ваш отзыв — лучшая награда для нас.",
    ],
    "Позитив доставка": ["Спасибо! Очень рады, что доставка прошла отлично."],
    "Позитив запах": ["Спасибо за отзыв! Приятно, что аромат вам понравился."],
    "Позитив конструкция": ["Спасибо! Рады, что конструкция товара вам подошла."],
    "Позитив упаковка": ["Спасибо! Рады, что упаковка вам понравилась."],
    "Позитив цвет": ["Спасибо за высокую оценку! Рады, что цвет вам подошел."],
    "Эффект": ["Спасибо за отзыв! Рады, что вы заметили хороший эффект."],
    "Брак и Б/У": ["Нам очень жаль, что вы получили товар в таком состоянии. Уже разбираемся."],
    "Высокая цена": ["Спасибо за обратную связь. Учтем ваш комментарий по стоимости."],
    "Качество": ["Нам жаль, что качество не оправдало ожиданий. Передали информацию в отдел качества."],
    "Негатив запах": ["Сожалеем о ситуации. Проверим партию и вернемся с ответом."],
    "Негатив конструкция": ["Спасибо за сигнал. Мы уже передали информацию в отдел разработки."],
    "Негатив цвет": ["Сожалеем, что цвет не совпал с ожиданиями. Проверим карточку товара."],
    "Не подошел лично мне": ["Спасибо за отзыв. Нам жаль, что товар вам не подошел."],
    "Не соответствует фото": ["Сожалеем о несоответствии. Передали информацию для проверки карточки."],
    "Не устраивает эффект": ["Спасибо за отзыв. Передали ваше замечание технологам."],
    "Общий негатив": ["Нам очень жаль, что вы остались недовольны. Уже разбираемся с ситуацией."],
    "Побочные эффекты": ["Сожалеем о ситуации. Рекомендуем прекратить использование и написать нам в поддержку."],
    "Подделка": ["Спасибо за сигнал. Мы проведем дополнительную проверку партии."],
    "Срок годности": ["Спасибо за отзыв. Мы проверим товар и условия хранения."],
    "Текстура, консистенция, материал": ["Спасибо за обратную связь. Передали замечание в отдел качества."],
    "Долгая доставка": ["Сожалеем о задержке доставки. Проверим логистику по вашему заказу."],
    "Испорченная упаковка": ["Нам очень жаль. Передали информацию в логистику и отдел упаковки."],
    "Наклейка": ["Спасибо за сигнал. Проверим корректность маркировки."],
    "Недостающая упаковка / грязное / поврежденное и сломанное": [
        "Сожалеем о состоянии товара. Уже разбираемся и улучшим контроль отгрузки."
    ],
    "Некомплект": ["Сожалеем о неполной комплектации. Мы уже передали информацию на склад."],
    "Не тот товар": ["Нам жаль, что пришел не тот товар. Уже разбираемся с отгрузкой."],
    "Общие доставка": ["Спасибо за отзыв о доставке. Учтем замечание и исправим процесс."],
    "Альтернативные измерения": ["Спасибо за отзыв. Дополним информацию по размерам в карточке товара."],
    "Большемерит/маломерит": ["Сожалеем, что размер не подошел. Передадим замечание по размерной сетке."],
    "Не подошел размер": ["Спасибо за обратную связь. Учтем это при обновлении размерной таблицы."],
    "Общие теги": [
        "Спасибо за оценку и выбор тегов {теги}! Нам очень приятно, что вы отметили эти преимущества.",
        "Благодарим за отзыв с тегами {теги}. Ваши отметки помогают нам становиться лучше!",
    ],
    "1-3 звезды": ["Спасибо за оценку. Нам важно ваше мнение — мы улучшаем сервис каждый день."],
    "4-5 звезд": [
        "Спасибо за высокую оценку! Будем рады снова видеть вас среди покупателей.",
        "Спасибо за 5 звезд! Очень рады, что вам все понравилось.",
    ],
}


def _is_protected_default_subgroup(group_id: str, subgroup: str) -> bool:
    clean_group = str(group_id or "").strip()
    clean_subgroup = str(subgroup or "").strip()
    if clean_group == TEXTLESS_RATINGS_GROUP_ID and clean_subgroup in TEXTLESS_LOCKED_SUBGROUPS:
        return True
    if clean_group in GENERAL_LOCKED_GROUP_IDS and clean_subgroup == GENERAL_LOCKED_SUBGROUP:
        return True
    return False


class SyncRequest(BaseModel):
    account_id: int | None = Field(default=None, description="Specific marketplace account ID")
    all_accounts: bool = Field(default=True, description="Sync all active accounts")
    account_ids: list[int] | None = Field(default=None, description="Specific account IDs to sync (from preview checkboxes)")
    total_expected: int | None = Field(default=None, ge=0, description="Expected total items from preview")


class SyncCapabilitiesRequest(BaseModel):
    account_id: int = Field(ge=1, description="Marketplace account ID for capabilities check")


class ChatQuickTemplateCreateRequest(BaseModel):
    template_name: str = Field(min_length=1, max_length=200)
    template_text: str = Field(min_length=1, max_length=2000)


class ChatQuickTemplateUpdateRequest(BaseModel):
    template_name: str = Field(min_length=1, max_length=200)
    template_text: str = Field(min_length=1, max_length=2000)


class ManualReplyRequest(BaseModel):
    operator_name: str = Field(min_length=2, max_length=120)
    response_text: str = Field(min_length=2, max_length=2000)


class AccountCreateRequest(BaseModel):
    marketplace: str = Field(description="wb|ozon|mock")
    account_name: str = Field(min_length=2, max_length=120)
    api_url: str | None = Field(default=None, max_length=2000)
    api_key: str | None = Field(default=None, max_length=2000)
    client_id: str | None = Field(default=None, max_length=200)
    integration: dict[str, object] | None = None


class AccountStatusRequest(BaseModel):
    is_active: bool


class ConversationStatusRequest(BaseModel):
    status: str = Field(description="open|waiting|closed")


class ConversationReplyRequest(BaseModel):
    response_text: str = Field(min_length=1, max_length=4000)
    idempotency_key: str | None = Field(default=None, max_length=120)


class TemplateUpsertRequest(BaseModel):
    category: str
    mode: str = Field(description="auto|manual|ignore")
    template_text: str = Field(max_length=4000)
    is_enabled: bool | None = None


class AISettingsRequest(BaseModel):
    provider: str = Field(description="rules|yandex")
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None
    yandex_model_uri: str | None = None
    group_processors: dict[str, str] | None = None
    default_sync_lookback_days: int = Field(default=7, ge=0, le=365)


class AIConnectionTestRequest(BaseModel):
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None


class AIReviewTestRequest(BaseModel):
    review_text: str = Field(min_length=1, max_length=8000)
    review_rating: int | None = Field(default=None, ge=1, le=5)
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None


class RoleUpdateRequest(BaseModel):
    role: str = Field(description="user|admin|feedback_manager")


class AdminUserCreateRequest(BaseModel):
    email: str = Field(min_length=5, max_length=255)
    password: str = Field(min_length=8, max_length=255)
    role: str = Field(default="user", description="user")
    plan_code: str = Field(default="starter", min_length=2, max_length=100)


class AdminUserPasswordUpdateRequest(BaseModel):
    password: str = Field(min_length=8, max_length=255)


class TenantUserCreateRequest(BaseModel):
    email: str = Field(min_length=5, max_length=255)
    password: str = Field(min_length=8, max_length=255)
    role: str = Field(default="feedback_manager", description="feedback_manager")
    full_name: str | None = Field(default=None, max_length=200)
    permissions: list["ManagerPermissionItemRequest"] = Field(default_factory=list)


class TenantUserRoleUpdateRequest(BaseModel):
    role: str = Field(description="admin|feedback_manager")


class ManagerPermissionItemRequest(BaseModel):
    account_id: int = Field(ge=1)
    can_reviews: bool = False
    can_questions: bool = False
    can_chats: bool = False


class ManagerPermissionsUpdateRequest(BaseModel):
    permissions: list[ManagerPermissionItemRequest] = Field(default_factory=list)


class UserBlockUpdateRequest(BaseModel):
    blocked: bool
    reason: str | None = Field(default=None, max_length=500)


class UserDeleteRequest(BaseModel):
    confirm: bool = False


class SuperAdminSettingsRequest(BaseModel):
    payment_provider: str = Field(default="manual", max_length=80)
    payment_api_key: str | None = Field(default=None, max_length=2000)
    ai_provider: str = Field(description="rules|yandex")
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None
    yandex_model_uri: str | None = None
    group_processors: dict[str, str] | None = None
    use_sync_start_date: bool = False
    sync_start_date: str | None = None
    default_sync_lookback_days: int = Field(default=7, ge=0, le=365)


class TemplateVariableUpsertRequest(BaseModel):
    var_key: str = Field(min_length=3, max_length=120)
    title: str = Field(min_length=1, max_length=255)
    description: str | None = Field(default=None, max_length=2000)
    is_user_editable: bool = False
    source_type: str = Field(default="manual", max_length=40)
    source_path: str | None = Field(default=None, max_length=255)
    default_value: str | None = Field(default=None, max_length=4000)
    is_active: bool = True


class TemplateVariableDeleteRequest(BaseModel):
    var_key: str = Field(min_length=3, max_length=120)


class CreateSupplySourceRequest(BaseModel):
    name: str
    api_key: str


class ToggleSupplySourceRequest(BaseModel):
    is_enabled: bool = True


class SyncSuppliesRequest(BaseModel):
    source_id: int | None = None


class SupplyManualFieldsRequest(BaseModel):
    pass_number: str | None = None
    pallets_count: str | None = None
    driver_name: str | None = None
    notes: str | None = None
    production: str | None = None


class CreateSupplyDriverRequest(BaseModel):
    full_name: str
    documents: str = ""


class CreateSupplyWarehouseRequest(BaseModel):
    warehouse_name: str
    address: str = ""


class UpdateSupplyWarehouseRequest(BaseModel):
    warehouse_name: str
    address: str = ""


class CreateSupplyLegalEntityRequest(BaseModel):
    short_name: str
    full_name: str = ""
    requisites: str = ""
    signatories: str = ""


class UpdateSupplyLegalEntityRequest(BaseModel):
    short_name: str
    full_name: str = ""
    requisites: str = ""
    signatories: str = ""


class UpdateSupplyDriverRequest(BaseModel):
    full_name: str
    documents: str = ""


class ManagerSuppliesAccessRequest(BaseModel):
    can_supplies: bool = False


class UserTemplateVariableValuesSaveRequest(BaseModel):
    values: dict[str, str] = Field(default_factory=dict)


TEMPLATE_VARIABLE_KEY_RE = re.compile(r"^%[A-Z0-9_]{2,50}%$")


class UserSyncSettingsRequest(BaseModel):
    use_sync_start_date: bool = True
    sync_start_date: str | None = None


class TariffPlanUpsertRequest(BaseModel):
    code: str = Field(min_length=2, max_length=100)
    title: str = Field(min_length=2, max_length=200)
    monthly_price: float = Field(default=0)
    limits: dict[str, object] = Field(default_factory=dict)
    is_active: bool = True


class TariffPlanDeleteRequest(BaseModel):
    code: str = Field(min_length=2, max_length=100)


class TenantPlanUpdateRequest(BaseModel):
    owner_user_id: int
    plan_code: str = Field(min_length=2, max_length=100)
    limits_override: dict[str, object] = Field(default_factory=dict)


class UserPlanUpdateRequest(BaseModel):
    plan_code: str = Field(min_length=2, max_length=100)


class PaymentRecordCreateRequest(BaseModel):
    owner_user_id: int
    amount: float
    currency: str = Field(default="RUB", max_length=10)
    status: str = Field(default="pending", max_length=80)
    external_payment_id: str | None = Field(default=None, max_length=255)
    details: dict[str, object] = Field(default_factory=dict)
    paid_at: str | None = None
    months: int = Field(default=1, ge=1, le=36)
    grace_days: int = Field(default=3, ge=0, le=30)


class PaymentRecordDeleteRequest(BaseModel):
    id: int = Field(ge=1)


class ProfileUpdateRequest(BaseModel):
    full_name: str | None = Field(default=None, max_length=200)
    email: str | None = Field(default=None, max_length=255)
    current_password: str | None = Field(default=None, max_length=255)
    new_password: str | None = Field(default=None, max_length=255)
    new_password_repeat: str | None = Field(default=None, max_length=255)
    use_sync_start_date: bool | None = None
    sync_start_date: str | None = None


class ClearReviewsRequest(BaseModel):
    user_id: int | None = None


class ClearConversationsRequest(BaseModel):
    user_id: int | None = None
    kind: str | None = None
    source: str | None = None


class TemplateSubgroupSaveRequest(BaseModel):
    templates: list[str] = Field(default_factory=list)


class TemplateVariantCreateRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    template_text: str = Field(min_length=1, max_length=4000)


class DefaultTemplateSubgroupSaveRequest(BaseModel):
    templates: list[str] = Field(default_factory=list)


class DefaultTemplateVariantCreateRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    template_text: str = Field(min_length=1, max_length=4000)


class DefaultTemplateSubgroupManageRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)


class DefaultTemplateSubgroupRenameRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    new_subgroup: str = Field(min_length=1, max_length=255)


class DefaultTemplateBulkImportRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    templates: list[str] = Field(default_factory=list)


class ProcessingRuleItemRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    action_mode: str = Field(description="template|manual")
    auto_send: bool = False


class ProcessingRulesApplyRequest(BaseModel):
    rules: list[ProcessingRuleItemRequest] = Field(default_factory=list)


class RecommendationRowRequest(BaseModel):
    source_article: str = Field(default="", max_length=255)
    targets_csv: str = Field(default="", max_length=4000)


class RecommendationsSaveRequest(BaseModel):
    rows: list[RecommendationRowRequest] = Field(default_factory=list)


class StockSourceCreateRequest(BaseModel):
    marketplace: str = Field(min_length=1, max_length=20)
    account_name: str = Field(min_length=1, max_length=200)
    api_url: str = Field(default="", max_length=500)
    api_key: str = Field(default="", max_length=2000)
    client_id: str = Field(default="", max_length=200)
    interval_hours: int = Field(default=24, ge=1, le=24)
    retention_days: int = Field(default=30, ge=1, le=365)


class StockSourceUpdateRequest(BaseModel):
    account_name: str | None = None
    api_key: str | None = None
    client_id: str | None = None
    interval_hours: int | None = Field(default=None, ge=1, le=24)
    retention_days: int | None = Field(default=None, ge=1, le=365)
    is_active: bool | None = None


ROLE_ADMIN = "admin"
ROLE_USER = "user"
ROLE_FEEDBACK_MANAGER = "feedback_manager"
ROLE_CAN_ACCESS_ANALYTICS = {ROLE_ADMIN, ROLE_USER}
ROLE_CAN_ACCESS_SETTINGS = {ROLE_ADMIN, ROLE_USER}
ROLE_ASSIGNABLE_BY_ADMIN = {ROLE_USER, ROLE_FEEDBACK_MANAGER}
TENANT_ROLE_OWNER = "admin"
TENANT_ROLE_MANAGER = "feedback_manager"
SESSION_TTL_SECONDS = 30 * 24 * 60 * 60
CSRF_COOKIE_NAME = "csrf_token"
CSRF_HEADER_NAME = "X-CSRF-Token"
RATE_LIMIT_API_READ_PER_MINUTE = 600
RATE_LIMIT_API_WRITE_PER_MINUTE = 180
RATE_LIMIT_SYNC_PER_MINUTE = 20
RATE_LIMIT_LOGIN_PER_10_MIN = 30
FAILED_LOGIN_LIMIT_PER_15_MIN = 10
AUTO_SYNC_INTERVAL_SECONDS = 60


def _normalize_role(raw_role: object) -> str:
    role = str(raw_role or "").strip().lower()
    if role == ROLE_ADMIN:
        return ROLE_ADMIN
    if role == ROLE_USER:
        return ROLE_USER
    if role == ROLE_FEEDBACK_MANAGER:
        return ROLE_FEEDBACK_MANAGER
    return ROLE_USER


def create_app(config: AppConfig | None = None) -> FastAPI:
    # Uvicorn sets the root logger level to WARNING, which silently drops our
    # INFO messages even if the child logger level is INFO.
    # Fix: attach a StreamHandler directly to our package logger and disable
    # propagation to root.  This guarantees INFO output to stderr → journald
    # regardless of uvicorn's root logger configuration.
    _rp_logger = logging.getLogger("review_processor")
    if not _rp_logger.handlers:
        _h = logging.StreamHandler()
        _h.setLevel(logging.DEBUG)
        _h.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)s %(name)s: %(message)s"
        ))
        _rp_logger.addHandler(_h)
    _rp_logger.setLevel(logging.INFO)
    _rp_logger.propagate = False  # bypass root logger whose level is WARNING
    app_config = config or load_app_config()
    repository = ReviewRepository(db_url=app_config.db_url)
    service = ReviewAutomationService(repository)
    self_registration_enabled = bool(app_config.self_registration_enabled)

    app = FastAPI(title="Marketplace Reviews Assistant", version="1.0.0")
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
    sync_stop_event = threading.Event()
    # Supply sync state — separate from main feedback sync
    supply_sync_lock = threading.Lock()
    supply_sync_state: dict[str, object] = {
        "in_progress": False,
        "page": 0,
        "synced": 0,
        "total": 0,
        "errors": [],
        "message": "",
        "started_at": None,
        "finished_at": None,
    }
    # TTN daily sequential counter — stored in DB, resets each new day
    @app.post("/api/ttn/next-number")
    def ttn_next_number(request: Request) -> dict[str, object]:
        _require_user(request)
        return {"number": repository.next_ttn_number()}

    sync_lock = threading.Lock()
    sync_state: dict[str, object] = {
        "in_progress": False,
        "cancel_requested": False,
        "last_started_at": None,
        "last_finished_at": None,
        "polling_enabled": False,
        "polling_user_id": None,
        "polling_account_ids": [],
        "polling_since_date": None,
        "polling_started_at": None,
        "last_poll_at": None,
        "last_poll_result": None,
        # Progress tracking (visible to all users via /api/sync/status)
        "progress_step": "",
        "progress_account": "",
        "progress_channel": "",
        "progress_loaded": 0,
        "progress_total_items": 0,
        "progress_total_accounts": 0,
        "progress_current_account": 0,
        # Sync result report (shown after completion)
        "last_sync_report": None,  # populated after manual sync finishes
        "sync_log": [],  # list of log lines accumulated during sync
    }
    auto_sync_stop_event = threading.Event()
    auto_sync_worker: dict[str, threading.Thread | None] = {"thread": None}
    rate_limit_lock = threading.Lock()
    rate_buckets: dict[str, list[float]] = {}
    failed_login_attempts: dict[str, list[float]] = {}
    stock_scheduler = StockScheduler(repository)

    def _client_ip(request: Request) -> str:
        forwarded = str(request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
        if forwarded:
            return forwarded
        if request.client and request.client.host:
            return str(request.client.host)
        return "unknown"

    def _allow_rate(scope: str, *, limit: int, window_seconds: int) -> bool:
        now = time.time()
        cutoff = now - float(window_seconds)
        with rate_limit_lock:
            bucket = [ts for ts in rate_buckets.get(scope, []) if ts >= cutoff]
            if len(bucket) >= limit:
                rate_buckets[scope] = bucket
                return False
            bucket.append(now)
            rate_buckets[scope] = bucket
            return True

    def _record_failed_login(login_key: str) -> None:
        now = time.time()
        cutoff = now - 15 * 60
        with rate_limit_lock:
            bucket = [ts for ts in failed_login_attempts.get(login_key, []) if ts >= cutoff]
            bucket.append(now)
            failed_login_attempts[login_key] = bucket

    def _clear_failed_login(login_key: str) -> None:
        with rate_limit_lock:
            failed_login_attempts.pop(login_key, None)

    def _is_login_blocked(login_key: str) -> bool:
        now = time.time()
        cutoff = now - 15 * 60
        with rate_limit_lock:
            bucket = [ts for ts in failed_login_attempts.get(login_key, []) if ts >= cutoff]
            failed_login_attempts[login_key] = bucket
            return len(bucket) >= FAILED_LOGIN_LIMIT_PER_15_MIN

    def _is_private_host(hostname: str) -> bool:
        host = hostname.strip().lower().rstrip(".")
        if not host:
            return True
        if host in {"localhost", "localhost.localdomain"} or host.endswith(".localhost") or host.endswith(".local"):
            return True
        try:
            ip = ipaddress.ip_address(host)
        except ValueError:
            return False
        return bool(
            ip.is_private
            or ip.is_loopback
            or ip.is_link_local
            or ip.is_multicast
            or ip.is_reserved
            or ip.is_unspecified
        )

    def _validate_account_api_url(marketplace: str, raw_url: str) -> str:
        normalized = raw_url.strip()
        parsed = urlparse(normalized)
        if parsed.scheme.lower() != "https":
            raise HTTPException(status_code=400, detail="Адрес интерфейса API должен начинаться с https://")
        host = (parsed.hostname or "").strip().lower()
        if not host:
            raise HTTPException(status_code=400, detail="Некорректный адрес интерфейса API")
        if _is_private_host(host):
            raise HTTPException(status_code=400, detail="Адрес интерфейса API указывает на недопустимый внутренний хост")
        if marketplace == "wb" and not (host == "feedbacks-api.wildberries.ru" or host.endswith(".wildberries.ru")):
            raise HTTPException(status_code=400, detail="Для WB разрешены только домены wildberries.ru")
        if marketplace == "ozon" and not (host == "api-seller.ozon.ru" or host.endswith(".ozon.ru")):
            raise HTTPException(status_code=400, detail="Для OZON разрешены только домены ozon.ru")
        return normalized

    def _set_session_cookie(response: RedirectResponse, token: str) -> None:
        secure_cookie = bool(app_config.is_production)
        response.set_cookie(
            "session_token",
            token,
            httponly=True,
            samesite="lax",
            secure=secure_cookie,
            max_age=SESSION_TTL_SECONDS,
        )
        response.set_cookie(
            CSRF_COOKIE_NAME,
            secrets.token_urlsafe(32),
            httponly=False,
            samesite="lax",
            secure=secure_cookie,
            max_age=SESSION_TTL_SECONDS,
        )

    def _ensure_csrf_cookie(response: HTMLResponse | RedirectResponse, request: Request) -> None:
        if not request.cookies.get("session_token"):
            return
        if request.cookies.get(CSRF_COOKIE_NAME):
            return
        response.set_cookie(
            CSRF_COOKIE_NAME,
            secrets.token_urlsafe(32),
            httponly=False,
            samesite="lax",
            secure=bool(app_config.is_production),
            max_age=SESSION_TTL_SECONDS,
        )

    def _is_same_origin(request: Request, origin_value: str) -> bool:
        parsed = urlparse(origin_value)
        if not parsed.scheme or not parsed.netloc:
            return False
        expected_scheme = str(request.headers.get("x-forwarded-proto") or request.url.scheme).lower()
        expected_host = str(request.url.hostname or "").lower()
        expected_port = request.url.port or (443 if expected_scheme == "https" else 80)
        origin_host = str(parsed.hostname or "").lower()
        origin_port = parsed.port or (443 if parsed.scheme.lower() == "https" else 80)
        return parsed.scheme.lower() == expected_scheme and origin_host == expected_host and origin_port == expected_port

    def _check_csrf(request: Request) -> None:
        method = request.method.upper()
        if method not in {"POST", "PUT", "PATCH", "DELETE"}:
            return
        path = request.url.path
        if not path.startswith("/api/"):
            return
        # Only enforce CSRF for authenticated browser requests.
        if not request.cookies.get("session_token"):
            return
        cookie_token = str(request.cookies.get(CSRF_COOKIE_NAME) or "")
        header_token = str(request.headers.get(CSRF_HEADER_NAME) or "").strip()
        if not cookie_token or not header_token or not secrets.compare_digest(cookie_token, header_token):
            raise HTTPException(status_code=403, detail="CSRF токен отсутствует или неверен")
        origin = str(request.headers.get("origin") or "").strip()
        referer = str(request.headers.get("referer") or "").strip()
        if origin and not _is_same_origin(request, origin):
            raise HTTPException(status_code=403, detail="Недопустимый origin запроса")
        if not origin and referer and not _is_same_origin(request, referer):
            raise HTTPException(status_code=403, detail="Недопустимый referer запроса")

    def _check_rate_limit(request: Request) -> None:
        path = request.url.path
        method = request.method.upper()
        ip = _client_ip(request)
        if path == "/login" and method == "POST":
            login_scope = f"login:{ip}"
            if not _allow_rate(login_scope, limit=RATE_LIMIT_LOGIN_PER_10_MIN, window_seconds=10 * 60):
                raise HTTPException(status_code=429, detail="Слишком много попыток входа. Попробуйте позже.")
        if path.startswith("/api/"):
            if method in {"GET", "HEAD", "OPTIONS"}:
                limit = RATE_LIMIT_API_READ_PER_MINUTE
            elif path == "/api/sync":
                limit = RATE_LIMIT_SYNC_PER_MINUTE
            else:
                limit = RATE_LIMIT_API_WRITE_PER_MINUTE
            scope = f"api:{method}:{path}:{ip}"
            if not _allow_rate(scope, limit=limit, window_seconds=60):
                raise HTTPException(status_code=429, detail="Слишком много запросов. Попробуйте позже.")

    @app.middleware("http")
    async def hardening_middleware(request: Request, call_next):
        try:
            _check_rate_limit(request)
            _check_csrf(request)
        except HTTPException as exc:
            if request.url.path == "/login":
                return HTMLResponse(build_login_html(error=str(exc.detail)), status_code=exc.status_code)
            return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
        response = await call_next(request)
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("Referrer-Policy", "same-origin")
        response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; connect-src 'self'; frame-ancestors 'none'; form-action 'self'; base-uri 'self'",
        )
        if app_config.is_production:
            response.headers.setdefault("Strict-Transport-Security", "max-age=63072000; includeSubDomains; preload")
        _ensure_csrf_cookie(response, request)
        return response

    def _now_iso() -> str:
        return datetime.now(UTC).isoformat()

    def _issue_session(user_id: int) -> str:
        token = create_session_token()
        expires = (datetime.now(UTC) + timedelta(days=30)).isoformat()
        repository.create_session(token=token, user_id=user_id, expires_at=expires)
        return token

    def _login_attempt_key(request: Request, email: str) -> str:
        return f"{_client_ip(request)}::{email.strip().lower()}"

    def _get_current_user(request: Request) -> dict[str, object] | None:
        token = request.cookies.get("session_token")
        if not token:
            return None
        repository.cleanup_expired_sessions(_now_iso())
        return repository.get_session_user(token)

    def _require_user(request: Request) -> dict[str, object]:
        user = _get_current_user(request)
        if user is None:
            raise HTTPException(status_code=401, detail="Требуется авторизация")
        return user

    def _require_admin(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if user.get("role") != ROLE_ADMIN:
            raise HTTPException(status_code=403, detail="Доступ только для администратора")
        return user

    def _is_super_admin(user: dict[str, object]) -> bool:
        return bool(user.get("is_super_admin"))

    def _tenant_owner_id(user: dict[str, object]) -> int:
        owner_raw = user.get("owner_user_id")
        if owner_raw is None:
            return int(user["id"])
        try:
            return int(owner_raw)
        except (TypeError, ValueError):
            return int(user["id"])

    def _require_super_admin(request: Request) -> dict[str, object]:
        user = _require_admin(request)
        if not _is_super_admin(user):
            raise HTTPException(status_code=403, detail="Доступ только для супер-администратора")
        return user

    def _require_tenant_owner(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        if _is_super_admin(user):
            owner_scope_user_id = _tenant_owner_id(user)
            if owner_scope_user_id <= 0:
                owner_scope_user_id = int(user["id"])
            owner_scope_user = repository.get_user_by_id(owner_scope_user_id)
            return owner_scope_user or user
        if _tenant_owner_id(user) != int(user["id"]):
            raise HTTPException(status_code=403, detail="Недостаточно прав для управления командой")
        return user

    def _parse_sync_start_date_or_none(value: str | None, *, enabled: bool) -> str | None:
        if not enabled:
            return None
        raw = (value or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="Укажите дату начала синхронизации")
        try:
            datetime.strptime(raw, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Дата должна быть в формате ГГГГ-ММ-ДД") from exc
        return raw

    def _target_user_for_admin_scope(*, actor: dict[str, object], target_user_id: int) -> dict[str, object]:
        target = repository.get_user_by_id(target_user_id)
        if target is None:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        if _is_super_admin(actor):
            return target
        actor_owner_id = _tenant_owner_id(actor)
        target_owner_id = _tenant_owner_id(target)
        if target_owner_id != actor_owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не относится к вашему кабинету")
        if bool(target.get("is_super_admin")):
            raise HTTPException(status_code=403, detail="Недостаточно прав для управления этим пользователем")
        return target

    def _normalize_tenant_role_or_400(raw_role: str) -> str:
        role = str(raw_role or "").strip().lower()
        if role not in {TENANT_ROLE_OWNER, TENANT_ROLE_MANAGER}:
            raise HTTPException(status_code=400, detail="Роль должна быть: администратор или менеджер обратной связи")
        return role

    def _manager_permissions_context_for_user(user: dict[str, object]) -> list[dict[str, object]]:
        if str(user.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            return []
        return repository.list_manager_permissions(manager_user_id=int(user["id"]))

    def _manager_allowed_review_account_ids(user: dict[str, object]) -> list[int] | None:
        if str(user.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            return None
        rows = _manager_permissions_context_for_user(user)
        ids: list[int] = []
        seen: set[int] = set()
        for row in rows:
            if not bool(row.get("can_reviews")):
                continue
            try:
                account_id = int(row.get("account_id"))
            except (TypeError, ValueError):
                continue
            if account_id <= 0 or account_id in seen:
                continue
            seen.add(account_id)
            ids.append(account_id)
        return ids

    def _manager_allowed_conversation_accounts(user: dict[str, object]) -> dict[str, list[int]] | None:
        if str(user.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            return None
        rows = _manager_permissions_context_for_user(user)
        scope: dict[str, list[int]] = {"question": [], "chat": []}
        seen: dict[str, set[int]] = {"question": set(), "chat": set()}
        for row in rows:
            try:
                account_id = int(row.get("account_id"))
            except (TypeError, ValueError):
                continue
            if account_id <= 0:
                continue
            if bool(row.get("can_questions")) and account_id not in seen["question"]:
                seen["question"].add(account_id)
                scope["question"].append(account_id)
            if bool(row.get("can_chats")) and account_id not in seen["chat"]:
                seen["chat"].add(account_id)
                scope["chat"].append(account_id)
        return scope

    def _manager_owner_account_ids(owner_user_id: int) -> set[int]:
        return {
            int(item.get("id"))
            for item in repository.list_marketplace_accounts(user_id=owner_user_id, include_secrets=False)
            if item.get("id") is not None
        }

    def _require_manager_scope_for_review(user: dict[str, object], review_uid: str) -> None:
        if str(user.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            return
        allowed = set(_manager_allowed_review_account_ids(user) or [])
        if not allowed:
            raise HTTPException(status_code=403, detail="Менеджеру не назначены доступы к отзывам")
        review = repository.get_review(user_id=_tenant_owner_id(user), review_uid=review_uid)
        if review is None:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        try:
            account_id = int(review.get("account_id"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=403, detail="Отзыв не привязан к разрешенному кабинету")
        if account_id not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к этому кабинету отзывов")

    def _require_manager_scope_for_conversation(user: dict[str, object], conversation_uid: str) -> None:
        if str(user.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            return
        scope = _manager_allowed_conversation_accounts(user) or {"question": [], "chat": []}
        conversation = repository.get_conversation(user_id=_tenant_owner_id(user), conversation_uid=conversation_uid)
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        kind = str(conversation.get("kind") or "").strip().lower()
        if kind not in {"question", "chat"}:
            raise HTTPException(status_code=403, detail="Нет доступа к этому типу диалога")
        allowed = set(scope.get(kind, []))
        if not allowed:
            raise HTTPException(status_code=403, detail="Менеджеру не назначены доступы к этому типу диалогов")
        try:
            account_id = int(conversation.get("account_id"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=403, detail="Диалог не привязан к разрешенному кабинету")
        if account_id not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к этому кабинету диалогов")

    def _require_analytics_access(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role")) not in ROLE_CAN_ACCESS_ANALYTICS:
            raise HTTPException(status_code=403, detail="Недостаточно прав для просмотра аналитики")
        return user

    def _require_settings_access(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role")) not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Недостаточно прав для раздела настроек")
        return user

    def _set_sync_in_progress(in_progress: bool) -> None:
        with sync_lock:
            sync_state["in_progress"] = in_progress
            if in_progress:
                sync_state["cancel_requested"] = False
                sync_state["last_started_at"] = _now_iso()
                sync_stop_event.clear()
            else:
                sync_state["last_finished_at"] = _now_iso()

    def _snapshot_active_account_ids_for_user(user_id: int) -> list[int]:
        ids: list[int] = []
        seen: set[int] = set()
        for account in repository.list_marketplace_accounts(user_id, include_secrets=False):
            try:
                account_id = int(account.get("id"))
            except (TypeError, ValueError):
                continue
            if account_id <= 0 or account_id in seen:
                continue
            if not bool(account.get("is_active")):
                continue
            seen.add(account_id)
            ids.append(account_id)
        return ids

    def _serialize_sync_error_details(raw_errors: object) -> list[dict[str, object]]:
        if not isinstance(raw_errors, list):
            return []
        result: list[dict[str, object]] = []
        for item in raw_errors:
            if not isinstance(item, dict):
                continue
            cleaned: dict[str, object] = {}
            for key, value in item.items():
                cleaned[str(key)] = value
            result.append(cleaned)
        return result

    def _channel_access_label(capabilities: Mapping[str, object]) -> str:
        reviews_ok = bool(capabilities.get("reviews"))
        questions_ok = bool(capabilities.get("questions"))
        chats_ok = bool(capabilities.get("chats"))
        if reviews_ok and questions_ok and chats_ok:
            return "По данному ключу доступны все каналы: отзывы, вопросы и чаты."
        if chats_ok and not reviews_ok and not questions_ok:
            return "По данному ключу вы можете работать только с чатами. К отзывам и вопросам нет доступа."
        if reviews_ok and questions_ok and not chats_ok:
            return "По данному ключу вы можете работать только с отзывами и вопросами, но у вас нет доступа к чатам."
        if reviews_ok and chats_ok and not questions_ok:
            return "По данному ключу вы можете работать только с отзывами и чатами, но у вас нет доступа к вопросам."
        if questions_ok and chats_ok and not reviews_ok:
            return "По данному ключу вы можете работать только с вопросами и чатами, но у вас нет доступа к отзывам."
        if reviews_ok and not questions_ok and not chats_ok:
            return "По данному ключу вы можете работать только с отзывами. К вопросам и чатам нет доступа."
        if questions_ok and not reviews_ok and not chats_ok:
            return "По данному ключу вы можете работать только с вопросами. К отзывам и чатам нет доступа."
        return "По данному ключу нет доступа к отзывам, вопросам и чатам."

    def _probe_account_capabilities(*, user_id: int, account_id: int, since_date: str | None) -> dict[str, object]:
        account = repository.get_marketplace_account(
            user_id=user_id,
            account_id=account_id,
            include_secrets=True,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        if not bool(account.get("is_active")):
            raise HTTPException(status_code=400, detail="Кабинет отключен")

        probe = service.probe_account_channels(account=account, since_date=since_date or None)
        channels_raw = probe.get("channels")
        channels: dict[str, dict[str, object]] = (
            channels_raw if isinstance(channels_raw, dict) else {}
        )
        capabilities: dict[str, bool] = {
            "reviews": bool((channels.get("reviews") or {}).get("available")),
            "questions": bool((channels.get("questions") or {}).get("available")),
            "chats": bool((channels.get("chats") or {}).get("available")),
        }
        channel_messages: dict[str, str] = {}
        all_errors: list[dict[str, object]] = []
        for channel in ("reviews", "questions", "chats"):
            channel_data = channels.get(channel)
            if not isinstance(channel_data, Mapping):
                continue
            if bool(channel_data.get("available")):
                continue
            message = str(channel_data.get("error") or "").strip()
            if message:
                channel_messages[channel] = message
            all_errors.append(
                {
                    "account_id": int(probe.get("account_id") or account_id),
                    "marketplace": str(probe.get("marketplace") or account.get("marketplace") or ""),
                    "channel": channel,
                    "scope": channel,
                    "error": message,
                    "access_denied": bool(channel_data.get("access_denied")),
                }
            )
        can_sync_any = any(capabilities.values())
        return {
            "account_id": int(probe.get("account_id") or account_id),
            "marketplace": str(probe.get("marketplace") or account.get("marketplace") or ""),
            "account_name": str(probe.get("account_name") or account.get("account_name") or ""),
            "is_active": bool(account.get("is_active")),
            "capabilities": capabilities,
            "can_sync_any": can_sync_any,
            "summary": _channel_access_label(capabilities),
            "channel_messages": channel_messages,
            "errors": all_errors,
            "all_channels_available": bool(probe.get("all_channels_available")),
        }

    def _update_sync_progress(
        *,
        step: str = "",
        account: str = "",
        channel: str = "",
        loaded: int = 0,
        total_accounts: int = 0,
        current_account: int = 0,
    ) -> None:
        with sync_lock:
            sync_state["progress_step"] = step
            sync_state["progress_account"] = account
            sync_state["progress_channel"] = channel
            if loaded:
                sync_state["progress_loaded"] = int(sync_state.get("progress_loaded") or 0) + loaded
            if total_accounts:
                sync_state["progress_total_accounts"] = total_accounts
            if current_account:
                sync_state["progress_current_account"] = current_account
            # Accumulate log line (keep last 200 lines)
            if step or account or channel:
                ts = _now_iso()[11:19]  # HH:MM:SS
                parts = [p for p in [account, channel, step] if p]
                line = f"{ts}  {' → '.join(parts)}"
                log_lines = list(sync_state.get("sync_log") or [])
                log_lines.append(line)
                sync_state["sync_log"] = log_lines[-200:]

    def _run_sync_for_user(
        *,
        user_id: int,
        since_date: str | None,
        account_ids: list[int] | None,
        run_started_at: str,
        apply_date_filter: bool = False,
    ) -> dict[str, object]:
        with sync_lock:
            if bool(sync_state.get("in_progress")):
                # If an auto-sync is running and this is a manual request,
                # cancel the auto-sync and let the manual one proceed after a short wait.
                if apply_date_filter and not bool(sync_state.get("is_manual")):
                    sync_stop_event.set()  # signal auto-sync to stop
                    # Will retry after releasing lock; auto-sync checks stop_requested
                else:
                    raise HTTPException(status_code=409, detail="Синхронизация уже выполняется")
        # Brief wait for auto-sync to see the stop signal before we acquire the slot
        if apply_date_filter and sync_stop_event.is_set():
            import time as _time
            _time.sleep(2)
        with sync_lock:
            if bool(sync_state.get("in_progress")):
                raise HTTPException(status_code=409, detail="Синхронизация уже выполняется. Попробуйте снова через несколько секунд.")
            sync_state["in_progress"] = True
            sync_state["is_manual"] = apply_date_filter  # True only for manual button clicks
            sync_state["cancel_requested"] = False
            sync_state["last_started_at"] = run_started_at
            sync_state["progress_step"] = "Подготовка..."
            sync_state["progress_account"] = ""
            sync_state["progress_channel"] = ""
            sync_state["progress_loaded"] = 0
            sync_state["progress_total_items"] = 0
            sync_state["progress_total_accounts"] = 0
            sync_state["progress_current_account"] = 0
            sync_state["sync_log"] = []  # reset log for new sync
        sync_stop_event.clear()
        try:
            result = service.sync_all_accounts(
                user_id=user_id,
                since_date=since_date or None,
                account_ids=account_ids,
                stop_requested=sync_stop_event.is_set,
                progress_callback=_update_sync_progress,
                apply_date_filter=apply_date_filter,
            )
            # Build detailed sync report for the completion modal
            if apply_date_filter and isinstance(result, dict):
                report_accounts = []
                for stat in (result.get("account_channel_stats") or []):
                    acct_id = stat.get("account_id")
                    acct_name = stat.get("account_name") or f"#{acct_id}"
                    channels = {}
                    for ch in ("reviews", "questions", "chats"):
                        ch_data = stat.get(ch) or {}
                        channels[ch] = {
                            "ok": bool(ch_data.get("ok")),
                            "loaded": int(ch_data.get("loaded") or 0),
                            "skipped": int(ch_data.get("skipped_old") or 0),
                            "error": str(ch_data.get("error") or ""),
                        }
                    report_accounts.append({
                        "account_id": acct_id,
                        "account_name": acct_name,
                        "channels": channels,
                    })
                with sync_lock:
                    sync_state["last_sync_report"] = {
                        "started_at": run_started_at,
                        "finished_at": _now_iso(),
                        "accounts": report_accounts,
                        "total_reviews": int(result.get("loaded_reviews") or result.get("loaded") or 0),
                        "total_questions": int(result.get("loaded_questions") or 0),
                        "total_chats": int(result.get("loaded_chats") or 0),
                        "cancelled": bool(result.get("cancelled")),
                        "errors": int(result.get("failed_accounts") or 0),
                        "log": list(sync_state.get("sync_log") or []),
                    }
            return result
        finally:
            with sync_lock:
                sync_state["in_progress"] = False
                sync_state["last_finished_at"] = _now_iso()
                sync_state["progress_step"] = "Завершено"
            sync_stop_event.clear()

    def _start_auto_sync_worker_if_needed() -> None:
        with sync_lock:
            existing = auto_sync_worker.get("thread")
            if isinstance(existing, threading.Thread) and existing.is_alive():
                return
            auto_sync_stop_event.clear()

            def _auto_sync_loop() -> None:
                _log.info("auto_sync_loop: started, first poll in %ds", AUTO_SYNC_INTERVAL_SECONDS)
                while not auto_sync_stop_event.is_set():
                    auto_sync_stop_event.wait(AUTO_SYNC_INTERVAL_SECONDS)
                    if auto_sync_stop_event.is_set():
                        break
                    _log.info("auto_sync_loop: poll iteration starting")
                    # Read sync target from DB on every iteration so the loop
                    # works even if in-memory sync_state was cleared (e.g. Stop
                    # button pressed, then the next auto-sync still fires).
                    # This makes polling resilient to manual Stop and restarts.
                    try:
                        owner_users_for_poll = repository.list_users(owner_only=True)
                    except Exception as exc:
                        _log.warning("auto_sync_loop: list_users failed: %s", exc)
                        continue
                    for poll_user in owner_users_for_poll:
                        try:
                            polling_user_id = int(poll_user.get("id") or 0)
                            if polling_user_id <= 0:
                                continue
                            poll_accounts = [
                                item for item in
                                repository.list_marketplace_accounts(polling_user_id, include_secrets=False)
                                if item.get("is_active")
                            ]
                            account_ids = [int(a["id"]) for a in poll_accounts if a.get("id")]
                            if not account_ids:
                                continue
                            poll_sync_settings = repository.get_user_sync_settings(user_id=polling_user_id)
                            polling_since_raw = (
                                str(poll_sync_settings.get("sync_start_date") or "").strip()
                                if bool(poll_sync_settings.get("use_sync_start_date"))
                                else None
                            )
                            # Update in-memory state so UI can see polling is active
                            with sync_lock:
                                sync_state["polling_enabled"] = True
                                sync_state["polling_user_id"] = polling_user_id
                                sync_state["polling_account_ids"] = account_ids
                                sync_state["polling_since_date"] = polling_since_raw
                        except Exception:
                            continue
                        run_started_at = _now_iso()
                        try:
                            result = _run_sync_for_user(
                                user_id=polling_user_id,
                                since_date=polling_since_raw or None,
                                account_ids=account_ids,
                                run_started_at=run_started_at,
                            )
                            with sync_lock:
                                sync_state["last_poll_at"] = _now_iso()
                                sync_state["last_poll_result"] = {
                                    "ok": True,
                                    "run_started_at": run_started_at,
                                    "accounts": int(result.get("accounts") or 0),
                                    "success_accounts": int(result.get("success_accounts") or 0),
                                    "failed_accounts": int(result.get("failed_accounts") or 0),
                                    "loaded": int(result.get("loaded") or 0),
                                    "loaded_conversations": int(result.get("loaded_conversations") or 0),
                                    "account_ids": list(account_ids),
                                    "errors": _serialize_sync_error_details(result.get("errors")),
                                    "cancelled": bool(result.get("cancelled")),
                                }
                        except HTTPException as exc:
                            with sync_lock:
                                sync_state["last_poll_at"] = _now_iso()
                                sync_state["last_poll_result"] = {
                                    "ok": False,
                                    "run_started_at": run_started_at,
                                    "error": str(exc.detail),
                                    "account_ids": list(account_ids),
                                }
                        except Exception as exc:
                            with sync_lock:
                                sync_state["last_poll_at"] = _now_iso()
                                sync_state["last_poll_result"] = {
                                    "ok": False,
                                    "run_started_at": run_started_at,
                                    "error": str(exc),
                                    "account_ids": list(account_ids),
                                }
                        # Continue to next owner user (no break — all tenants polled)

            worker = threading.Thread(
                target=_auto_sync_loop,
                name="feedpilot-auto-sync-worker",
                daemon=True,
            )
            auto_sync_worker["thread"] = worker
            worker.start()

    def _template_group_by_id(group_id: str) -> dict[str, object] | None:
        for item in TEMPLATE_GROUPS:
            if str(item.get("id")) == group_id:
                return item
        return None

    def _is_protected_subgroup(group_id: str, subgroup: str) -> bool:
        """Return True if this subgroup must not be deleted by anyone.

        The 'textless_ratings' group has fixed per-star subgroups that are
        required for the review processing pipeline and cannot be removed.
        """
        from .service import ReviewAutomationService as _RAS
        if str(group_id or "").strip() != _RAS.TEXTLESS_GROUP_ID:
            return False
        return str(subgroup or "").strip() in _RAS.TEXTLESS_SUBGROUPS

    def _base_subgroups_for_group(group_id: str) -> list[str]:
        group = _template_group_by_id(group_id)
        if group is None:
            return []
        subgroups_raw = group.get("subgroups")
        if not isinstance(subgroups_raw, list):
            return []
        result: list[str] = []
        for value in subgroups_raw:
            name = str(value).strip()
            if name and name not in result:
                result.append(name)
        return result

    def _all_subgroups_for_group(group_id: str) -> list[dict[str, object]]:
        custom_rows = repository.list_default_template_subgroups(group_id=group_id)
        result: list[dict[str, object]] = []
        seen: set[str] = set()
        for row in custom_rows:
            clean = str((row or {}).get("subgroup") or "").strip()
            if not clean or clean in seen:
                continue
            seen.add(clean)
            subgroup_id = str((row or {}).get("subgroup_id") or "").strip() or None
            result.append({"name": clean, "subgroup_id": subgroup_id})
        if group_id in GENERAL_LOCKED_GROUP_IDS:
            general = GENERAL_LOCKED_SUBGROUP
            reordered = [item for item in result if str(item.get("name") or "") != general]
            general_item = next((item for item in result if str(item.get("name") or "") == general), None)
            if general_item is None:
                general_item = {"name": general, "subgroup_id": None}
            result = [general_item, *reordered]
        if result:
            return result
        # Backward-compatible fallback for old datasets where subgroup registry
        # might still be empty.
        return [{"name": name, "subgroup_id": None} for name in _base_subgroups_for_group(group_id)]
        

    def _validate_subgroup(group_id: str, subgroup: str) -> bool:
        clean_group_id = str(group_id or "").strip()
        clean_subgroup = str(subgroup or "").strip()
        if not clean_group_id or not clean_subgroup:
            return False
        return any(
            str(item.get("name") or "") == clean_subgroup
            for item in _all_subgroups_for_group(clean_group_id)
        )

    def _default_template_seed_rows() -> tuple[list[dict[str, str]], list[dict[str, object]]]:
        rows: list[dict[str, str]] = []
        subgroup_rows: list[dict[str, object]] = []
        for group in TEMPLATE_GROUPS:
            group_id = str(group.get("id") or "")
            subgroups = group.get("subgroups")
            if not group_id or not isinstance(subgroups, list):
                continue
            for subgroup in subgroups:
                name = str(subgroup).strip()
                if not name:
                    continue
                subgroup_rows.append({"group_id": group_id, "subgroup": name, "is_system": True})
                defaults = DEFAULT_TEMPLATE_CONTENT.get(name) or [f"Спасибо за отзыв! Категория: {name}."]
                for text in defaults:
                    clean = str(text or "").strip()
                    if not clean:
                        continue
                    rows.append(
                        {
                            "group_id": group_id,
                            "subgroup": name,
                            "template_text": clean,
                        }
                    )
        return rows, subgroup_rows

    def _ensure_platform_default_templates() -> None:
        seed_rows, subgroup_rows = _default_template_seed_rows()
        if repository.count_default_template_subgroups() == 0:
            repository.ensure_default_template_subgroups(subgroup_rows)
        for group_id in GENERAL_LOCKED_GROUP_IDS:
            repository.ensure_default_template_subgroups(
                [{"group_id": group_id, "subgroup": GENERAL_LOCKED_SUBGROUP}]
            )
            existing = repository.list_default_template_variants(group_id=group_id, subgroup=GENERAL_LOCKED_SUBGROUP)
            if not existing:
                repository.replace_default_subgroup_templates(
                    group_id=group_id,
                    subgroup=GENERAL_LOCKED_SUBGROUP,
                    templates=DEFAULT_TEMPLATE_CONTENT.get(GENERAL_LOCKED_SUBGROUP)
                    or ["Спасибо за ваш отзыв! Мы ценим обратную связь и уже работаем над улучшениями."],
                )
        repository.sync_default_template_subgroups_from_variants()
        if repository.count_default_template_variants(include_inactive=True) > 0:
            return
        seeded = repository.seed_default_templates_from_user_templates()
        if seeded > 0:
            return
        repository.seed_default_template_variants(seed_rows)

    def _build_template_group_items(counts: dict[tuple[str, str], int]) -> list[dict[str, object]]:
        items: list[dict[str, object]] = []
        for group in TEMPLATE_GROUPS:
            group_id = str(group.get("id") or "")
            title = str(group.get("title") or group_id)
            base_subgroups = set(_base_subgroups_for_group(group_id))
            all_subgroups = _all_subgroups_for_group(group_id)
            subgroups: list[dict[str, object]] = []
            for subgroup_item in all_subgroups:
                subgroup_name = str(subgroup_item.get("name") or "").strip()
                if not subgroup_name:
                    continue
                subgroups.append(
                    {
                        "name": subgroup_name,
                        "count": counts.get((group_id, subgroup_name), 0),
                        "subgroup_id": str(subgroup_item.get("subgroup_id") or "").strip() or None,
                        "is_system": subgroup_name in base_subgroups,
                    }
                )
            items.append(
                {
                    "id": group_id,
                    "title": title,
                    "subgroups": subgroups,
                }
            )
        return items

    def _ensure_default_template_variants(user_id: int) -> None:
        _ensure_platform_default_templates()
        repository.copy_default_templates_to_user(user_id=user_id, only_if_empty=True)

    def _parse_recommendation_targets(raw_csv: str) -> list[str]:
        values = str(raw_csv or "").replace(";", ",").replace("\n", ",").split(",")
        result: list[str] = []
        seen: set[str] = set()
        for value in values:
            article = value.strip()
            if not article or article in seen:
                continue
            seen.add(article)
            result.append(article)
        return result

    def _registration_disabled_response() -> HTMLResponse:
        return HTMLResponse(
            build_login_html(error="Самостоятельная регистрация отключена. Пользователей добавляет администратор."),
            status_code=403,
        )

    @app.get("/", response_class=HTMLResponse)
    def landing(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is not None:
            return RedirectResponse("/app", status_code=302)
        return HTMLResponse(build_landing_html())

    @app.get("/login", response_class=HTMLResponse)
    def login_page(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is not None:
            return RedirectResponse("/app", status_code=302)
        return HTMLResponse(build_login_html())

    @app.post("/login")
    def login(request: Request, email: str = Form(...), password: str = Form(...)) -> HTMLResponse:
        login_key = _login_attempt_key(request, email)
        if _is_login_blocked(login_key):
            return HTMLResponse(build_login_html(error="Слишком много неудачных попыток входа. Повторите позже."), status_code=429)
        user = repository.get_user_by_email(email)
        if user is None or not verify_password(password, str(user["password_hash"])):
            _record_failed_login(login_key)
            return HTMLResponse(build_login_html(error="Неверная эл. почта или пароль"), status_code=401)

        _clear_failed_login(login_key)
        token = _issue_session(int(user["id"]))
        response = RedirectResponse("/app", status_code=302)
        _set_session_cookie(response, token)
        return response

    @app.get("/register", response_class=HTMLResponse)
    def register_page(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is not None:
            return RedirectResponse("/app", status_code=302)
        if not self_registration_enabled:
            return _registration_disabled_response()
        return HTMLResponse(build_register_html())

    @app.post("/register")
    def register(email: str = Form(...), password: str = Form(...), password_repeat: str = Form(...)) -> HTMLResponse:
        if not self_registration_enabled:
            return _registration_disabled_response()
        email = email.strip().lower()
        if len(email) < 5 or "@" not in email:
            return HTMLResponse(build_register_html(error="Введите корректную эл. почту"), status_code=400)
        if len(password) < 8:
            return HTMLResponse(build_register_html(error="Пароль должен быть не короче 8 символов"), status_code=400)
        if password != password_repeat:
            return HTMLResponse(build_register_html(error="Пароли не совпадают"), status_code=400)
        if repository.get_user_by_email(email) is not None:
            return HTMLResponse(build_register_html(error="Пользователь уже существует"), status_code=409)

        role = ROLE_ADMIN if repository.count_users() == 0 else ROLE_USER
        user = repository.create_user(email=email, password_hash=hash_password(password), role=role)
        token = _issue_session(int(user["id"]))
        response = RedirectResponse("/app", status_code=302)
        _set_session_cookie(response, token)
        return response

    @app.get("/logout")
    def logout(request: Request) -> RedirectResponse:
        token = request.cookies.get("session_token")
        if token:
            repository.delete_session(token)
        response = RedirectResponse("/", status_code=302)
        secure_cookie = bool(app_config.is_production)
        response.delete_cookie("session_token", samesite="lax", secure=secure_cookie)
        response.delete_cookie(CSRF_COOKIE_NAME, samesite="lax", secure=secure_cookie)
        return response

    @app.get("/app", response_class=HTMLResponse)
    def app_dashboard(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is None:
            return RedirectResponse("/login", status_code=302)
        return HTMLResponse(build_app_html(user, repository=repository))

    @app.get("/admin", response_class=HTMLResponse)
    def admin_page(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is None:
            return RedirectResponse("/login", status_code=302)
        if user.get("role") != ROLE_ADMIN:
            return HTMLResponse("<h1>Доступ запрещен</h1><p>Нужны права администратора.</p>", status_code=403)
        return HTMLResponse(build_admin_html(user))

    @app.get("/api/me")
    def get_me(request: Request) -> dict[str, object]:
        user = _require_user(request)
        return {
            "id": user["id"],
            "email": user["email"],
            "full_name": user.get("full_name") or "",
            "role": user["role"],
        }

    @app.get("/api/profile")
    def get_profile(request: Request) -> dict[str, object]:
        user = _require_user(request)
        sync_settings = repository.get_user_sync_settings(user_id=int(user["id"]))
        template_variables = repository.list_user_template_variable_values(user_id=int(user["id"]))
        editable_template_variables = [
            item
            for item in template_variables
            if bool(item.get("is_user_editable")) and bool(item.get("is_active"))
        ]
        return {
            "full_name": user.get("full_name") or "",
            "email": user["email"],
            "use_sync_start_date": bool(sync_settings.get("use_sync_start_date")),
            "sync_start_date": str(sync_settings.get("sync_start_date") or "") or None,
            "default_sync_lookback_days": int(sync_settings.get("default_sync_lookback_days") or 7),
            "editable_template_variables": editable_template_variables,
        }

    @app.get("/api/user-sync-settings")
    def get_user_sync_settings(request: Request) -> dict[str, object]:
        user = _require_user(request)
        return repository.get_user_sync_settings(user_id=int(user["id"]))

    @app.put("/api/user-sync-settings")
    def update_user_sync_settings(request: Request, payload: UserSyncSettingsRequest) -> dict[str, object]:
        user = _require_user(request)
        # User-level settings always operate with explicit sync start date.
        # The checkbox toggle was removed from UI, so force enabled mode.
        enabled_sync_start_date = True
        sync_start_date = _parse_sync_start_date_or_none(
            payload.sync_start_date,
            enabled=enabled_sync_start_date,
        )
        updated = repository.save_user_sync_settings(
            user_id=int(user["id"]),
            use_sync_start_date=enabled_sync_start_date,
            sync_start_date=sync_start_date,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        settings = repository.get_user_sync_settings(user_id=int(user["id"]))
        return {"ok": True, "settings": settings}

    @app.get("/api/user/template-variables")
    def user_list_template_variables(request: Request) -> dict[str, object]:
        user = _require_user(request)
        rows = repository.list_user_template_variable_values(user_id=int(user["id"]))
        items = [
            item
            for item in rows
            if bool(item.get("is_active")) and bool(item.get("is_user_editable"))
        ]
        return {"items": items, "count": len(items)}

    @app.put("/api/user/template-variables")
    def user_save_template_variables(
        payload: UserTemplateVariableValuesSaveRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_user(request)
        saved = repository.save_user_template_variable_values(
            user_id=int(user["id"]),
            values={str(k): str(v) for k, v in dict(payload.values).items()},
        )
        return {"ok": True, "saved": int(saved)}

    @app.put("/api/profile")
    def update_profile(request: Request, payload: ProfileUpdateRequest) -> dict[str, object]:
        user = _require_user(request)
        user_id = int(user["id"])
        stored_user = repository.get_user_by_id(user_id)
        if stored_user is None:
            raise HTTPException(status_code=404, detail="Пользователь не найден")

        new_email = (payload.email or str(stored_user.get("email") or "")).strip().lower()
        if not new_email or "@" not in new_email:
            raise HTTPException(status_code=400, detail="Введите корректную электронную почту")

        full_name = (payload.full_name or "").strip() or None

        wants_password_change = any(
            value is not None and value != ""
            for value in (payload.current_password, payload.new_password, payload.new_password_repeat)
        )
        password_hash: str | None = None
        if wants_password_change:
            if not payload.current_password:
                raise HTTPException(status_code=400, detail="Введите текущий пароль")
            if not verify_password(payload.current_password, str(stored_user.get("password_hash") or "")):
                raise HTTPException(status_code=400, detail="Текущий пароль неверный")
            if not payload.new_password or len(payload.new_password) < 8:
                raise HTTPException(status_code=400, detail="Новый пароль должен быть не короче 8 символов")
            if payload.new_password != (payload.new_password_repeat or ""):
                raise HTTPException(status_code=400, detail="Новый пароль и подтверждение не совпадают")
            password_hash = hash_password(payload.new_password)

        try:
            updated = repository.update_user_profile(
                user_id=user_id,
                email=new_email,
                full_name=full_name,
                password_hash=password_hash,
            )
        except Exception as exc:
            is_duplicate_error = False
            if psycopg is not None and isinstance(exc, getattr(psycopg, "IntegrityError", (Exception,))):
                is_duplicate_error = True
            if "integrityerror" in str(type(exc)).lower():
                is_duplicate_error = True
            if not is_duplicate_error:
                raise
            raise HTTPException(status_code=409, detail="Эта электронная почта уже используется другим аккаунтом") from exc
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/reviews")
    def list_reviews(
        request: Request,
        source: str | None = None,
        priority: str | None = None,
        status: str | None = None,
        category: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        sort: str = "newest",
        page: int = 1,
        page_size: int = 30,
        bucket: str = "all",
        product_search: str = "",
        has_contradiction: int = 0,
    ) -> dict[str, object]:
        user = _require_user(request)
        allowed_page_sizes = {10, 30, 50, 100}
        normalized_page_size = page_size if page_size in allowed_page_sizes else 30
        normalized_bucket = bucket.strip().lower()
        if normalized_bucket not in {"all", "new", "processed"}:
            normalized_bucket = "all"
        normalized_sort = sort.strip().lower()
        if normalized_sort not in {"newest", "oldest", "rating_asc", "rating_desc", "category"}:
            normalized_sort = "newest"

        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")

        normalized_source = source.strip().lower() if source else None
        if normalized_source in {"all", "all_sources"}:
            normalized_source = None

        status_key = status.strip().lower() if status else ""
        status_map: dict[str, list[str] | None] = {
            "": None,
            "all": None,
            "waiting_send": ["waiting_send"],
            "processed_outside_spix": ["processed_outside_spix"],
            "rejected": ["rejected", "ignored"],
            "answered": ["answered_auto", "answered_manual", "answered"],
            "waiting_processing": ["queued_for_operator", "waiting_processing"],
            "generating_answer": ["generating_answer"],
        }
        status_values = status_map.get(status_key)
        if status_values is None and status_key not in {"", "all"}:
            status_values = [status_key]

        account_ids_filter = _manager_allowed_review_account_ids(user)
        # Use owner's user_id for data queries — managers share owner's data
        owner_user_id = _tenant_owner_id(user)
        page_data = service.list_reviews_paginated(
            user_id=owner_user_id,
            source=normalized_source,
            priority=priority,
            status=None,
            statuses=status_values,
            category=category,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
            sort=normalized_sort,
            page=max(page, 1),
            page_size=normalized_page_size,
            bucket=normalized_bucket,
            account_ids=account_ids_filter,
            product_search=product_search.strip() or None,
            has_contradiction=bool(has_contradiction),
        )
        # Enrich each review with a suggested template text for the reply column.
        # Uses a single batch query instead of N separate queries (eliminates N+1).
        user_id_int = owner_user_id
        items = page_data.get("items") or []
        # First pass: collect pairs and expose classified_subgroup
        pairs: list[tuple[str, str | None]] = []
        for item in items:
            meta = item.get("metadata") or {}
            subgroup = str(meta.get("classified_subgroup") or "")
            item["classified_subgroup"] = subgroup
            group_id = str(item.get("category") or "")
            if group_id:
                pairs.append((group_id, subgroup or None))
        # Single DB call — load ALL templates for relevant groups
        # Each review picks independently so different reviews get different templates
        import random as _rnd
        group_ids_needed = list({p[0] for p in pairs if p[0]})
        try:
            tmpl_pool = repository.get_template_pool_for_reviews(
                user_id=user_id_int, group_ids=group_ids_needed
            )
        except Exception:
            tmpl_pool = {}
        # Load contradiction rules once for dynamic check on existing reviews
        try:
            contradiction_map = repository.get_review_contradiction_map(user_id=user_id_int)
        except Exception:
            contradiction_map = {}
        # Second pass: each review gets its own random pick from the pool
        for item in items:
            group_id = str(item.get("category") or "")
            subgroup = str(item.get("classified_subgroup") or "")
            rating_val = item.get("rating")
            # Skip template if contradiction rule matches (flag in metadata OR dynamic check)
            meta = item.get("metadata") or {}
            has_contradiction_flag = bool(meta.get("rating_contradiction"))
            has_contradiction_rule = bool(
                group_id and rating_val is not None
                and int(rating_val) in contradiction_map.get(group_id, set())
            )
            if has_contradiction_flag or has_contradiction_rule:
                item["suggested_reply"] = ""
            elif group_id:
                texts = tmpl_pool.get((group_id, subgroup)) or tmpl_pool.get((group_id, "")) or []
                raw_tpl = _rnd.choice(texts) if texts else ""
                if raw_tpl:
                    # Apply basic variable substitution for UI display
                    _author = str(item.get("author") or "").strip()
                    # If author is empty — remove preceding comma/space to avoid ", !" artifacts
                    if not _author:
                        for _ph in ("%USER%", "%AUTHOR%"):
                            raw_tpl = raw_tpl.replace(f", {_ph}", "").replace(f" {_ph}", "")
                    raw_tpl = raw_tpl.replace("%USER%", _author)
                    raw_tpl = raw_tpl.replace("%AUTHOR%", _author)
                    # Apply user-defined template variables context
                    try:
                        _vars_ctx = repository.build_template_variables_context(
                            user_id=user_id_int,
                            review_author=_author,
                            review_rating=item.get("rating"),
                            review_category=group_id,
                            review_sentiment="",
                            review_tags=None,
                            review_metadata=item.get("metadata") if isinstance(item.get("metadata"), dict) else {},
                        )
                        for _vk, _vv in _vars_ctx.items():
                            raw_tpl = raw_tpl.replace(str(_vk), str(_vv or ""))
                    except Exception:
                        pass
                    # Remove any remaining unreplaced %VAR% placeholders
                    import re as _re2
                    raw_tpl = _re2.sub(r'%[A-Z0-9_]{2,50}%', '', raw_tpl)
                item["suggested_reply"] = raw_tpl
            else:
                item["suggested_reply"] = ""

        source_options = service.list_review_sources(user_id=owner_user_id)
        # Enrich with product photo URLs
        try:
            _photo_map = repository.get_product_photo_map(user_id=owner_user_id)
            if _photo_map:
                for _item in page_data["items"]:
                    _meta = _item.get("metadata") or {}
                    _raw = _meta.get("raw") or {} if isinstance(_meta, dict) else {}
                    _pd = (_raw.get("productDetails") or {}) if isinstance(_raw, dict) else {}
                    _keys = [
                        str(_pd.get("supplierArticle") or "").strip(),
                        str(_pd.get("nmId") or "").strip(),
                        str(_raw.get("supplierArticle") or "").strip(),
                    ]
                    _item["product_photo_url"] = next(
                        (_photo_map[k] for k in _keys if k and k in _photo_map), None
                    )
        except Exception:
            pass
        return {
            "items": page_data["items"],
            "count": len(page_data["items"]),
            "total": page_data["total"],
            "page": page_data["page"],
            "page_size": page_data["page_size"],
            "pages": page_data["pages"],
            "new_count": page_data["new_count"],
            "processed_count": page_data["processed_count"],
            "bucket": normalized_bucket,
            "sort": normalized_sort,
            "date_from": normalized_date_from,
            "date_to": normalized_date_to,
            "source": normalized_source,
            "status": status_key or "all",
            "source_options": source_options,
        }

    @app.get("/api/reviews/random-template")
    def reviews_random_template(
        request: Request,
        group_id: str,
        subgroup: str = "",
        review_uid: str = "",
    ) -> dict[str, object]:
        """Return a random template for the given group/subgroup — used by the reply refresh button."""
        user = _require_user(request)
        owner_uid = _tenant_owner_id(user)
        tmpl = repository.get_random_template_variant(
            user_id=owner_uid,
            group_id=group_id.strip(),
            subgroup=subgroup.strip() or None,
        )
        raw_text = str(tmpl.get("template_text") or "") if tmpl else ""
        if raw_text and review_uid.strip():
            # Apply variable substitution using the specific review's data
            review_obj = repository.get_review(user_id=owner_uid, review_uid=review_uid.strip())
            if review_obj:
                _author = str(review_obj.get("author") or "").strip()
                if not _author:
                    for _ph in ("%USER%", "%AUTHOR%"):
                        raw_text = raw_text.replace(f", {_ph}", "").replace(f" {_ph}", "")
                raw_text = raw_text.replace("%USER%", _author)
                raw_text = raw_text.replace("%AUTHOR%", _author)
                try:
                    _vars_ctx = repository.build_template_variables_context(
                        user_id=owner_uid,
                        review_author=_author,
                        review_rating=review_obj.get("rating"),
                        review_category=group_id.strip(),
                        review_sentiment="",
                        review_tags=None,
                        review_metadata=review_obj.get("metadata") if isinstance(review_obj.get("metadata"), dict) else {},
                    )
                    for _vk, _vv in _vars_ctx.items():
                        raw_text = raw_text.replace(str(_vk), str(_vv or ""))
                except Exception:
                    pass
        # Remove any remaining unreplaced %VAR% placeholders
        import re as _re_tpl
        raw_text = _re_tpl.sub(r'%[A-Z0-9_]{2,50}%', '', raw_text)
        return {"template_text": raw_text}

    @app.post("/api/reviews/{review_uid}/reply")
    def reply_to_review(review_uid: str, request: Request, payload: ConversationReplyRequest) -> dict[str, object]:
        """Send a reply to a WB review directly from the review table."""
        user = _require_user(request)
        owner_uid = _tenant_owner_id(user)
        review_obj = repository.get_review(user_id=owner_uid, review_uid=review_uid)
        if review_obj is None:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        response_text = str(payload.response_text or "").strip()
        if not response_text:
            raise HTTPException(status_code=400, detail="Текст ответа не может быть пустым")
        account_id = review_obj.get("account_id")
        account = repository.get_marketplace_account(
            user_id=owner_uid,
            account_id=int(account_id),
            include_secrets=True,
        ) if account_id else None
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет не найден")
        client = service._build_client(account)
        review_input = ReviewInput(
            review_id=str(review_obj.get("external_review_id") or ""),
            text=str(review_obj.get("text") or ""),
            metadata=review_obj.get("metadata") or {},
        )
        try:
            sent = client.send_review_reply(review=review_input, response_text=response_text)
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось отправить ответ: {exc}")
        if not sent:
            raise HTTPException(status_code=502, detail="Ответ не был отправлен")
        repository.update_review_manual_reply(
            user_id=owner_uid,
            review_uid=review_uid,
            operator_name=str(user.get("full_name") or user.get("email") or "Оператор"),
            reply_text=response_text,
        )
        return {"ok": True}

    @app.post("/api/reviews/{review_uid}/retry-send")
    def retry_review_send(review_uid: str, request: Request) -> dict[str, object]:
        """Retry a previously failed auto-reply using the saved auto_reply text."""
        user = _require_user(request)
        owner_uid = _tenant_owner_id(user)
        review_obj = repository.get_review(user_id=owner_uid, review_uid=review_uid)
        if review_obj is None:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        auto_reply_text = str(review_obj.get("auto_reply") or "").strip()
        if not auto_reply_text:
            raise HTTPException(status_code=400, detail="Нет сохранённого текста для повторной отправки")
        account_id = review_obj.get("account_id")
        account = repository.get_marketplace_account(
            user_id=owner_uid,
            account_id=int(account_id),
            include_secrets=True,
        ) if account_id else None
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет не найден")
        client = service._build_client(account)
        review_input = ReviewInput(
            review_id=str(review_obj.get("external_review_id") or ""),
            text=str(review_obj.get("text") or ""),
            metadata=review_obj.get("metadata") or {},
        )
        try:
            sent = client.send_review_reply(review=review_input, response_text=auto_reply_text)
        except Exception as exc:
            repository.mark_review_send_error(
                user_id=owner_uid,
                review_uid=review_uid,
                error_message=str(exc),
            )
            raise HTTPException(status_code=502, detail=f"Не удалось отправить ответ: {exc}")
        if not sent:
            raise HTTPException(status_code=502, detail="Ответ не был отправлен")
        repository.clear_review_send_error(user_id=owner_uid, review_uid=review_uid)
        repository.update_review_processing_result(
            user_id=owner_uid,
            review_uid=review_uid,
            status="answered_auto",
            auto_reply=auto_reply_text,
        )
        return {"ok": True}

    @app.get("/api/conversations")
    def list_conversations(
        request: Request,
        source: str | None = None,
        account_id: int | None = None,
        kind: str | None = None,
        status: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        sort: str = "newest",
        bucket: str = "new",
        page: int = 1,
        page_size: int = 30,
        search: str | None = None,
    ) -> dict[str, object]:
        user = _require_user(request)
        normalized_source = (source or "").strip().lower()
        if not normalized_source or normalized_source == "all":
            normalized_source = None
        kind_key = (kind or "").strip().lower()
        if not kind_key or kind_key == "all":
            normalized_kind = None
        elif kind_key in {"question", "chat"}:
            normalized_kind = kind_key
        else:
            raise HTTPException(status_code=400, detail="Тип должен быть: вопрос, чат или все")
        status_key = (status or "").strip().lower()
        if not status_key or status_key == "all":
            normalized_status = None
            status_key = "all"
        elif status_key in {"open", "waiting", "closed"}:
            normalized_status = status_key
        else:
            raise HTTPException(status_code=400, detail="Статус должен быть: открыт, ожидает, закрыт или все")
        normalized_sort = sort.strip().lower()
        if normalized_sort not in {"newest", "oldest"}:
            normalized_sort = "newest"
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        normalized_bucket = bucket.strip().lower()
        if normalized_bucket not in {"all", "new", "processed"}:
            normalized_bucket = "new"
        normalized_page_size = page_size if page_size in {10, 30, 50, 100, 200, 500, 1000} else 30
        manager_conversation_scope = _manager_allowed_conversation_accounts(user)
        # Use owner's user_id for data queries — managers share owner's data
        conv_owner_user_id = _tenant_owner_id(user)
        page_data = repository.list_conversations_paginated(
            user_id=conv_owner_user_id,
            source=normalized_source,
            account_id=account_id,
            kind=normalized_kind,
            status=normalized_status,
            statuses=None,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
            sort=normalized_sort,
            page=max(page, 1),
            page_size=normalized_page_size,
            bucket=normalized_bucket,
            search=str(search).strip() if search else None,
            account_permissions=manager_conversation_scope,
        )
        source_options = repository.list_conversation_sources(user_id=conv_owner_user_id)
        account_options = repository.list_conversation_accounts(
            user_id=conv_owner_user_id, kind=normalized_kind
        )
        # For answered questions: enrich with last sent text from conversation_messages
        # (text answered via our system) and portal reply from metadata.raw.answer.text
        items = page_data["items"]
        if normalized_kind == "question" and normalized_bucket == "processed":
            answered_uids = [
                item["conversation_uid"] for item in items
                if item.get("processed_by_operator") or item.get("last_sent_at")
            ]
            if answered_uids:
                try:
                    sent_texts = repository.get_last_sent_text_for_conversations(
                        user_id=conv_owner_user_id, conversation_uids=answered_uids
                    )
                    for item in items:
                        uid = item["conversation_uid"]
                        # Priority: our system reply > portal reply (metadata.raw.answer.text)
                        item["last_sent_text"] = sent_texts.get(uid, "")
                except Exception:
                    pass
        # Enrich questions with product photo URLs
        if normalized_kind == "question":
            try:
                _photo_map = repository.get_product_photo_map(user_id=conv_owner_user_id)
                if _photo_map:
                    for _item in items:
                        _meta = _item.get("metadata") or {}
                        _raw = (_meta.get("raw") or {}) if isinstance(_meta, dict) else {}
                        _pd = (_raw.get("productDetails") or {}) if isinstance(_raw, dict) else {}
                        _keys = [
                            str(_raw.get("sku") or "").strip(),           # Ozon SKU
                            str(_pd.get("supplierArticle") or "").strip(), # WB supplier article
                            str(_pd.get("nmId") or "").strip(),            # WB nmId
                            str(_raw.get("supplierArticle") or "").strip(),# WB supplier article alt
                        ]
                        _item["product_photo_url"] = next(
                            (_photo_map[k] for k in _keys if k and k in _photo_map), None
                        )
            except Exception:
                pass
        return {
            "items": items,
            "count": len(items),
            "total": page_data["total"],
            "page": page_data["page"],
            "page_size": page_data["page_size"],
            "pages": page_data["pages"],
            "new_count": page_data["new_count"],
            "processed_count": page_data["processed_count"],
            "bucket": normalized_bucket,
            "sort": normalized_sort,
            "date_from": normalized_date_from,
            "date_to": normalized_date_to,
            "source": normalized_source or "all",
            "status": status_key,
            "kind": normalized_kind or "all",
            "source_options": source_options,
            "account_options": account_options,
        }

    @app.post("/api/conversations/{conversation_uid}/status")
    def set_conversation_status(
        conversation_uid: str,
        payload: ConversationStatusRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        status_value = payload.status.strip().lower()
        if status_value not in {"open", "waiting", "closed"}:
            raise HTTPException(status_code=400, detail="Статус должен быть: открыт, ожидает или закрыт")
        updated = repository.update_conversation_status(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
            status=status_value,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        repository.log_review_action(
            user_id=owner_uid,
            review_uid=conversation_uid,
            action_type="conversation_status",
            actor=str(user["email"]),
            details={"status": status_value},
        )
        return {"ok": True}

    @app.post("/api/conversations/{conversation_uid}/mark-answered")
    def mark_conversation_answered(
        conversation_uid: str,
        request: Request,
    ) -> dict[str, object]:
        """Move a chat to the 'answered' bucket by setting last_sent_at = now.

        Useful for ad/promo chats where the seller does not need to reply but
        wants to clear them from the 'needs reply' queue.
        """
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        conversation = repository.get_conversation(
            user_id=owner_uid, conversation_uid=conversation_uid
        )
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        updated = repository.mark_conversation_answered(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        repository.log_review_action(
            user_id=owner_uid,
            review_uid=conversation_uid,
            action_type="conversation_mark_answered",
            actor=str(user["email"]),
            details={"manual": True},
        )
        return {"ok": True}

    @app.post("/api/conversations/{conversation_uid}/move-to-new")
    def move_conversation_to_new(
        conversation_uid: str,
        request: Request,
    ) -> dict[str, object]:
        """Move a chat to the 'new' bucket by clearing last_sent_at.

        Used when the operator manually moves an answered chat back to New,
        e.g. if they want to re-process it.
        """
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        conversation = repository.get_conversation(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
        )
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        updated = repository.move_conversation_to_new(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        repository.log_review_action(
            user_id=owner_uid,
            review_uid=conversation_uid,
            action_type="conversation_move_to_new",
            actor=str(user["email"]),
            details={"manual": True},
        )
        return {"ok": True}

    @app.post("/api/conversations/{conversation_uid}/reply")
    def reply_conversation(
        conversation_uid: str,
        payload: ConversationReplyRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        operator_name = str(user.get("full_name") or user.get("email") or "").strip() or "Продавец"
        idempotency_key = (payload.idempotency_key or "").strip() or f"{conversation_uid}:{int(time.time() * 1000)}"
        result = service.send_conversation_reply(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
            response_text=payload.response_text,
            operator_name=operator_name,
            idempotency_key=idempotency_key,
        )
        if not bool(result.get("ok")):
            raise HTTPException(status_code=502, detail=str(result.get("error") or "Не удалось отправить ответ в диалог"))
        return {
            "ok": True,
            "status": result.get("status"),
            "deduplicated": bool(result.get("deduplicated")),
            "idempotency_key": idempotency_key,
        }

    @app.get("/api/wb-image")
    def wb_image_proxy(request: Request, id: str, account_id: int) -> object:
        """Proxy WB chat images via /api/v1/seller/download/{id}.

        WB returns image URLs pointing to internal K8s addresses that are
        not publicly reachable. The downloadID field allows fetching via
        the public buyer-chat-api endpoint with Authorization header.
        """
        from fastapi.responses import Response as _Response
        user = _require_user(request)
        clean_id = str(id or "").strip()
        if not clean_id:
            raise HTTPException(status_code=400, detail="Missing image id")
        account = repository.get_marketplace_account(
            user_id=_tenant_owner_id(user),
            account_id=account_id,
            include_secrets=True,
        )
        if account is None or str(account.get("marketplace") or "") != "wb":
            raise HTTPException(status_code=404, detail="WB account not found")
        api_key = str(account.get("api_key") or "").strip()
        extra = account.get("extra") if isinstance(account.get("extra"), dict) else {}
        chats_api_url = str(extra.get("chats_api_url") or "https://buyer-chat-api.wildberries.ru").rstrip("/")
        if not api_key:
            raise HTTPException(status_code=400, detail="WB credentials missing")
        download_url = f"{chats_api_url}/api/v1/seller/download/{clean_id}"
        try:
            req = urllib.request.Request(
                download_url,
                method="GET",
                headers={"Authorization": api_key},
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                content = resp.read()
                content_type = resp.headers.get("Content-Type", "image/jpeg")
            return _Response(content=content, media_type=content_type)
        except urllib.error.HTTPError as exc:
            _log.warning("wb_image_proxy: HTTP %d for id=%s", exc.code, clean_id)
            raise HTTPException(status_code=502, detail=f"WB returned HTTP {exc.code}")
        except Exception as exc:
            _log.warning("wb_image_proxy: error %s for id=%s", exc, clean_id)
            raise HTTPException(status_code=502, detail="Failed to fetch WB image")

    @app.get("/api/ozon-image")
    def ozon_image_proxy(request: Request, url: str, account_id: int) -> object:
        """Proxy Ozon chat images that require Client-Id/Api-Key authentication.

        The browser cannot load Ozon image URLs directly (they return 401).
        This endpoint fetches the image server-side using the stored credentials
        and streams it back to the browser.
        """
        from fastapi.responses import Response as _Response
        user = _require_user(request)
        if not url.startswith("https://api-seller.ozon.ru/"):
            raise HTTPException(status_code=400, detail="Invalid Ozon image URL")
        account = repository.get_marketplace_account(
            user_id=_tenant_owner_id(user),
            account_id=account_id,
            include_secrets=True,
        )
        if account is None or str(account.get("marketplace") or "") != "ozon":
            raise HTTPException(status_code=404, detail="Ozon account not found")
        api_key = str(account.get("api_key") or "").strip()
        extra = account.get("extra") if isinstance(account.get("extra"), dict) else {}
        client_id = str(extra.get("client_id") or "").strip()
        _log.info(
            "ozon_image_proxy: account_id=%d client_id=%r api_key_len=%d",
            account_id, client_id, len(api_key),
        )
        if not api_key or not client_id:
            raise HTTPException(status_code=400, detail="Ozon credentials missing")
        try:
            req = urllib.request.Request(
                url,
                method="GET",
                headers={"Client-Id": client_id, "Api-Key": api_key},
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                content = resp.read()
                content_type = resp.headers.get("Content-Type", "image/jpeg")
            return _Response(content=content, media_type=content_type)
        except urllib.error.HTTPError as exc:
            _log.warning("ozon_image_proxy: HTTP %d for %s", exc.code, url[:80])
            raise HTTPException(status_code=502, detail=f"Ozon returned HTTP {exc.code}")
        except Exception as exc:
            _log.warning("ozon_image_proxy: error %s for %s", exc, url[:80])
            raise HTTPException(status_code=502, detail="Failed to fetch Ozon image")

    @app.get("/api/conversations/{conversation_uid}/messages")
    def conversation_messages(conversation_uid: str, request: Request, limit: int = 200, refresh: int = 0) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        conversation = repository.get_conversation(user_id=owner_uid, conversation_uid=conversation_uid)
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        messages = repository.list_conversation_messages(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
            limit=limit,
        )
        # For WB chats: fetch events from WB API when:
        # - messages table is empty (first load), OR
        # - refresh=1 parameter passed (force reload of full history), OR
        # - conversation.last_message_at is newer than the newest message in DB
        #   (buyer sent a new message AFTER our reply), OR
        # - no inbound (buyer) messages in DB at all — buyer wrote BEFORE our reply
        #   and auto-sync never saved their message text, OR
        # - unread_count > 0 means buyer has messages we haven't shown yet
        _should_refresh = not messages or bool(refresh)
        if not _should_refresh and messages and str(conversation.get("source") or "") == "wb":
            conv_last_msg = str(conversation.get("last_message_at") or "").strip()
            db_newest = str(messages[-1].get("created_at") or "").strip() if messages else ""
            # If conversation updated more recently than newest DB message → refresh
            if conv_last_msg and db_newest and conv_last_msg > db_newest:
                _should_refresh = True
            # If no inbound (buyer) messages in DB → buyer wrote before our reply
            # and auto-sync never stored their message text → always fetch history
            if not _should_refresh:
                has_inbound = any(str(m.get("direction") or "") == "inbound" for m in messages)
                if not has_inbound:
                    _should_refresh = True
            # If WB still reports unread messages → buyer has messages not yet in DB
            if not _should_refresh:
                unread = int(conversation.get("unread_count") or 0)
                if unread > 0:
                    _should_refresh = True
        if _should_refresh and str(conversation.get("source") or "") == "wb":
            try:
                account_id = conversation.get("account_id")
                ext_id = str(conversation.get("external_conversation_id") or "")
                if account_id and ext_id:
                    account = repository.get_marketplace_account(
                        user_id=owner_uid,
                        account_id=int(account_id),
                        include_secrets=True,
                    )
                    if account:
                        client = service._build_client(account)
                        # Fetch events starting from since_date (or 30 days ago fallback).
                        # This gives only RECENT events — fast (1-2 requests instead of 83).
                        # Use sync_start_date from user settings as the start cursor.
                        user_sync_settings = repository.get_user_sync_settings(user_id=owner_uid)
                        event_since = (
                            str(user_sync_settings.get("sync_start_date") or "").strip()
                            if bool(user_sync_settings.get("use_sync_start_date"))
                            else None
                        )
                        if not event_since:
                            # Default: events from last 30 days
                            event_since = (datetime.now(UTC) - timedelta(days=30)).date().isoformat()
                        # Convert since date to ms timestamp for cursor
                        if hasattr(client, "_to_wb_unix_timestamp"):
                            since_ts = client._to_wb_unix_timestamp(event_since)  # type: ignore[attr-defined]
                            # WB events cursor is in milliseconds
                            resume_cursor = str(since_ts * 1000) if since_ts else None
                        else:
                            resume_cursor = None
                        sender_map = client._fetch_last_sender_map(  # type: ignore[attr-defined]
                            resume_cursor=resume_cursor,
                            stop_requested=None,
                        )
                        sender_map.pop("_final_cursor", None)
                        entry = sender_map.get(ext_id, {})
                        wb_events = entry.get("events") or []
                        # Status update is handled by the 60s auto-sync; skip here.
                        history: list[dict[str, object]] = []
                        for ev in wb_events:
                            if not isinstance(ev, dict):
                                continue
                            ev_id = str(ev.get("eventID") or "").strip()
                            ev_sender = str(ev.get("sender") or "").strip().lower()
                            msg = ev.get("message") or {}
                            ev_text = str(msg.get("text") or "").strip()
                            if not ev_text:
                                attachments = msg.get("attachments") or {}
                                images = attachments.get("images") or []
                                if images:
                                    img_parts = [f"[img:{_wb_image_url(img)}]" for img in images if img.get("url") or img.get("downloadID")]
                                    ev_text = " ".join(img_parts) if img_parts else f"[Фото: {len(images)} шт.]"
                                elif attachments.get("goodCard"):
                                    ev_text = f"[Товар: {attachments['goodCard'].get('name', '')}]"
                            ev_ts_raw = ev.get("addTimestamp")
                            ev_ts_ms = int(ev_ts_raw) if ev_ts_raw is not None else 0
                            # Fallback: images[0]['date'] if addTimestamp is 0
                            if not ev_ts_ms and images and images[0].get("date"):
                                ev_iso = _normalize_timestamp(str(images[0]["date"])) or datetime.now(UTC).isoformat()
                            else:
                                ev_iso = _normalize_timestamp(ev_ts_ms) or datetime.now(UTC).isoformat()
                            client_name = str(ev.get("clientName") or "").strip()
                            if not ev_id or not ev_text:
                                continue
                            history.append({
                                "direction": "inbound" if ev_sender == "client" else "outbound",
                                "message_text": ev_text,
                                "idempotency_key": f"wb-event-{ev_id}",
                                "created_at": ev_iso,
                                "operator_name": client_name if ev_sender == "client" else "Продавец",
                            })
                        # Migrate old internal K8s URLs to wb-download tokens
                        repository.fix_wb_internal_photo_urls(
                            user_id=owner_uid,
                            conversation_uid=conversation_uid,
                        )
                        if history:
                            repository.bulk_insert_chat_history_messages(
                                user_id=owner_uid,
                                conversation_uid=conversation_uid,
                                messages=history,
                            )
                            # Move chat to "New" bucket if buyer replied after our last reply
                            try:
                                repository.move_chat_to_new_if_buyer_replied(
                                    user_id=owner_uid,
                                    conversation_uid=conversation_uid,
                                )
                            except Exception:
                                pass
                            messages = repository.list_conversation_messages(
                                user_id=owner_uid,
                                conversation_uid=conversation_uid,
                                limit=limit,
                            )
            except Exception:
                pass
        # For Ozon chats: fetch history from /v3/chat/history when empty or refresh=1
        if _should_refresh and str(conversation.get("source") or "") == "ozon":
            try:
                account_id = conversation.get("account_id")
                ext_id = str(conversation.get("external_conversation_id") or "")
                if account_id and ext_id:
                    account = repository.get_marketplace_account(
                        user_id=owner_uid,
                        account_id=int(account_id),
                        include_secrets=True,
                    )
                    if account:
                            client = service._build_client(account)
                            if hasattr(client, "_request_json") and hasattr(client, "chats_history_path"):
                                hist_body = client._request_json(  # type: ignore[attr-defined]
                                    path=client.chats_history_path,  # type: ignore[attr-defined]
                                    payload={"chat_id": ext_id, "limit": 100, "direction": "Backward"},
                                )
                                ozon_msgs = hist_body.get("messages") or []
                                history_ozon: list[dict[str, object]] = []
                                buyer_uid_web: str = ""
                                order_num_web: str = ""
                                for msg in ozon_msgs:
                                    if not isinstance(msg, dict):
                                        continue
                                    user_info = msg.get("user") or {}
                                    user_type = str(user_info.get("type") or "").lower()
                                    msg_id = str(msg.get("message_id") or "").strip()
                                    msg_ts = str(msg.get("created_at") or "")
                                    msg_text = _parse_ozon_message_text(
                                        msg.get("data"), bool(msg.get("is_image"))
                                    )
                                    if user_type == "customer":
                                        uid = str(user_info.get("id") or "").strip()
                                        if uid and not buyer_uid_web:
                                            buyer_uid_web = uid
                                        ctx = msg.get("context") or {}
                                        on = str(ctx.get("order_number") or "").strip()
                                        if on and not order_num_web:
                                            order_num_web = on
                                    if not msg_id or not msg_text:
                                        continue
                                    direction = "inbound" if user_type == "customer" else "outbound"
                                    history_ozon.append({
                                        "direction": direction,
                                        "message_text": msg_text,
                                        "idempotency_key": f"ozon-msg-{msg_id}",
                                        "created_at": msg_ts,
                                        "operator_name": "" if direction == "inbound" else "Продавец",
                                    })
                                # Update customer_name if still missing
                                if not conversation.get("customer_name"):
                                    new_name = (
                                        f"Заказ {order_num_web}" if order_num_web
                                        else (f"Покупатель {buyer_uid_web}" if buyer_uid_web else None)
                                    )
                                    if new_name:
                                        repository.update_conversation_customer_name(
                                            user_id=owner_uid,
                                            conversation_uid=conversation_uid,
                                            customer_name=new_name,
                                        )
                                # Fix any previously saved messages with old Markdown format
                                repository.fix_ozon_photo_messages(
                                    user_id=owner_uid,
                                    conversation_uid=conversation_uid,
                                )
                                if history_ozon:
                                    repository.bulk_insert_chat_history_messages(
                                        user_id=owner_uid,
                                        conversation_uid=conversation_uid,
                                        messages=history_ozon,
                                    )
                                messages = repository.list_conversation_messages(
                                    user_id=owner_uid,
                                    conversation_uid=conversation_uid,
                                    limit=limit,
                                )
            except Exception:
                pass
        return {
            "conversation": conversation,
            "messages": messages,
            "count": len(messages),
        }

    @app.get("/api/chat-quick-templates")
    def list_chat_quick_templates(request: Request) -> dict[str, object]:
        user = _require_user(request)
        items = repository.list_chat_quick_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/chat-quick-templates")
    def create_chat_quick_template(request: Request, payload: ChatQuickTemplateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.add_chat_quick_template(
            user_id=int(user["id"]), template_name=name, template_text=text
        )
        return {"ok": True, "item": item}

    @app.put("/api/chat-quick-templates/{template_id}")
    def update_chat_quick_template(
        template_id: int, request: Request, payload: ChatQuickTemplateUpdateRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.update_chat_quick_template(
            user_id=int(user["id"]),
            template_id=int(template_id),
            template_name=name,
            template_text=text,
        )
        if item is None:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "item": item}

    @app.delete("/api/chat-quick-templates/{template_id}")
    def delete_chat_quick_template(template_id: int, request: Request) -> dict[str, object]:
        user = _require_user(request)
        deleted = repository.delete_chat_quick_template(user_id=int(user["id"]), template_id=int(template_id))
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "deleted": True}

    # ── Review contradiction rules endpoints ─────────────────────────────────

    @app.get("/api/contradiction-rules")
    def list_contradiction_rules(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        try:
            repository._ensure_contradiction_rules_table()
            items = repository.list_review_contradiction_rules(user_id=int(user["id"]))
        except Exception:
            items = []
        return {"items": items}

    @app.post("/api/contradiction-rules")
    def save_contradiction_rule(
        request: Request,
        group_id: str,
        ratings: str,
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        try:
            import json as _json
            ratings_list = [int(r) for r in _json.loads(ratings) if 1 <= int(r) <= 5]
        except Exception:
            raise HTTPException(status_code=400, detail="ratings must be JSON array of ints 1-5")
        if not group_id.strip():
            raise HTTPException(status_code=400, detail="group_id is required")
        repository._ensure_contradiction_rules_table()
        repository.save_review_contradiction_rule(
            user_id=int(user["id"]),
            group_id=group_id.strip(),
            ratings=ratings_list,
        )
        return {"ok": True}

    @app.delete("/api/contradiction-rules")
    def delete_contradiction_rule(request: Request, group_id: str) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_review_contradiction_rule(
            user_id=int(user["id"]),
            group_id=group_id.strip(),
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Правило не найдено")
        return {"ok": True}

    # ── Question quick templates endpoints ───────────────────────────────────

    @app.get("/api/question-quick-templates")
    def list_question_quick_templates(request: Request) -> dict[str, object]:
        user = _require_user(request)
        items = repository.list_question_quick_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/question-quick-templates")
    def create_question_quick_template(request: Request, payload: ChatQuickTemplateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.add_question_quick_template(
            user_id=int(user["id"]), template_name=name, template_text=text
        )
        return {"ok": True, "item": item}

    @app.put("/api/question-quick-templates/{template_id}")
    def update_question_quick_template(
        template_id: int, request: Request, payload: ChatQuickTemplateUpdateRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.update_question_quick_template(
            user_id=int(user["id"]),
            template_id=int(template_id),
            template_name=name,
            template_text=text,
        )
        if item is None:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "item": item}

    @app.delete("/api/question-quick-templates/{template_id}")
    def delete_question_quick_template(template_id: int, request: Request) -> dict[str, object]:
        user = _require_user(request)
        deleted = repository.delete_question_quick_template(
            user_id=int(user["id"]), template_id=int(template_id)
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "deleted": True}

    # ── Product photos catalog ────────────────────────────────────────────────

    import os as _os
    _PHOTO_DIR = str(_os.path.join(_os.path.dirname(__file__), "..", "product_photos")).rstrip("/")

    def _ensure_product_photos_table() -> None:
        try:
            with repository._connect() as conn:
                repository._migrate_product_photos(conn)
        except Exception:
            pass

    @app.get("/api/products")
    def _list_products_ensure(request: Request) -> dict[str, object]:
        _ensure_product_photos_table()
        user = _require_settings_access(request)
        items = repository.list_product_photos(user_id=_tenant_owner_id(user))
        for item in items:
            item["photo_url"] = f"/api/products/photo/{item['id']}" if item.get("photo_path") else None
        return {"items": items}

    @app.post("/api/products")
    async def add_product(
        request: Request,
        name: str = Form(""),
        supplier_article: str = Form(""),
        wb_nmid: str = Form(""),
        ozon_sku: str = Form(""),
        photo: UploadFile | None = File(None),
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        _ensure_product_photos_table()
        owner_uid = _tenant_owner_id(user)
        photo_path: str | None = None
        if photo and photo.filename:
            import io as _io
            try:
                from PIL import Image as _PilImage
                content = await photo.read()
                img = _PilImage.open(_io.BytesIO(content)).convert("RGB")
                img.thumbnail((200, 200), _PilImage.LANCZOS)
                _os.makedirs(_PHOTO_DIR, exist_ok=True)
                import uuid as _uuid
                fname = f"{_uuid.uuid4().hex}.webp"
                fpath = _os.path.join(_PHOTO_DIR, fname)
                img.save(fpath, "WEBP", quality=85)
                photo_path = fname
            except Exception as _e:
                _log.warning("add_product: photo processing failed: %s", _e)
        item = repository.add_product_photo(
            user_id=owner_uid, name=name.strip(), supplier_article=supplier_article.strip(),
            wb_nmid=wb_nmid.strip(), ozon_sku=ozon_sku.strip(), photo_path=photo_path,
        )
        if item:
            item["photo_url"] = f"/api/products/photo/{item['id']}" if item.get("photo_path") else None
        return {"ok": True, "item": item}

    @app.put("/api/products/{product_id}")
    async def update_product(
        product_id: int,
        request: Request,
        name: str = Form(""),
        supplier_article: str = Form(""),
        wb_nmid: str = Form(""),
        ozon_sku: str = Form(""),
        photo: UploadFile | None = File(None),
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        owner_uid = _tenant_owner_id(user)
        new_photo_path: str | None = None
        if photo and photo.filename:
            import io as _io
            try:
                from PIL import Image as _PilImage
                content = await photo.read()
                img = _PilImage.open(_io.BytesIO(content)).convert("RGB")
                img.thumbnail((200, 200), _PilImage.LANCZOS)
                _os.makedirs(_PHOTO_DIR, exist_ok=True)
                import uuid as _uuid
                fname = f"{_uuid.uuid4().hex}.webp"
                fpath = _os.path.join(_PHOTO_DIR, fname)
                img.save(fpath, "WEBP", quality=85)
                new_photo_path = fname
            except Exception as _e:
                _log.warning("update_product: photo processing failed: %s", _e)
        ok = repository.update_product_photo(
            user_id=owner_uid, product_id=product_id, name=name.strip(),
            supplier_article=supplier_article.strip(), wb_nmid=wb_nmid.strip(),
            ozon_sku=ozon_sku.strip(), photo_path=new_photo_path,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Товар не найден")
        return {"ok": True}

    @app.delete("/api/products/{product_id}")
    def delete_product(product_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_product_photo(user_id=_tenant_owner_id(user), product_id=product_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Товар не найден")
        # Delete physical file
        if deleted.get("photo_path"):
            try:
                _os.remove(_os.path.join(_PHOTO_DIR, deleted["photo_path"]))
            except Exception:
                pass
        return {"ok": True}

    @app.get("/api/products/photo/{product_id}")
    def product_photo(product_id: int, request: Request) -> object:
        from fastapi.responses import FileResponse as _FileResp
        _require_user(request)
        items = repository.list_product_photos(user_id=_tenant_owner_id(_require_user(request)))
        item = next((i for i in items if i.get("id") == product_id), None)
        if not item or not item.get("photo_path"):
            raise HTTPException(status_code=404, detail="Фото не найдено")
        fpath = _os.path.join(_PHOTO_DIR, item["photo_path"])
        if not _os.path.exists(fpath):
            raise HTTPException(status_code=404, detail="Файл не найден")
        return _FileResp(fpath, media_type="image/webp")

    # Enrich /api/reviews and /api/conversations with product_photo_url
    # (done inline in list_reviews and list_conversations endpoints)

    # ── Review quick templates ────────────────────────────────────────────────

    def _ensure_review_quick_templates_table() -> None:
        try:
            with repository._connect() as conn:
                repository._migrate_review_quick_templates(conn)
        except Exception:
            pass

    @app.get("/api/review-quick-templates")
    def list_review_quick_templates(request: Request) -> dict[str, object]:
        user = _require_user(request)
        _ensure_review_quick_templates_table()
        items = repository.list_review_quick_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/review-quick-templates")
    def create_review_quick_template(request: Request, payload: ChatQuickTemplateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        _ensure_review_quick_templates_table()
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.add_review_quick_template(
            user_id=int(user["id"]), template_name=name, template_text=text
        )
        return {"ok": True, "item": item}

    @app.put("/api/review-quick-templates/{template_id}")
    def update_review_quick_template(
        template_id: int, request: Request, payload: ChatQuickTemplateUpdateRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        _ensure_review_quick_templates_table()
        with repository._connect() as conn:
            conn.execute(
                repository._sql("""
                UPDATE review_quick_templates
                SET template_name = ?, template_text = ?, updated_at = ?
                WHERE id = ? AND user_id = ?
                """),
                (name, text, _now_iso(), int(template_id), int(user["id"])),
            )
        items = repository.list_review_quick_templates(user_id=int(user["id"]))
        updated = next((i for i in items if i["id"] == int(template_id)), None)
        if updated is None:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "item": updated}

    @app.delete("/api/review-quick-templates/{template_id}")
    def delete_review_quick_template(template_id: int, request: Request) -> dict[str, object]:
        user = _require_user(request)
        deleted = repository.delete_review_quick_template(
            user_id=int(user["id"]), template_id=int(template_id)
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "deleted": True}

    @app.post("/api/admin/actions-purge-sync")
    def admin_purge_sync_actions(request: Request) -> dict[str, object]:
        """Delete all sync_review and sync_conversation entries — they are no longer logged."""
        _require_admin(request)
        deleted = repository.purge_sync_action_logs()
        _log.info("admin_purge_sync_actions: deleted %d rows", deleted)
        return {"ok": True, "deleted": deleted}

    @app.post("/api/admin/conversations-clear")
    def admin_clear_conversations(request: Request, payload: ClearConversationsRequest) -> dict[str, object]:
        actor = _require_admin(request)
        if payload.user_id is None:
            target_user_id = _tenant_owner_id(actor) if not _is_super_admin(actor) else int(actor["id"])
        else:
            target_user_id = int(payload.user_id)
            _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        deleted = repository.clear_conversations(user_id=target_user_id, kind=payload.kind, source=payload.source)
        return {"ok": True, "deleted": deleted, "user_id": target_user_id}

    @app.get("/api/analytics")
    def user_analytics(
        request: Request,
        source: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> dict[str, object]:
        user = _require_analytics_access(request)
        return repository.get_user_analytics(
            user_id=int(user["id"]),
            source=source.strip() or None,
            date_from=date_from.strip() or None,
            date_to=date_to.strip() or None,
        )

    @app.post("/api/sync")
    def sync_reviews(request: Request, payload: SyncRequest) -> dict[str, object]:
        user = _require_user(request)
        user_id = int(user["id"])
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        if payload.all_accounts or payload.account_ids:
            all_active = _snapshot_active_account_ids_for_user(user_id)
            if not all_active:
                raise HTTPException(status_code=400, detail="Нет активных кабинетов для синхронизации")
            # Filter to selected accounts if checkboxes were used
            if payload.account_ids:
                selected = set(int(x) for x in payload.account_ids if x)
                account_ids_snapshot = [aid for aid in all_active if aid in selected]
                if not account_ids_snapshot:
                    raise HTTPException(status_code=400, detail="Ни один из выбранных кабинетов не найден")
            else:
                account_ids_snapshot = all_active
            # Store expected total from preview for the progress bar
            if payload.total_expected is not None and payload.total_expected > 0:
                with sync_lock:
                    sync_state["progress_total_items"] = int(payload.total_expected)
            run_started_at = _now_iso()
            result = _run_sync_for_user(
                user_id=user_id,
                since_date=since_date or None,
                account_ids=account_ids_snapshot,
                run_started_at=run_started_at,
                apply_date_filter=True,  # manual sync applies date filter
            )
            with sync_lock:
                sync_state["polling_enabled"] = True
                sync_state["polling_user_id"] = user_id
                sync_state["polling_account_ids"] = list(account_ids_snapshot)
                sync_state["polling_since_date"] = since_date or None
                sync_state["polling_started_at"] = run_started_at
                sync_state["last_poll_at"] = run_started_at
                sync_state["last_poll_result"] = {
                    "ok": True,
                    "run_started_at": run_started_at,
                    "accounts": int(result.get("accounts") or 0),
                    "success_accounts": int(result.get("success_accounts") or 0),
                    "failed_accounts": int(result.get("failed_accounts") or 0),
                    "loaded": int(result.get("loaded") or 0),
                    "loaded_conversations": int(result.get("loaded_conversations") or 0),
                    "account_ids": list(result.get("account_ids") or account_ids_snapshot),
                    "errors": _serialize_sync_error_details(result.get("errors")),
                    "cancelled": bool(result.get("cancelled")),
                }
            _start_auto_sync_worker_if_needed()
            return result

        if payload.account_id is None:
            raise HTTPException(status_code=400, detail="Необходимо указать идентификатор кабинета")
        account = repository.get_marketplace_account(
            user_id=user_id,
            account_id=payload.account_id,
            include_secrets=True,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        marketplace = str(account["marketplace"])
        client = service._build_client(account)
        account_id_val = int(account["id"])
        loaded = 0
        loaded_conversations = 0
        errors: list[str] = []
        try:
            loaded = service.sync_reviews(
                user_id=user_id,
                source=marketplace,
                account_id=account_id_val,
                client=client,
                since_date=since_date or None,
            )
        except MarketplaceSyncError as exc:
            if not service._is_access_error(exc):
                raise HTTPException(status_code=502, detail=f"Ошибка синхронизации отзывов: {exc}") from exc
            errors.append(str(exc))
        try:
            loaded_conversations = service.sync_conversations(
                user_id=user_id,
                source=marketplace,
                account_id=account_id_val,
                client=client,
                since_date=since_date or None,
            )
        except MarketplaceSyncError as exc:
            if not service._is_access_error(exc):
                raise HTTPException(status_code=502, detail=f"Ошибка синхронизации диалогов: {exc}") from exc
            errors.append(str(exc))
        return {
            "accounts": 1,
            "loaded": loaded,
            "loaded_conversations": loaded_conversations,
            "errors": errors,
        }

    @app.post("/api/sync/capabilities")
    def sync_capabilities(request: Request, payload: SyncCapabilitiesRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        result = _probe_account_capabilities(
            user_id=user_id,
            account_id=int(payload.account_id),
            since_date=since_date or None,
        )
        return {"ok": True, "item": result}

    @app.get("/api/sync/capabilities")
    def sync_capabilities_all(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        account_ids = _snapshot_active_account_ids_for_user(user_id)
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        items: list[dict[str, object]] = []
        aggregate_errors: list[dict[str, object]] = []
        any_syncable = False
        for account_id in account_ids:
            item = _probe_account_capabilities(
                user_id=user_id,
                account_id=account_id,
                since_date=since_date or None,
            )
            items.append(item)
            any_syncable = any_syncable or bool(item.get("can_sync_any"))
            raw_item_errors = item.get("errors")
            if isinstance(raw_item_errors, list):
                aggregate_errors.extend(_serialize_sync_error_details(raw_item_errors))
        return {
            "ok": True,
            "items": items,
            "count": len(items),
            "any_syncable": any_syncable,
            "errors": aggregate_errors,
        }

    @app.get("/api/sync/preview")
    def sync_preview(request: Request) -> dict[str, object]:
        """Return estimated counts of items available to sync for all active accounts.

        Uses lightweight count endpoints (no full data download).  Results are
        used to populate the confirmation modal before starting a sync.
        """
        user = _require_user(request)
        user_id = int(user["id"])
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        accounts = [
            item
            for item in repository.list_marketplace_accounts(user_id=user_id, include_secrets=True)
            if item["is_active"]
        ]
        items: list[dict[str, object]] = []
        total_reviews = 0
        total_questions = 0
        total_chats = 0
        for account in accounts:
            try:
                result = service.count_pending_for_account(
                    account=account,
                    since_date=since_date,
                )
            except Exception as _exc:
                _log.warning(
                    "sync_preview: count_pending_for_account failed account_id=%s: %s",
                    account.get("id"), _exc,
                )
                result = {
                    "account_id": int(account.get("id") or 0),
                    "account_name": str(account.get("account_name") or ""),
                    "marketplace": str(account.get("marketplace") or ""),
                    "reviews": 0,
                    "questions": 0,
                    "chats": 0,
                    "total": 0,
                }
            items.append(result)
            total_reviews += int(result.get("reviews") or 0)
            total_questions += int(result.get("questions") or 0)
            total_chats += int(result.get("chats") or 0)
        return {
            "ok": True,
            "since_date": since_date,
            "accounts": len(items),
            "items": items,
            "total_reviews": total_reviews,
            "total_questions": total_questions,
            "total_chats": total_chats,
            "total": total_reviews + total_questions + total_chats,
        }

    @app.get("/api/sync/status")
    def sync_status_public(request: Request) -> dict[str, object]:
        """Public sync progress endpoint accessible to all logged-in users."""
        _require_user(request)
        with sync_lock:
            return {
                "in_progress": bool(sync_state.get("in_progress")),
                "is_manual": bool(sync_state.get("is_manual")),
                "cancel_requested": bool(sync_state.get("cancel_requested")),
                "last_started_at": sync_state.get("last_started_at"),
                "last_finished_at": sync_state.get("last_finished_at"),
                "step": str(sync_state.get("progress_step") or ""),
                "account": str(sync_state.get("progress_account") or ""),
                "channel": str(sync_state.get("progress_channel") or ""),
                "loaded": int(sync_state.get("progress_loaded") or 0),
                "total_items": int(sync_state.get("progress_total_items") or 0),
                "total_accounts": int(sync_state.get("progress_total_accounts") or 0),
                "current_account": int(sync_state.get("progress_current_account") or 0),
                "last_sync_report": sync_state.get("last_sync_report"),
            }

    @app.get("/api/accounts")
    def list_accounts(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        items = repository.list_marketplace_accounts(user_id=int(user["id"]), include_secrets=True)
        return {"items": items, "count": len(items)}

    @app.post("/api/accounts")
    def create_account(request: Request, payload: AccountCreateRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        marketplace = payload.marketplace.strip().lower()
        if marketplace not in {"wb", "ozon", "mock"}:
            raise HTTPException(status_code=400, detail="Некорректный маркетплейс")
        integration = payload.integration if isinstance(payload.integration, dict) else {}
        default_api_urls = {
            "wb": "https://feedbacks-api.wildberries.ru/api/v1/feedbacks",
            "ozon": "https://api-seller.ozon.ru",
            "mock": "https://example.local/api/reviews",
        }
        api_url = (payload.api_url or "").strip() or str(integration.get("api_url") or default_api_urls[marketplace])
        api_url = _validate_account_api_url(marketplace, api_url)
        if marketplace in {"wb", "ozon"} and not (payload.api_key or "").strip():
            raise HTTPException(status_code=400, detail="Для WB/OZON требуется ключ доступа")
        client_id_value = (payload.client_id or "").strip() or str(integration.get("client_id") or "").strip()
        if marketplace == "ozon" and not client_id_value:
            raise HTTPException(status_code=400, detail="Для OZON требуется идентификатор клиента")
        if client_id_value:
            integration["client_id"] = client_id_value
        if marketplace == "ozon":
            page_size = integration.get("page_size")
            if page_size is not None and (not isinstance(page_size, int) or page_size <= 0):
                raise HTTPException(status_code=400, detail="Размер страницы должен быть положительным целым числом")
        if marketplace == "wb":
            max_pages = integration.get("max_pages")
            if max_pages is not None and (not isinstance(max_pages, int) or max_pages <= 0):
                raise HTTPException(status_code=400, detail="Лимит страниц должен быть положительным целым числом")

        account = repository.create_marketplace_account(
            user_id=int(user["id"]),
            marketplace=marketplace,
            account_name=payload.account_name.strip(),
            api_url=api_url,
            api_key=(payload.api_key or "").strip() or None,
            extra=integration,
        )
        return {"ok": True, "item": account}

    @app.post("/api/accounts/{account_id}/status")
    def update_account_status(account_id: int, request: Request, payload: AccountStatusRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        account = repository.get_marketplace_account(
            user_id=int(user["id"]),
            account_id=account_id,
            include_secrets=False,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        updated = repository.update_marketplace_account_status(
            user_id=int(user["id"]),
            account_id=account_id,
            is_active=payload.is_active,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        return {"ok": True}

    @app.delete("/api/accounts/{account_id}")
    def delete_account(account_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        account = repository.get_marketplace_account(
            user_id=int(user["id"]),
            account_id=account_id,
            include_secrets=False,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        repository.update_marketplace_account_status(
            user_id=int(user["id"]),
            account_id=account_id,
            is_active=False,
        )
        deleted = repository.delete_marketplace_account(
            user_id=int(user["id"]),
            account_id=account_id,
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        return {"ok": True}

    @app.get("/api/templates")
    def list_templates(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        items = repository.list_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.get("/api/template-groups")
    def list_template_groups(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        _ensure_default_template_variants(user_id)
        rows = repository.list_template_variants(user_id=user_id)
        counts: dict[tuple[str, str], int] = {}
        for row in rows:
            key = (str(row.get("group_id") or ""), str(row.get("subgroup") or ""))
            counts[key] = counts.get(key, 0) + 1
        items = _build_template_group_items(counts)
        return {"items": items, "count": len(items)}

    @app.get("/api/processing-rules")
    def list_processing_rules(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        _ensure_default_template_variants(user_id)
        existing_rows = repository.list_processing_rules(user_id=user_id)
        existing_map = {str(row.get("group_id") or ""): row for row in existing_rows}
        items: list[dict[str, object]] = []
        for group in TEMPLATE_GROUPS:
            group_id = str(group.get("id") or "")
            title = str(group.get("title") or group_id)
            row = existing_map.get(group_id)
            mode = str((row or {}).get("action_mode") or "manual")
            if mode == "auto":
                mode = "template"
            if mode in {"ai", "ignore"}:
                mode = "manual"
            if mode not in {"template", "manual"}:
                mode = "manual"
            items.append(
                {
                    "group_id": group_id,
                    "title": title,
                    "action_mode": mode,
                    "auto_send": bool((row or {}).get("auto_send")),
                }
            )
        return {"items": items, "count": len(items)}

    @app.put("/api/processing-rules/apply")
    def apply_processing_rules(payload: ProcessingRulesApplyRequest, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        normalized_rules: list[dict[str, object]] = []
        for item in payload.rules:
            group_id = item.group_id.strip()
            if _template_group_by_id(group_id) is None:
                raise HTTPException(status_code=400, detail=f"Неизвестная группа правил: {group_id}")
            mode = item.action_mode.strip().lower()
            if mode in {"ai", "ignore"}:
                mode = "manual"
            if mode not in {"template", "manual"}:
                raise HTTPException(status_code=400, detail=f"Некорректный режим правила: {mode}")
            normalized_rules.append(
                {
                    "group_id": group_id,
                    "action_mode": mode,
                    "auto_send": bool(item.auto_send),
                }
            )
        repository.replace_processing_rules(user_id=user_id, rules=normalized_rules)
        stats = service.apply_processing_rules_to_unprocessed(user_id=user_id)
        return {"ok": True, "applied": len(normalized_rules), "updated_reviews": stats}

    @app.get("/api/recommendations")
    def list_recommendations(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        items = repository.list_recommendations(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.put("/api/recommendations")
    def save_recommendations(payload: RecommendationsSaveRequest, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        normalized_rows: list[dict[str, object]] = []
        unique_sources: set[str] = set()
        for row in payload.rows:
            source_article = row.source_article.strip()
            targets = _parse_recommendation_targets(row.targets_csv)
            if not source_article:
                continue
            if source_article in unique_sources:
                continue
            unique_sources.add(source_article)
            normalized_rows.append(
                {
                    "source_article": source_article,
                    "target_articles": targets,
                }
            )
        inserted_pairs = repository.replace_all_recommendations(
            user_id=int(user["id"]),
            rows=normalized_rows,
        )
        return {"ok": True, "sources": len(normalized_rows), "pairs": inserted_pairs}

    @app.post("/api/recommendations/import")
    async def import_recommendations(request: Request, file: UploadFile = File(...)) -> dict[str, object]:
        user = _require_settings_access(request)
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - protected by dependency
            raise HTTPException(status_code=500, detail="Библиотека Excel не установлена") from exc

        filename = (file.filename or "").lower()
        if filename and not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel формата .xlsx")
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Файл пустой")
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc
        sheet = workbook.active
        normalized_rows: list[dict[str, object]] = []
        unique_sources: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_article = str(row[0] or "").strip() if len(row) > 0 else ""
            targets_csv = str(row[1] or "").strip() if len(row) > 1 else ""
            targets = _parse_recommendation_targets(targets_csv)
            if not source_article:
                continue
            if source_article in unique_sources:
                continue
            unique_sources.add(source_article)
            normalized_rows.append(
                {
                    "source_article": source_article,
                    "target_articles": targets,
                }
            )
        inserted_pairs = repository.replace_all_recommendations(
            user_id=int(user["id"]),
            rows=normalized_rows,
        )
        return {"ok": True, "sources": len(normalized_rows), "pairs": inserted_pairs}

    @app.get("/api/recommendations/export")
    def export_recommendations(request: Request) -> StreamingResponse:
        try:
            from openpyxl import Workbook
        except Exception as exc:  # pragma: no cover - protected by dependency
            raise HTTPException(status_code=500, detail="Библиотека Excel не установлена") from exc
        user = _require_settings_access(request)
        items = repository.list_recommendations(user_id=int(user["id"]))
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Рекомендации"
        sheet.append(["Артикул товара", "Рекомендуемые артикулы"])
        for item in items:
            sheet.append(
                [
                    str(item.get("source_article") or ""),
                    str(item.get("targets_csv") or ""),
                ]
            )
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="recommendations.xlsx"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @app.get("/api/template-subgroup")
    def get_template_subgroup(group_id: str, subgroup: str, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        _ensure_default_template_variants(user_id)
        items = repository.list_template_variants(
            user_id=user_id,
            group_id=group_id,
            subgroup=subgroup,
        )
        return {"items": items, "count": len(items), "group_id": group_id, "subgroup": subgroup}

    @app.put("/api/template-subgroup")
    def save_template_subgroup(
        group_id: str,
        subgroup: str,
        payload: TemplateSubgroupSaveRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        repository.replace_subgroup_templates(
            user_id=int(user["id"]),
            group_id=group_id,
            subgroup=subgroup,
            templates=payload.templates,
        )
        return {"ok": True, "saved": len([x for x in payload.templates if x and x.strip()])}

    @app.post("/api/template-subgroup/item")
    def add_template_subgroup_item(payload: TemplateVariantCreateRequest, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        item = repository.add_template_variant(
            user_id=int(user["id"]),
            group_id=group_id,
            subgroup=subgroup,
            template_text=payload.template_text,
        )
        return {"ok": True, "item": item}

    @app.delete("/api/template-subgroup/item/{template_id}")
    def delete_template_subgroup_item(template_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        # Check if the template belongs to a protected subgroup before deleting
        existing = repository.get_template_variant_by_id(
            user_id=int(user["id"]),
            template_id=template_id,
        )
        if existing and _is_protected_subgroup(
            str(existing.get("group_id") or ""),
            str(existing.get("subgroup") or ""),
        ):
            raise HTTPException(status_code=403, detail="Шаблоны этой подгруппы защищены от удаления")
        deleted = repository.delete_template_variant(
            user_id=int(user["id"]),
            template_id=template_id,
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True}

    @app.post("/api/templates/reset-to-defaults")
    def reset_templates_to_defaults(request: Request) -> dict[str, object]:
        """Reset all user templates to the current admin defaults.

        Only available to admin/owner — not managers.
        Deletes all existing user templates and copies from default_template_variants.
        """
        user = _require_admin(request)
        owner_id = _tenant_owner_id(user)
        deleted_and_replaced = repository.reset_templates_to_defaults(user_id=owner_id)
        return {"ok": True, "replaced": deleted_and_replaced}

    @app.put("/api/templates")
    def upsert_template(request: Request, payload: TemplateUpsertRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        category = payload.category.strip().lower()
        mode = payload.mode.strip().lower()
        if category not in CATEGORIES:
            raise HTTPException(status_code=400, detail=f"Неизвестная категория: {category}")
        if mode not in {"auto", "manual", "ignore"}:
            raise HTTPException(status_code=400, detail="Режим должен быть: авто, вручную или игнор")
        repository.upsert_template(
            user_id=int(user["id"]),
            category=category,
            mode=mode,
            template_text=payload.template_text.strip(),
            is_enabled=payload.is_enabled,
        )
        return {"ok": True}

    @app.delete("/api/templates/{category}")
    def delete_template(category: str, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        normalized = category.strip().lower()
        if normalized not in CATEGORIES:
            raise HTTPException(status_code=400, detail=f"Неизвестная категория: {normalized}")
        deleted = repository.delete_template(user_id=int(user["id"]), category=normalized)
        if not deleted:
            raise HTTPException(status_code=404, detail="Правило не найдено")
        return {"ok": True}

    @app.post("/api/reviews/{review_id}/queue-manual")
    def queue_manual(review_id: str, request: Request) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_review(user, review_id)
        updated = service.queue_for_manual_processing_with_actor(
            actor_email=str(user.get("email") or ""),
            owner_user_id=_tenant_owner_id(user),
            review_uid=review_id,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        return {"ok": True}

    @app.post("/api/reviews/{review_id}/auto-reply")
    def auto_reply(review_id: str, request: Request) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_review(user, review_id)
        try:
            reply = service.generate_auto_reply(user_id=_tenant_owner_id(user), review_uid=review_id)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail=str(exc) or "Отзыв не найден") from exc
        except MarketplaceSyncError as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось отправить ответ в маркетплейс: {exc}") from exc
        return {"ok": True, "reply": reply}

    @app.post("/api/reviews/{review_id}/manual-reply")
    def manual_reply(review_id: str, payload: ManualReplyRequest, request: Request) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_review(user, review_id)
        updated = service.save_manual_reply_with_actor(
            actor_email=str(user.get("email") or ""),
            owner_user_id=_tenant_owner_id(user),
            review_uid=review_id,
            response_text=payload.response_text,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        return {"ok": True}

    @app.get("/api/admin/ai-settings")
    def get_ai_settings(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        return repository.get_ai_settings()

    @app.put("/api/admin/ai-settings")
    def update_ai_settings(request: Request, payload: AISettingsRequest) -> dict[str, object]:
        _require_super_admin(request)
        provider = payload.provider.strip().lower()
        if provider not in {"rules", "yandex"}:
            raise HTTPException(status_code=400, detail="Провайдер должен быть: встроенные правила или Яндекс")
        lookback_days = int(payload.default_sync_lookback_days)
        repository.update_ai_settings(
            provider=provider,
            yandex_api_key=payload.yandex_api_key.strip() if payload.yandex_api_key is not None else None,
            yandex_folder_id=(payload.yandex_folder_id or "").strip() or None,
            yandex_model_uri=(payload.yandex_model_uri or "").strip() or None,
            group_processors=payload.group_processors,
            use_sync_start_date=False,
            sync_start_date=None,
        )
        repository.set_default_sync_lookback_days(days=lookback_days)
        return {"ok": True}

    @app.post("/api/admin/ai-settings/check")
    def check_ai_settings_connection(request: Request, payload: AIConnectionTestRequest) -> dict[str, object]:
        _require_super_admin(request)
        stored = repository.get_ai_settings(include_secrets=True)
        api_key = (payload.yandex_api_key or "").strip() or str(stored.get("yandex_api_key") or "").strip()
        folder_id = (payload.yandex_folder_id or "").strip() or str(stored.get("yandex_folder_id") or "").strip()
        if not api_key:
            raise HTTPException(status_code=400, detail="Укажите API-ключ Yandex Cloud.")
        if not folder_id:
            raise HTTPException(status_code=400, detail="Укажите ID каталога (folderId).")
        try:
            result = service.check_yandex_connection(api_key=api_key, folder_id=folder_id)
            return {
                "ok": True,
                "status": "ok",
                "message": str(result.get("message") or "Подключение успешно"),
                "model_uri": result.get("model_uri"),
                "response_preview": result.get("response_preview"),
            }
        except MarketplaceSyncError as exc:
            detail = str(exc).lower()
            error_code = "connection"
            if any(code in detail for code in ["401", "403", "unauthorized", "forbidden", "invalid api"]):
                error_code = "auth"
            elif any(code in detail for code in ["400", "404", "folder", "modeluri", "not found"]):
                error_code = "config"
            elif "429" in detail or "rate" in detail or "quota" in detail:
                error_code = "rate_limit"
            elif "timeout" in detail or "network" in detail:
                error_code = "network"
            return {"ok": False, "status": "error", "error_code": error_code, "error": str(exc)}

    @app.post("/api/admin/ai-settings/test-review")
    def test_ai_review_classification(request: Request, payload: AIReviewTestRequest) -> dict[str, object]:
        user = _require_super_admin(request)
        stored = repository.get_ai_settings(include_secrets=True)
        api_key = (payload.yandex_api_key or "").strip() or str(stored.get("yandex_api_key") or "").strip()
        folder_id = (payload.yandex_folder_id or "").strip() or str(stored.get("yandex_folder_id") or "").strip()
        review_text = str(payload.review_text or "").strip()
        if not api_key:
            raise HTTPException(status_code=400, detail="Укажите API-ключ Yandex Cloud.")
        if not folder_id:
            raise HTTPException(status_code=400, detail="Укажите ID каталога (folderId).")
        if not review_text:
            raise HTTPException(status_code=400, detail="Введите текст тестового отзыва.")
        try:
            result = service.classify_test_review_with_yandex(
                user_id=int(user["id"]),
                review_text=review_text,
                review_rating=payload.review_rating,
                settings={
                    "yandex_api_key": api_key,
                    "yandex_folder_id": folder_id,
                    "yandex_model_uri": str(stored.get("yandex_model_uri") or "") or None,
                },
            )
            return {
                "ok": True,
                "status": "ok",
                "group_id": result.get("group_id"),
                "group_title": result.get("group_title"),
                "subgroup_id": result.get("subgroup_id"),
                "subgroup": result.get("subgroup"),
                "model_uri": result.get("model_uri"),
                "raw_response": result.get("raw_response"),
            }
        except MarketplaceSyncError as exc:
            detail = str(exc).lower()
            error_code = "classification"
            if any(code in detail for code in ["401", "403", "unauthorized", "forbidden", "invalid api"]):
                error_code = "auth"
            elif any(code in detail for code in ["400", "404", "folder", "modeluri", "not found"]):
                error_code = "config"
            elif "429" in detail or "rate" in detail or "quota" in detail:
                error_code = "rate_limit"
            elif "timeout" in detail or "network" in detail:
                error_code = "network"
            return {
                "ok": False,
                "status": "error",
                "error_code": error_code,
                "error": str(exc),
                "debug": exc.details if isinstance(exc.details, dict) else {},
            }
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось выполнить тестовый запрос к Yandex GPT: {exc}") from exc

    @app.get("/api/admin/ai-settings/active-ids")
    def list_ai_active_ids(request: Request) -> dict[str, object]:
        user = _require_super_admin(request)
        owner_user_id = _tenant_owner_id(user)
        options = service._list_group_subgroups_for_review_classification(
            repository=repository,
            user_id=owner_user_id,
        )
        items: list[dict[str, object]] = []
        for group in options:
            group_id = str(group.get("group_id") or "").strip()
            if not group_id or group_id == service.TEXTLESS_GROUP_ID:
                continue
            subgroup_items_raw = group.get("subgroup_items")
            subgroup_items: list[dict[str, str]] = []
            if isinstance(subgroup_items_raw, list):
                for subgroup_item in subgroup_items_raw:
                    if not isinstance(subgroup_item, dict):
                        continue
                    subgroup_id = str(subgroup_item.get("subgroup_id") or "").strip()
                    subgroup_title = str(subgroup_item.get("subgroup") or "").strip()
                    if not subgroup_id or not subgroup_title:
                        continue
                    subgroup_items.append(
                        {
                            "subgroup_id": subgroup_id,
                            "subgroup": subgroup_title,
                        }
                    )
            if not subgroup_items:
                continue
            items.append(
                {
                    "group_id": group_id,
                    "group_title": str(group.get("group_title") or group_id),
                    "subgroup_items": subgroup_items,
                }
            )
        return {"ok": True, "items": items, "count": len(items)}

    @app.get("/api/admin/ai-usage-stats")
    def get_ai_usage_stats(request: Request, days: int = 30) -> dict[str, object]:
        """Return daily Yandex GPT usage statistics for the last N days."""
        user = _require_admin(request)
        owner_id = _tenant_owner_id(user)
        rows = repository.get_ai_usage_stats(user_id=owner_id, days=min(max(days, 1), 90))
        # Estimate cost: Yandex YandexGPT Lite ≈ 0.20₽ per 1000 tokens input, 0.60₽ output
        # (approximate — actual pricing may differ)
        total_requests = sum(int(r.get("requests") or 0) for r in rows)
        total_input = sum(int(r.get("input_tokens") or 0) for r in rows)
        total_output = sum(int(r.get("output_tokens") or 0) for r in rows)
        return {
            "ok": True,
            "rows": rows,
            "totals": {
                "requests": total_requests,
                "input_tokens": total_input,
                "output_tokens": total_output,
                "total_tokens": total_input + total_output,
            },
        }

    @app.get("/api/admin/ai-request-log")
    def get_ai_request_log(request: Request, limit: int = 200) -> dict[str, object]:
        """Return recent Yandex GPT requests (last 1 day) for debugging."""
        user = _require_admin(request)
        owner_id = _tenant_owner_id(user)
        try:
            repository.purge_old_ai_request_logs(user_id=owner_id)
        except Exception:
            pass
        try:
            logs = repository.list_ai_request_logs(user_id=owner_id, limit=min(max(limit, 1), 500))
        except Exception as exc:
            _log.warning("ai-request-log: table not ready yet: %s", exc)
            logs = []
        return {"ok": True, "logs": logs, "count": len(logs)}

    @app.get("/api/admin/context")
    def get_admin_context(request: Request) -> dict[str, object]:
        user = _require_admin(request)
        user_id = int(user["id"])
        owner_user_id = _tenant_owner_id(user)
        manager_permissions: list[dict[str, object]] = []
        if str(user.get("role") or "").strip().lower() == TENANT_ROLE_MANAGER:
            manager_permissions = _manager_permissions_context_for_user(user)
        return {
            "user_id": user_id,
            "owner_user_id": owner_user_id,
            "is_super_admin": _is_super_admin(user),
            "is_tenant_owner": user_id == owner_user_id and not _is_super_admin(user),
            "manager_permissions": manager_permissions,
        }

    @app.get("/api/super-admin/settings")
    def super_admin_settings(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        return repository.get_super_admin_settings()

    @app.put("/api/super-admin/settings")
    def super_admin_update_settings(payload: SuperAdminSettingsRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        ai_provider = payload.ai_provider.strip().lower()
        if ai_provider not in {"rules", "yandex"}:
            raise HTTPException(status_code=400, detail="Провайдер должен быть: встроенные правила или Яндекс")
        repository.save_super_admin_settings(
            payment_provider=(payload.payment_provider or "").strip() or "manual",
            payment_api_key=payload.payment_api_key.strip() if payload.payment_api_key is not None else None,
            ai_provider=ai_provider,
            yandex_api_key=payload.yandex_api_key.strip() if payload.yandex_api_key is not None else None,
            yandex_folder_id=(payload.yandex_folder_id or "").strip() or None,
            yandex_model_uri=(payload.yandex_model_uri or "").strip() or None,
            group_processors=payload.group_processors,
            use_sync_start_date=False,
            sync_start_date=None,
            default_sync_lookback_days=int(payload.default_sync_lookback_days),
        )
        return {"ok": True}

    @app.get("/api/super-admin/template-variables")
    def super_admin_list_template_variables(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        items = repository.list_template_variables(only_active=False)
        return {"items": items, "count": len(items)}

    @app.put("/api/super-admin/template-variables")
    def super_admin_upsert_template_variable(
        payload: TemplateVariableUpsertRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        normalized_key = payload.var_key.strip().upper()
        if not TEMPLATE_VARIABLE_KEY_RE.fullmatch(normalized_key):
            raise HTTPException(
                status_code=400,
                detail="Ключ переменной должен быть в формате %NAME% и содержать только A-Z, 0-9 и _ (2-50 символов).",
            )
        source_type = (payload.source_type or "").strip().lower() or "manual"
        if source_type not in {"manual", "review_field", "system"}:
            raise HTTPException(status_code=400, detail="source_type должен быть manual, review_field или system")
        item = repository.upsert_template_variable(
            var_key=normalized_key,
            title=payload.title.strip(),
            description=(payload.description or "").strip() or None,
            is_user_editable=bool(payload.is_user_editable),
            source_type=source_type,
            source_path=(payload.source_path or "").strip() or None,
            default_value=(payload.default_value or "").strip() or None,
            is_active=bool(payload.is_active),
        )
        return {"ok": True, "item": item}

    @app.delete("/api/super-admin/template-variables")
    def super_admin_delete_template_variable(
        payload: TemplateVariableDeleteRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        deleted = repository.delete_template_variable(var_key=payload.var_key.strip().upper())
        if not deleted:
            raise HTTPException(status_code=404, detail="Переменная шаблона не найдена")
        return {"ok": True}

    @app.get("/api/super-admin/tariffs")
    def super_admin_list_tariffs(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        items = repository.list_tariff_plans()
        return {"items": items, "count": len(items)}

    @app.put("/api/super-admin/tariffs")
    def super_admin_upsert_tariff(payload: TariffPlanUpsertRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        code = payload.code.strip().lower()
        if not code:
            raise HTTPException(status_code=400, detail="Код тарифа обязателен")
        repository.upsert_tariff_plan(
            code=code,
            title=payload.title.strip(),
            monthly_price=float(payload.monthly_price),
            limits=dict(payload.limits),
            is_active=bool(payload.is_active),
        )
        return {"ok": True}

    @app.delete("/api/super-admin/tariffs")
    def super_admin_delete_tariff(payload: TariffPlanDeleteRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        code = payload.code.strip().lower()
        if not code:
            raise HTTPException(status_code=400, detail="Код тарифа обязателен")
        deleted, in_use_count = repository.delete_tariff_plan(code=code)
        if not deleted and in_use_count > 0:
            raise HTTPException(
                status_code=409,
                detail=f"Тариф используется у {in_use_count} клиентов. Сначала смените им тариф.",
            )
        if not deleted:
            raise HTTPException(status_code=404, detail="Тариф не найден")
        return {"ok": True}

    @app.get("/api/super-admin/tenants")
    def super_admin_list_tenants(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        items = repository.list_tenants_overview()
        return {"items": items, "count": len(items)}

    @app.get("/api/super-admin/default-template-groups")
    def super_admin_list_default_template_groups(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        _ensure_platform_default_templates()
        rows = repository.list_default_template_variants()
        counts: dict[tuple[str, str], int] = {}
        for row in rows:
            key = (str(row.get("group_id") or ""), str(row.get("subgroup") or ""))
            counts[key] = counts.get(key, 0) + 1
        items = _build_template_group_items(counts)
        return {"items": items, "count": len(items)}

    @app.get("/api/super-admin/default-template-subgroup")
    def super_admin_get_default_template_subgroup(
        group_id: str,
        subgroup: str,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        _ensure_platform_default_templates()
        items = repository.list_default_template_variants(group_id=group_id, subgroup=subgroup)
        return {"items": items, "count": len(items), "group_id": group_id, "subgroup": subgroup}

    @app.put("/api/super-admin/default-template-subgroup")
    def super_admin_save_default_template_subgroup(
        group_id: str,
        subgroup: str,
        payload: DefaultTemplateSubgroupSaveRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        repository.replace_default_subgroup_templates(
            group_id=group_id,
            subgroup=subgroup,
            templates=payload.templates,
        )
        return {"ok": True, "saved": len([x for x in payload.templates if x and x.strip()])}

    @app.post("/api/super-admin/default-template-subgroup")
    def super_admin_add_default_template_subgroup(
        payload: DefaultTemplateSubgroupManageRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if _template_group_by_id(group_id) is None:
            raise HTTPException(status_code=404, detail="Группа шаблонов не найдена")
        if not subgroup:
            raise HTTPException(status_code=400, detail="Название подгруппы обязательно")
        existing = {
            str(item.get("name") or "").strip()
            for item in _all_subgroups_for_group(group_id)
            if str(item.get("name") or "").strip()
        }
        if subgroup in existing:
            raise HTTPException(status_code=409, detail="Подгруппа с таким названием уже существует")
        repository.add_default_template_subgroup(group_id=group_id, subgroup=subgroup)
        return {"ok": True, "group_id": group_id, "subgroup": subgroup}

    @app.delete("/api/super-admin/default-template-subgroup")
    def super_admin_delete_default_template_subgroup(
        group_id: str,
        subgroup: str,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        clean_group_id = str(group_id or "").strip()
        clean_subgroup = str(subgroup or "").strip()
        if _template_group_by_id(clean_group_id) is None:
            raise HTTPException(status_code=404, detail="Группа шаблонов не найдена")
        if _is_protected_subgroup(clean_group_id, clean_subgroup):
            raise HTTPException(status_code=403, detail="Эта подгруппа защищена и не может быть удалена")
        if not clean_subgroup:
            raise HTTPException(status_code=400, detail="Название подгруппы обязательно")
        if _is_protected_default_subgroup(clean_group_id, clean_subgroup):
            raise HTTPException(
                status_code=400,
                detail="Подгруппы '1-3 звезды' и '4-5 звезд' в блоке 'Оценки без текста' удалять нельзя",
            )
        deleted = repository.delete_default_template_subgroup(group_id=clean_group_id, subgroup=clean_subgroup)
        if not deleted:
            raise HTTPException(status_code=404, detail="Подгруппа не найдена")
        return {"ok": True}

    @app.patch("/api/super-admin/default-template-subgroup")
    def super_admin_rename_default_template_subgroup(
        payload: DefaultTemplateSubgroupRenameRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        new_subgroup = payload.new_subgroup.strip()
        if _template_group_by_id(group_id) is None:
            raise HTTPException(status_code=404, detail="Группа шаблонов не найдена")
        if not subgroup or not new_subgroup:
            raise HTTPException(status_code=400, detail="Название подгруппы обязательно")
        if subgroup == new_subgroup:
            current = repository.get_default_template_subgroup(group_id=group_id, subgroup=subgroup)
            return {
                "ok": True,
                "group_id": group_id,
                "subgroup": subgroup,
                "new_subgroup": new_subgroup,
                "subgroup_id": str((current or {}).get("subgroup_id") or "").strip() or None,
            }
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Подгруппа не найдена")
        if _is_protected_default_subgroup(group_id, subgroup):
            raise HTTPException(status_code=400, detail="Эту системную подгруппу переименовывать нельзя")
        if _validate_subgroup(group_id, new_subgroup):
            raise HTTPException(status_code=409, detail="Подгруппа с таким названием уже существует")
        current = repository.get_default_template_subgroup(group_id=group_id, subgroup=subgroup)
        if current is None:
            raise HTTPException(status_code=404, detail="Подгруппа не найдена")
        preserved_subgroup_id = str(current.get("subgroup_id") or "").strip() or None
        renamed = repository.rename_default_template_subgroup(
            group_id=group_id,
            subgroup=subgroup,
            new_subgroup=new_subgroup,
        )
        if not renamed:
            raise HTTPException(status_code=409, detail="Не удалось переименовать подгруппу")
        return {
            "ok": True,
            "group_id": group_id,
            "subgroup": subgroup,
            "new_subgroup": new_subgroup,
            "subgroup_id": preserved_subgroup_id,
        }

    @app.post("/api/super-admin/default-template-subgroup/item")
    def super_admin_add_default_template_subgroup_item(
        payload: DefaultTemplateVariantCreateRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        item = repository.add_default_template_variant(
            group_id=group_id,
            subgroup=subgroup,
            template_text=payload.template_text,
        )
        return {"ok": True, "item": item}

    @app.post("/api/super-admin/default-template-subgroup/bulk-import")
    def super_admin_bulk_import_default_template_subgroup_items(
        payload: DefaultTemplateBulkImportRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        templates = [str(item or "").strip() for item in payload.templates]
        clean_templates = [item for item in templates if item]
        if not clean_templates:
            raise HTTPException(status_code=400, detail="Передайте хотя бы один непустой шаблон")
        added = repository.add_default_template_variants_bulk(
            group_id=group_id,
            subgroup=subgroup,
            templates=clean_templates,
        )
        return {"ok": True, "added": int(added), "group_id": group_id, "subgroup": subgroup}

    @app.delete("/api/super-admin/default-template-subgroup/item/{template_id}")
    def super_admin_delete_default_template_subgroup_item(template_id: int, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        deleted = repository.delete_default_template_variant(template_id=template_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True}

    @app.post("/api/super-admin/tenant-plan")
    def super_admin_set_tenant_plan(payload: TenantPlanUpdateRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        tenant = repository.get_user_by_id(int(payload.owner_user_id))
        if tenant is None:
            raise HTTPException(status_code=404, detail="Пользователь кабинета не найден")
        if bool(tenant.get("is_super_admin")):
            raise HTTPException(status_code=400, detail="Нельзя назначать тариф супер-администратору")
        if _tenant_owner_id(tenant) != int(tenant["id"]):
            raise HTTPException(status_code=400, detail="Тариф можно назначать только владельцу кабинета")
        updated = repository.set_tenant_plan(
            owner_user_id=int(payload.owner_user_id),
            plan_code=payload.plan_code.strip().lower(),
            limits_override=dict(payload.limits_override),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь кабинета не найден")
        return {"ok": True}

    @app.get("/api/super-admin/payments")
    def super_admin_list_payments(
        request: Request,
        owner_user_id: int | None = None,
        limit: int = 200,
    ) -> dict[str, object]:
        _require_super_admin(request)
        safe_limit = min(max(limit, 1), 1000)
        items = repository.list_billing_records(owner_user_id=owner_user_id, limit=safe_limit)
        return {"items": items, "count": len(items)}

    @app.post("/api/super-admin/payments")
    def super_admin_create_payment(payload: PaymentRecordCreateRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        tenant = repository.get_user_by_id(int(payload.owner_user_id))
        if tenant is None:
            raise HTTPException(status_code=404, detail="Пользователь кабинета не найден")
        if bool(tenant.get("is_super_admin")):
            raise HTTPException(status_code=400, detail="Нельзя привязывать оплату к супер-администратору")
        item, subscription = repository.save_payment_record_with_subscription_update(
            owner_user_id=int(payload.owner_user_id),
            amount=float(payload.amount),
            currency=payload.currency.strip().upper(),
            status=payload.status.strip().lower(),
            external_payment_id=(payload.external_payment_id or "").strip() or None,
            details=dict(payload.details),
            paid_at=(payload.paid_at or "").strip() or None,
            months=int(payload.months),
            grace_days=int(payload.grace_days),
        )
        return {"ok": True, "item": item, "subscription": subscription}

    @app.delete("/api/super-admin/payments")
    def super_admin_delete_payment(payload: PaymentRecordDeleteRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        deleted = repository.delete_payment_record(payment_id=int(payload.id))
        if not deleted:
            raise HTTPException(status_code=404, detail="Платеж не найден")
        return {"ok": True}

    @app.post("/api/super-admin/users/{target_user_id}/block")
    def super_admin_block_user(target_user_id: int, payload: UserBlockUpdateRequest, request: Request) -> dict[str, object]:
        actor = _require_super_admin(request)
        target = repository.get_user_by_id(target_user_id)
        if target is None:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        if int(target["id"]) == int(actor["id"]) and payload.blocked:
            raise HTTPException(status_code=400, detail="Нельзя заблокировать собственный аккаунт")
        updated = repository.set_user_blocked(
            user_id=target_user_id,
            blocked=bool(payload.blocked),
            reason=(payload.reason or "").strip() or None,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/super-admin/users/{target_user_id}/delete")
    def super_admin_delete_user(target_user_id: int, payload: UserDeleteRequest, request: Request) -> dict[str, object]:
        actor = _require_super_admin(request)
        if not payload.confirm:
            raise HTTPException(status_code=400, detail="Требуется подтверждение удаления")
        if int(target_user_id) == int(actor["id"]):
            raise HTTPException(status_code=400, detail="Нельзя удалить собственный аккаунт")
        deleted = repository.soft_delete_user(user_id=target_user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/tenant/team")
    def tenant_list_team(request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(owner)
        items = repository.list_tenant_users(owner_user_id=owner_id)
        for item in items:
            if str(item.get("role") or "").strip().lower() == TENANT_ROLE_MANAGER:
                item["manager_permissions"] = repository.list_manager_permissions(manager_user_id=int(item["id"]))
                item["can_supplies"] = bool(item.get("can_supplies"))
        return {"items": items, "count": len(items)}

    @app.post("/api/tenant/team")
    def tenant_create_team_user(payload: TenantUserCreateRequest, request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(owner)
        email = payload.email.strip().lower()
        if len(email) < 5 or "@" not in email:
            raise HTTPException(status_code=400, detail="Введите корректную эл. почту")
        if repository.get_user_by_email(email) is not None:
            raise HTTPException(status_code=409, detail="Пользователь с такой почтой уже существует")
        role = TENANT_ROLE_MANAGER
        created = repository.create_tenant_user(
            owner_user_id=owner_id,
            email=email,
            password_hash=hash_password(payload.password),
            role=role,
            full_name=(payload.full_name or "").strip() or None,
        )
        owner_account_ids = _manager_owner_account_ids(owner_id)
        normalized_permissions: list[dict[str, object]] = []
        for item in payload.permissions:
            account_id = int(item.account_id)
            if account_id not in owner_account_ids:
                raise HTTPException(
                    status_code=400,
                    detail=f"Кабинет {account_id} не относится к вашему профилю или недоступен",
                )
            normalized_permissions.append(
                {
                    "account_id": account_id,
                    "can_reviews": bool(item.can_reviews),
                    "can_questions": bool(item.can_questions),
                    "can_chats": bool(item.can_chats),
                }
            )
        saved_permissions = repository.replace_manager_permissions(
            manager_user_id=int(created["id"]),
            permissions=normalized_permissions,
        )
        return {
            "ok": True,
            "item": {
                "id": created.get("id"),
                "email": created.get("email"),
                "full_name": created.get("full_name"),
                "role": created.get("role"),
                "is_blocked": created.get("is_blocked"),
                "created_at": created.get("created_at"),
                "manager_permissions": repository.list_manager_permissions(manager_user_id=int(created["id"])),
                "permissions_saved": saved_permissions,
            },
        }

    @app.get("/api/tenant/team/{target_user_id}/permissions")
    def tenant_get_manager_permissions(target_user_id: int, request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца кабинета отдельные права менеджера не назначаются")
        if str(target.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            raise HTTPException(status_code=400, detail="Права можно настраивать только для менеджера")
        permissions = repository.list_manager_permissions(manager_user_id=target_user_id)
        return {"items": permissions, "count": len(permissions)}

    @app.put("/api/tenant/team/{target_user_id}/permissions")
    def tenant_update_manager_permissions(
        target_user_id: int,
        payload: ManagerPermissionsUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца кабинета отдельные права менеджера не назначаются")
        if str(target.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            raise HTTPException(status_code=400, detail="Права можно настраивать только для менеджера")
        owner_account_ids = _manager_owner_account_ids(_tenant_owner_id(owner))
        normalized_permissions: list[dict[str, object]] = []
        for item in payload.permissions:
            account_id = int(item.account_id)
            if account_id not in owner_account_ids:
                raise HTTPException(
                    status_code=400,
                    detail=f"Кабинет {account_id} не относится к вашему профилю или недоступен",
                )
            normalized_permissions.append(
                {
                    "account_id": account_id,
                    "can_reviews": bool(item.can_reviews),
                    "can_questions": bool(item.can_questions),
                    "can_chats": bool(item.can_chats),
                }
            )
        saved = repository.replace_manager_permissions(
            manager_user_id=target_user_id,
            permissions=normalized_permissions,
        )
        return {
            "ok": True,
            "saved": saved,
            "items": repository.list_manager_permissions(manager_user_id=target_user_id),
        }

    @app.put("/api/tenant/team/{target_user_id}/supplies-access")
    def tenant_set_manager_supplies_access(
        target_user_id: int,
        payload: ManagerSuppliesAccessRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца права не меняются")
        if str(target.get("role") or "").strip().lower() != TENANT_ROLE_MANAGER:
            raise HTTPException(status_code=400, detail="Применимо только для менеджера")
        repository._ensure_supply_tables()
        repository.set_user_can_supplies(user_id=target_user_id, can_supplies=payload.can_supplies)
        return {"ok": True, "can_supplies": payload.can_supplies}

    @app.post("/api/tenant/team/{target_user_id}/role")
    def tenant_update_team_role(
        target_user_id: int,
        payload: TenantUserRoleUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        role = _normalize_tenant_role_or_400(payload.role)
        if int(target["id"]) == int(owner["id"]) and role != TENANT_ROLE_OWNER:
            raise HTTPException(status_code=400, detail="Нельзя снять роль администратора у владельца кабинета")
        updated = repository.update_user_role(user_id=target_user_id, role=role)
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/tenant/team/{target_user_id}/password")
    def tenant_update_team_password(
        target_user_id: int,
        payload: AdminUserPasswordUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        updated = repository.update_user_password(
            user_id=target_user_id,
            password_hash=hash_password(payload.password),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/tenant/team/{target_user_id}/block")
    def tenant_set_user_block(
        target_user_id: int,
        payload: UserBlockUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]) and payload.blocked:
            raise HTTPException(status_code=400, detail="Нельзя заблокировать собственный аккаунт")
        updated = repository.set_user_blocked(
            user_id=target_user_id,
            blocked=bool(payload.blocked),
            reason=(payload.reason or "").strip() or None,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/tenant/team/{target_user_id}/delete")
    def tenant_delete_user(
        target_user_id: int,
        payload: UserDeleteRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        if not payload.confirm:
            raise HTTPException(status_code=400, detail="Требуется подтверждение удаления")
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Нельзя удалить владельца кабинета")
        deleted = repository.soft_delete_user(user_id=target_user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/tenant/me/plan")
    def tenant_me_plan(request: Request) -> dict[str, object]:
        user = _require_admin(request)
        owner_user_id = _tenant_owner_id(user)
        owner = repository.get_user_by_id(owner_user_id)
        if owner is None:
            raise HTTPException(status_code=404, detail="Владелец кабинета не найден")
        plans = repository.list_tariff_plans()
        current_plan_code = str(owner.get("plan_code") or "").strip().lower()
        current_plan = next((item for item in plans if str(item.get("code") or "").strip().lower() == current_plan_code), None)
        effective_limits: dict[str, object] = {}
        if current_plan and isinstance(current_plan.get("limits"), dict):
            effective_limits.update(current_plan["limits"])
        owner_override = owner.get("limits_override")
        if isinstance(owner_override, dict):
            effective_limits.update(owner_override)
        return {
            "owner_user_id": owner_user_id,
            "plan_code": current_plan_code,
            "plan": current_plan,
            "limits_override": owner_override if isinstance(owner_override, dict) else {},
            "effective_limits": effective_limits,
        }

    @app.get("/api/admin/users")
    def admin_list_users(request: Request) -> dict[str, object]:
        actor = _require_admin(request)
        if _is_super_admin(actor):
            items = repository.list_users(super_admin_only=False, owner_only=True)
        else:
            owner = _require_tenant_owner(request)
            items = repository.list_tenant_users(owner_user_id=int(owner["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/admin/users")
    def admin_create_user(payload: AdminUserCreateRequest, request: Request) -> dict[str, object]:
        actor = _require_admin(request)
        email = payload.email.strip().lower()
        if len(email) < 5 or "@" not in email:
            raise HTTPException(status_code=400, detail="Введите корректную эл. почту")
        password = payload.password
        if len(password) < 8:
            raise HTTPException(status_code=400, detail="Пароль должен быть не короче 8 символов")
        if repository.get_user_by_email(email) is not None:
            raise HTTPException(status_code=409, detail="Пользователь с такой почтой уже существует")
        if _is_super_admin(actor):
            role = ROLE_USER
            plan_code = payload.plan_code.strip().lower()
            plans = repository.list_tariff_plans()
            all_codes = {str(item.get("code") or "").strip().lower() for item in plans}
            all_codes = {code for code in all_codes if code}
            if all_codes and plan_code not in all_codes:
                # Keep user creation resilient to stale UI state:
                # if selected plan was removed, fall back to any existing tariff.
                plan_code = sorted(all_codes)[0]
            if not plan_code:
                # Keep super-admin flow operational even when tariff catalog
                # is temporarily empty or not yet configured.
                plan_code = "starter"
            created = repository.create_user(
                email=email,
                password_hash=hash_password(password),
                role=role,
                plan_code=plan_code,
            )
        else:
            owner = _require_tenant_owner(request)
            created = repository.create_tenant_user(
                owner_user_id=int(owner["id"]),
                email=email,
                password_hash=hash_password(password),
                role=_normalize_tenant_role_or_400(payload.role),
                full_name=None,
            )
        return {
            "ok": True,
            "item": {
                "id": created.get("id"),
                "email": created.get("email"),
                "full_name": created.get("full_name"),
                "role": created.get("role"),
                "is_blocked": created.get("is_blocked"),
                "created_at": created.get("created_at"),
            },
        }

    @app.post("/api/admin/users/{target_user_id}/role")
    def admin_update_user_role(target_user_id: int, payload: RoleUpdateRequest, request: Request) -> dict[str, object]:
        current_user = _require_admin(request)
        target_user = _target_user_for_admin_scope(actor=current_user, target_user_id=target_user_id)
        if _is_super_admin(current_user):
            role = payload.role.strip().lower()
            if role not in ROLE_ASSIGNABLE_BY_ADMIN:
                raise HTTPException(
                    status_code=400,
                    detail="Роль должна быть: пользователь, менеджер обратной связи или администратор",
                )
            if role != ROLE_ADMIN:
                admin_rows = repository.raw_fetch(
                    """
                    SELECT id
                    FROM users
                    WHERE role = 'admin'
                      AND is_deleted = 0
                      AND is_super_admin = 0
                      AND owner_user_id = ?
                    """,
                    (_tenant_owner_id(target_user),),
                )
                if len(admin_rows) <= 1 and any(int(item["id"]) == target_user_id for item in admin_rows):
                    raise HTTPException(status_code=400, detail="Нельзя снять роль последнего администратора клиента")
        else:
            owner = _require_tenant_owner(request)
            role = _normalize_tenant_role_or_400(payload.role)
            if int(target_user["id"]) == int(owner["id"]) and role != TENANT_ROLE_OWNER:
                raise HTTPException(status_code=400, detail="Нельзя снять роль администратора у владельца кабинета")
        updated = repository.update_user_role(user_id=target_user_id, role=role)
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True, "by_admin": current_user["email"]}

    @app.post("/api/admin/users/{target_user_id}/password")
    def admin_update_user_password(
        target_user_id: int,
        payload: AdminUserPasswordUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        updated = repository.update_user_password(
            user_id=target_user_id,
            password_hash=hash_password(payload.password),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/admin/users/{target_user_id}/plan")
    def admin_update_user_plan(
        target_user_id: int,
        payload: UserPlanUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        target_user = _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        if bool(target_user.get("is_super_admin")):
            raise HTTPException(status_code=400, detail="Нельзя менять тариф супер-администратора")
        plan_code = payload.plan_code.strip().lower()
        if not plan_code:
            raise HTTPException(status_code=400, detail="Код тарифа обязателен")
        plans = repository.list_tariff_plans()
        available_codes = {str(item.get("code") or "").strip().lower() for item in plans}
        if plan_code not in available_codes:
            raise HTTPException(status_code=404, detail="Тариф не найден")
        owner_user_id = _tenant_owner_id(target_user)
        updated = repository.set_tenant_plan(
            owner_user_id=owner_user_id,
            plan_code=plan_code,
            limits_override={},
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/admin/users/{target_user_id}/block")
    def admin_block_user(
        target_user_id: int,
        payload: UserBlockUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        target_user = _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        actor_owner_id = _tenant_owner_id(actor)
        if payload.blocked and int(target_user["id"]) == actor_owner_id:
            raise HTTPException(status_code=400, detail="Нельзя заблокировать собственный аккаунт владельца")
        updated = repository.set_user_blocked(
            user_id=target_user_id,
            blocked=bool(payload.blocked),
            reason=(payload.reason or "").strip() or None,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/admin/users/{target_user_id}/delete")
    def admin_delete_user(
        target_user_id: int,
        payload: UserDeleteRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        if not payload.confirm:
            raise HTTPException(status_code=400, detail="Требуется подтверждение удаления")
        target_user = _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        actor_owner_id = _tenant_owner_id(actor)
        if int(target_user["id"]) == actor_owner_id:
            raise HTTPException(status_code=400, detail="Нельзя удалить владельца кабинета")
        deleted = repository.soft_delete_user(user_id=target_user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/admin/metrics")
    def admin_metrics(request: Request) -> dict[str, object]:
        actor = _require_admin(request)
        if _is_super_admin(actor):
            return repository.get_sla_metrics(user_id=None)
        return repository.get_sla_metrics(user_id=_tenant_owner_id(actor))

    @app.get("/api/admin/actions")
    def admin_actions(
        request: Request,
        page: int = 1,
        page_size: int = 50,
        action_type: str | None = None,
        actor: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        search: str | None = None,
    ) -> dict[str, object]:
        admin_user = _require_admin(request)
        safe_page = max(page, 1)
        safe_page_size = min(max(page_size, 1), 200)
        safe_offset = (safe_page - 1) * safe_page_size
        normalized_action_type = (action_type or "").strip() or None
        normalized_actor = (actor or "").strip() or None
        normalized_search = (search or "").strip() or None
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        owner_scope_user_id = None if _is_super_admin(admin_user) else _tenant_owner_id(admin_user)
        if _is_super_admin(admin_user):
            rows, total = repository.list_recent_actions(
                user_id=None,
                limit=safe_page_size,
                offset=safe_offset,
                action_type=normalized_action_type,
                actor=normalized_actor,
                date_from=normalized_date_from,
                date_to=normalized_date_to,
                search=normalized_search,
            )
        else:
            rows, total = repository.list_recent_actions(
                user_id=owner_scope_user_id,
                limit=safe_page_size,
                offset=safe_offset,
                action_type=normalized_action_type,
                actor=normalized_actor,
                date_from=normalized_date_from,
                date_to=normalized_date_to,
                search=normalized_search,
            )
        filter_options = repository.list_action_filter_options(user_id=owner_scope_user_id)
        return {
            "items": rows,
            "count": len(rows),
            "total": int(total),
            "page": safe_page,
            "page_size": safe_page_size,
            "offset": safe_offset,
            "has_more": (safe_offset + len(rows)) < int(total),
            "filters": {
                "action_type": normalized_action_type or "all",
                "actor": normalized_actor or "all",
                "date_from": normalized_date_from,
                "date_to": normalized_date_to,
                "search": normalized_search or "",
            },
            "filter_options": filter_options,
        }

    @app.get("/api/admin/actions/export")
    def admin_actions_export(
        request: Request,
        format: str = "csv",
        action_type: str | None = None,
        actor: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        search: str | None = None,
    ) -> StreamingResponse:
        actor_user = _require_admin(request)
        export_format = format.strip().lower()
        if export_format not in {"csv", "xlsx"}:
            raise HTTPException(status_code=400, detail="Формат экспорта должен быть csv или xlsx")
        normalized_action_type = (action_type or "").strip() or None
        normalized_actor = (actor or "").strip() or None
        normalized_search = (search or "").strip() or None
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        scope_user_id = None if _is_super_admin(actor_user) else _tenant_owner_id(actor_user)
        items, _total = repository.list_recent_actions(
            user_id=scope_user_id,
            limit=200000,
            offset=0,
            action_type=normalized_action_type,
            actor=normalized_actor,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
            search=normalized_search,
        )
        normalized_rows: list[dict[str, str]] = []
        for item in items:
            details = item.get("details")
            details_text = ""
            if isinstance(details, dict):
                pairs: list[str] = []
                for key, value in details.items():
                    pairs.append(f"{key}={value}")
                details_text = "; ".join(pairs)
            row = {
                "created_at": str(item.get("created_at") or ""),
                "actor": str(item.get("actor") or ""),
                "review_uid": str(item.get("review_uid") or ""),
                "action_type": str(item.get("action_type") or ""),
                "details": details_text,
            }
            normalized_rows.append(row)
        if export_format == "csv":
            out = io.StringIO()
            writer = csv.DictWriter(out, fieldnames=["created_at", "actor", "review_uid", "action_type", "details"])
            writer.writeheader()
            for row in normalized_rows:
                writer.writerow(row)
            payload = io.BytesIO(out.getvalue().encode("utf-8-sig"))
            filename = f"admin-actions-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.csv"
            return StreamingResponse(
                payload,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Для экспорта Excel нужен пакет openpyxl. Установите: pip install openpyxl",
            ) from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Лента действий"
        sheet.append(["Время", "Пользователь", "Идентификатор", "Действие", "Детали"])
        for row in normalized_rows:
            sheet.append(
                [
                    row["created_at"],
                    row["actor"],
                    row["review_uid"],
                    row["action_type"],
                    row["details"],
                ]
            )
        out_xlsx = io.BytesIO()
        workbook.save(out_xlsx)
        out_xlsx.seek(0)
        xlsx_name = f"admin-actions-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.xlsx"
        return StreamingResponse(
            out_xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{xlsx_name}"'},
        )

    @app.get("/api/admin/sync-status")
    def admin_sync_status(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        with sync_lock:
            return {
                "in_progress": bool(sync_state.get("in_progress")),
                "cancel_requested": bool(sync_state.get("cancel_requested")),
                "last_started_at": sync_state.get("last_started_at"),
                "last_finished_at": sync_state.get("last_finished_at"),
                "polling_enabled": bool(sync_state.get("polling_enabled")),
                "polling_user_id": sync_state.get("polling_user_id"),
                "polling_account_ids": list(sync_state.get("polling_account_ids") or []),
                "polling_since_date": sync_state.get("polling_since_date"),
                "polling_started_at": sync_state.get("polling_started_at"),
                "last_poll_at": sync_state.get("last_poll_at"),
                "last_poll_result": sync_state.get("last_poll_result"),
            }

    @app.post("/api/admin/sync-stop")
    def admin_stop_sync(request: Request) -> dict[str, object]:
        _require_admin(request)  # Any admin (not just super-admin) can stop sync
        was_running = False
        was_polling = False
        with sync_lock:
            was_running = bool(sync_state.get("in_progress"))
            was_polling = bool(sync_state.get("polling_enabled"))
        sync_stop_event.set()
        auto_sync_stop_event.set()
        with sync_lock:
            sync_state["cancel_requested"] = True
            sync_state["polling_enabled"] = False
            sync_state["polling_user_id"] = None
            sync_state["polling_account_ids"] = []
            sync_state["polling_since_date"] = None
            sync_state["polling_started_at"] = None
            sync_state["last_poll_result"] = {
                "ok": True,
                "cancelled": True,
                "message": "Синхронизация остановлена администратором",
                "run_started_at": _now_iso(),
            }
        return {
            "ok": True,
            "was_running": bool(was_running or was_polling),
            "already_stopped": not bool(was_running or was_polling),
        }

    @app.on_event("startup")
    def restore_auto_sync_on_startup() -> None:
        """Resume background polling for all users who have active accounts.

        When the server restarts the in-memory sync_state is lost.  This hook
        reads the database to find tenant-owner users with at least one active
        marketplace account, initialises sync_state, and starts the auto-sync
        worker.  The first poll fires after AUTO_SYNC_INTERVAL_SECONDS (60 s)
        to avoid hammering the marketplace APIs right at startup.
        """
        _log.info("restore_auto_sync_on_startup: starting")
        try:
            owner_users = repository.list_users(owner_only=True)
            _log.info("restore_auto_sync_on_startup: found %d owner users", len(owner_users))
            for user in owner_users:
                uid = int(user.get("id") or 0)
                if uid <= 0:
                    continue
                try:
                    accounts = [
                        item
                        for item in repository.list_marketplace_accounts(uid, include_secrets=False)
                        if item.get("is_active")
                    ]
                    _log.info(
                        "restore_auto_sync_on_startup: user %d has %d active accounts",
                        uid, len(accounts),
                    )
                    if not accounts:
                        continue
                    account_ids = [int(a["id"]) for a in accounts if a.get("id")]
                    sync_settings = repository.get_user_sync_settings(user_id=uid)
                    since_date = (
                        str(sync_settings.get("sync_start_date") or "").strip()
                        if bool(sync_settings.get("use_sync_start_date"))
                        else None
                    )
                    # Set the first user with active accounts as the polling target.
                    # If more tenants exist they will activate their own polling when
                    # they manually trigger a sync.
                    with sync_lock:
                        if not bool(sync_state.get("polling_enabled")):
                            sync_state["polling_enabled"] = True
                            sync_state["polling_user_id"] = uid
                            sync_state["polling_account_ids"] = account_ids
                            sync_state["polling_since_date"] = since_date
                            sync_state["polling_started_at"] = _now_iso()
                    _log.info(
                        "restore_auto_sync_on_startup: starting auto-sync worker for user %d "
                        "accounts=%s since=%s",
                        uid, account_ids, since_date,
                    )
                    _start_auto_sync_worker_if_needed()
                    # Repair any chats whose answered status got lost
                    try:
                        service.repair_all_chat_statuses(user_id=uid)
                    except Exception:
                        pass
                    # Purge review_actions older than 30 days to prevent
                    # unbounded table growth with 200k+ reviews per sync
                    try:
                        purged = repository.purge_old_review_actions(keep_days=30)
                        if purged:
                            _log.info("startup: purged %d old review_actions (>30 days)", purged)
                    except Exception:
                        pass
                    # Purge AI request debug logs older than 1 day
                    try:
                        repository.purge_old_ai_request_logs()
                    except Exception:
                        pass
                    # Purge AI usage stats older than 30 days
                    try:
                        repository.purge_old_ai_usage_logs(keep_days=30)
                    except Exception:
                        pass
                    # Add manually_closed_at column if missing
                    try:
                        with repository._connect() as _conn:
                            repository._migrate_manually_closed_at(_conn)
                    except Exception:
                        pass
                    break
                except Exception as _inner_exc:
                    _log.warning("restore_auto_sync_on_startup: inner error: %s", _inner_exc)
                    continue
        except Exception as _outer_exc:
            _log.error("restore_auto_sync_on_startup: fatal error: %s", _outer_exc)
        # Start stock scheduler
        try:
            stock_scheduler.start()
        except Exception as _exc:
            _log.warning("restore_auto_sync_on_startup: stock_scheduler start failed: %s", _exc)

    # ── Stock module endpoints ────────────────────────────────────────────────

    @app.get("/api/stock/sources")
    def list_stock_sources(request: Request) -> dict[str, object]:
        user = _require_user(request)
        sources = repository.list_stock_sources(user_id=int(user["id"]))
        return {"items": sources, "count": len(sources)}

    @app.post("/api/stock/sources")
    def create_stock_source(request: Request, payload: StockSourceCreateRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        extra = {}
        if payload.client_id:
            extra["client_id"] = payload.client_id
        source = repository.create_stock_source(
            user_id=int(user["id"]),
            marketplace=payload.marketplace.lower().strip(),
            account_name=payload.account_name.strip(),
            api_url=payload.api_url.strip(),
            api_key=payload.api_key.strip(),
            extra=extra,
            interval_hours=payload.interval_hours,
            retention_days=payload.retention_days,
        )
        return {"ok": True, "item": source}

    @app.put("/api/stock/sources/{source_id}")
    def update_stock_source(source_id: int, request: Request, payload: StockSourceUpdateRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        uid = int(user["id"])
        extra_update = None
        if payload.client_id is not None:
            src = repository.get_stock_source(user_id=uid, source_id=source_id, include_secrets=False)
            if src:
                ex = dict(src.get("extra") or {})
                ex["client_id"] = payload.client_id
                extra_update = ex
        updated = repository.update_stock_source(
            user_id=uid,
            source_id=source_id,
            account_name=payload.account_name,
            api_key=payload.api_key,
            interval_hours=payload.interval_hours,
            retention_days=payload.retention_days,
            is_active=payload.is_active,
            extra=extra_update,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True}

    @app.delete("/api/stock/sources/{source_id}")
    def delete_stock_source(source_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_stock_source(user_id=int(user["id"]), source_id=source_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True}

    @app.post("/api/stock/sync")
    def sync_stock_sources(request: Request) -> dict[str, object]:
        """Manually trigger stock sync for all active sources."""
        user = _require_settings_access(request)
        uid = int(user["id"])
        sources = repository.list_stock_sources(user_id=uid, include_secrets=True)
        results = []
        for source in sources:
            if not source.get("is_active"):
                continue
            result = sync_stock_source(source, repository)
            results.append(result)
        return {
            "ok": True,
            "synced": len(results),
            "results": results,
        }

    @app.get("/api/stock/reports")
    def list_stock_reports(request: Request, source_id: int | None = None) -> dict[str, object]:
        user = _require_user(request)
        reports = repository.list_stock_reports(user_id=int(user["id"]), source_id=source_id, limit=100)
        return {"items": reports, "count": len(reports)}

    @app.delete("/api/stock/reports")
    def delete_stock_reports(request: Request, source_id: int | None = None) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_all_stock_reports(user_id=int(user["id"]), source_id=source_id)
        return {"ok": True, "deleted": deleted}

    @app.get("/api/stock/reports/{report_id}/download")
    def download_stock_report(report_id: int, request: Request) -> object:
        from fastapi.responses import FileResponse as _FileResp
        user = _require_user(request)
        reports = repository.list_stock_reports(user_id=int(user["id"]), limit=1000)
        report = next((r for r in reports if int(r.get("id") or 0) == report_id), None)
        if not report:
            raise HTTPException(status_code=404, detail="Отчёт не найден")
        file_path = str(report.get("file_path") or "")
        if not file_path or not Path(file_path).exists():
            raise HTTPException(status_code=404, detail="Файл отчёта не найден на сервере")
        return _FileResp(file_path, filename=Path(file_path).name, media_type="application/json")

    @app.get("/api/stock/data")
    def get_stock_data(request: Request, source_id: int) -> dict[str, object]:
        """Return pivot stock data enriched with product catalog names and zero-fill."""
        user = _require_user(request)
        user_id = int(user["id"])
        owner_id = _tenant_owner_id(user)
        dates = repository.get_stock_data_dates(user_id=user_id, source_id=source_id)
        rows = repository.get_stock_data_pivot(user_id=user_id, source_id=source_id)

        # Determine source marketplace to choose correct catalog lookup key
        source = repository.get_stock_source(user_id=user_id, source_id=source_id)
        is_ozon = str((source or {}).get("marketplace") or "").lower() == "ozon"

        if is_ozon:
            catalog = repository.get_product_catalog_map_ozon(user_id=owner_id)
            # For Ozon: match by wb_article field which stores ozon seller article
            art_key = "wb_article"
        else:
            catalog = repository.get_product_catalog_map(user_id=owner_id)
            art_key = "wb_article"

        if catalog:
            # Build existing (warehouse, article) pairs from report
            existing: set[tuple[str, str]] = {
                (r["warehouse_name"], r.get(art_key, "")) for r in rows
            }
            # Substitute product names
            for r in rows:
                art = r.get(art_key, "")
                if art in catalog:
                    r["seller_article"] = catalog[art]["product_name"] or art
            # Zero-fill: for each warehouse that has data, add missing catalog articles
            warehouses: list[str] = []
            seen_wh: set[str] = set()
            for r in rows:
                wh = r["warehouse_name"]
                if wh not in seen_wh:
                    warehouses.append(wh)
                    seen_wh.add(wh)
            for wh in warehouses:
                for art, cat_item in catalog.items():
                    if (wh, art) not in existing:
                        rows.append({
                            "warehouse_name": wh,
                            "wb_article": art,
                            "seller_article": cat_item["product_name"] or art,
                            "dates": {d: 0 for d in dates},
                        })
        return {"dates": dates, "rows": rows, "count": len(rows)}

    # ── Product catalog endpoints ─────────────────────────────────────────────

    class ProductCatalogItemRequest(BaseModel):
        product_name: str = ""
        wb_article: str = ""
        ozon_article: str = ""

    @app.get("/api/stock/products")
    def list_products(request: Request) -> dict[str, object]:
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        items = repository.list_product_catalog(user_id=owner_id)
        return {"ok": True, "items": items, "count": len(items)}

    @app.post("/api/stock/products")
    def upsert_product(request: Request, payload: ProductCatalogItemRequest) -> dict[str, object]:
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        if not str(payload.wb_article or "").strip():
            raise HTTPException(status_code=400, detail="wb_article is required")
        item = repository.upsert_product_catalog_item(
            user_id=owner_id,
            wb_article=payload.wb_article,
            product_name=payload.product_name,
            ozon_article=payload.ozon_article,
        )
        return {"ok": True, "item": item}

    @app.delete("/api/stock/products/{item_id}")
    def delete_product(request: Request, item_id: int) -> dict[str, object]:
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        deleted = repository.delete_product_catalog_item(user_id=owner_id, item_id=item_id)
        return {"ok": deleted}

    @app.post("/api/stock/products/import")
    async def import_products_excel(
        request: Request,
        file: UploadFile = File(...),
    ) -> dict[str, object]:
        """Import product catalog from Excel. Columns: Наименование товара, Артикул ВБ, Артикул ОЗОН."""
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        try:
            import openpyxl  # type: ignore
            data = await file.read()
            import io
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            # Detect header row
            header = None
            col_name = col_wb = col_ozon = None
            for row in rows_iter:
                cells = [str(c or "").strip().lower() for c in row]
                for i, c in enumerate(cells):
                    if any(x in c for x in ("наименование", "название", "name", "товар")):
                        col_name = i
                    elif any(x in c for x in ("артикул вб", "wb_article", "wb article", "артикул wb", "артикул ВБ".lower())):
                        col_wb = i
                    elif any(x in c for x in ("артикул ozon", "артикул озон", "ozon_article", "ozon article")):
                        col_ozon = i
                if col_name is not None or col_wb is not None:
                    break
            # Fallback: assume columns 0=name, 1=wb, 2=ozon
            if col_name is None: col_name = 0
            if col_wb is None: col_wb = 1
            if col_ozon is None: col_ozon = 2
            imported = 0
            for row in rows_iter:
                if not row or all(c is None for c in row):
                    continue
                def _cell(idx: int) -> str:
                    if idx >= len(row): return ""
                    return str(row[idx] or "").strip()
                wb_art = _cell(col_wb)
                if not wb_art:
                    continue
                repository.upsert_product_catalog_item(
                    user_id=owner_id,
                    wb_article=wb_art,
                    product_name=_cell(col_name),
                    ozon_article=_cell(col_ozon),
                )
                imported += 1
            return {"ok": True, "imported": imported}
        except Exception as exc:
            _log.warning("import_products_excel error: %s", exc)
            raise HTTPException(status_code=400, detail=f"Ошибка парсинга файла: {exc}")

    # ── End stock endpoints ───────────────────────────────────────────────────

    # ── Supply module endpoints ───────────────────────────────────────────────

    def _supply_owner_id(user: dict[str, object]) -> int:
        """Return the owner's user_id for supply queries (same tenant logic)."""
        return _tenant_owner_id(user)

    def _can_view_supplies(user: dict[str, object]) -> bool:
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return True
        return bool(user.get("can_supplies"))

    @app.get("/api/supply-sources")
    def list_supply_sources(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_sources(user_id=owner_id)

    @app.post("/api/supply-sources")
    def create_supply_source(request: Request, payload: CreateSupplySourceRequest) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Только владелец может добавлять источники")
        if not payload.name.strip():
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        if not payload.api_key.strip():
            raise HTTPException(status_code=400, detail="API-ключ не может быть пустым")
        try:
            repository._ensure_supply_tables()
            return repository.create_supply_source(
                user_id=int(user["id"]),
                name=payload.name.strip(),
                api_key=payload.api_key.strip(),
            )
        except Exception as exc:
            _log.error("create_supply_source error: %s", exc, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {exc}")

    @app.patch("/api/supply-sources/{source_id}/toggle")
    def toggle_supply_source(request: Request, source_id: int, payload: ToggleSupplySourceRequest) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.toggle_supply_source(
            user_id=int(user["id"]), source_id=source_id, is_enabled=payload.is_enabled
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True, "source_id": source_id, "is_enabled": payload.is_enabled}

    @app.delete("/api/supply-sources/{source_id}")
    def delete_supply_source(request: Request, source_id: int) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_source(user_id=int(user["id"]), source_id=source_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True}

    @app.get("/api/supplies")
    def list_supplies(
        request: Request,
        source_id: int | None = None,
        status_id: int | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        production: str | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_items(
            user_id=owner_id,
            source_id=source_id,
            status_id=status_id,
            date_from=date_from,
            production=production or None,
            search=search or None,
            date_to=date_to,
            page=page,
            page_size=page_size,
        )

    @app.get("/api/supplies/{supply_id}/goods")
    def get_supply_goods(request: Request, supply_id: int) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        name_map = repository.get_product_name_by_article(user_id=owner_id)

        def _enrich_goods(goods: list[dict]) -> list[dict]:
            for g in goods:
                vc = str(g.get("vendor_code") or "").strip()
                g["product_name"] = name_map.get(vc) or vc or ""
            return goods

        # Check if we have goods cached; if not, fetch from WB and cache
        cached = repository.get_supply_goods(user_id=owner_id, supply_id=supply_id)
        if cached:
            return _enrich_goods(cached)
        # Lazy-fetch from WB API
        try:
            import urllib.request as _ul, json as _jm, ssl as _sl
            def _wb_get(url: str, key: str):
                req = _ul.Request(url, headers={
                    "Authorization": key,
                    "Content-Type": "application/json",
                    "User-Agent": "FeedPilot/1.0",
                }, method="GET")
                ctx = _sl.create_default_context()
                for attempt in range(3):
                    try:
                        with _ul.urlopen(req, timeout=15, context=ctx) as r:
                            return r.status, _jm.loads(r.read() or b"{}")
                    except Exception as e:
                        code = getattr(e, "code", None)
                        if code in (429, 503):
                            import time as _t; _t.sleep((attempt + 1) * 2)
                            continue
                        return (int(code) if code else 0), {}
                return 0, {}

            row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
            if not row:
                return []
            src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(row["source_id"]))
            if not src or not src.get("api_key"):
                return []
            api_key = str(src["api_key"])
            # Fetch details (warehouse, quantity)
            det_status, det_data = _wb_get(
                f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}", api_key
            )
            if det_status == 200 and isinstance(det_data, dict):
                det_data["supplyID"] = supply_id
                repository.upsert_supply_item(source_id=int(row["source_id"]), data=det_data)
            # Fetch goods
            g_status, goods = _wb_get(
                f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}/goods", api_key
            )
            if g_status == 200 and isinstance(goods, list):
                item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
                if item_row:
                    repository.upsert_supply_goods(supply_item_id=int(item_row["id"]), goods=goods)
                    return _enrich_goods(repository.get_supply_goods(user_id=owner_id, supply_id=supply_id))
        except Exception as exc:
            _log.warning("lazy supply goods fetch error supply_id=%d: %s", supply_id, exc)
        return []

    @app.get("/api/supplies/{supply_id}/nm-prices")
    def get_supply_nm_prices(request: Request, supply_id: int) -> dict[str, object]:
        """Return {nmID: discountedPrice} for all goods using the source api_key (same token, prices scope)."""
        import urllib.request as _ul, json as _jm, ssl as _sl
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
        if not row:
            return {"prices": {}}
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(row["source_id"]))
        if not src or not src.get("api_key"):
            return {"prices": {}}
        api_key = str(src["api_key"])
        ctx = _sl.create_default_context()
        prices: dict[str, float] = {}
        offset = 0
        try:
            while True:
                url = f"https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter?limit=1000&offset={offset}"
                req = _ul.Request(url, method="GET", headers={
                    "Authorization": api_key, "User-Agent": "Mozilla/5.0"
                })
                with _ul.urlopen(req, context=ctx, timeout=15) as r:
                    data = _jm.loads(r.read())
                page = data.get("data", {}).get("listGoods", [])
                for g in page:
                    nm = g.get("nmID")
                    sizes = g.get("sizes") or []
                    dp = float(sizes[0].get("discountedPrice", 0)) if sizes else 0.0
                    if nm and dp > 0:
                        prices[str(nm)] = dp
                offset += len(page)
                if len(page) < 1000:
                    break
        except Exception as exc:
            _log.warning("nm-prices fetch error supply_id=%d: %s", supply_id, exc)
        return {"prices": prices}

    @app.get("/api/supplies/{supply_id}/ttn.pdf")
    def get_ttn_pdf(request: Request, supply_id: int):
        """Generate TTN DOCX from same template as download button, convert to PDF via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, zipfile as _zf, io as _io
        import pathlib as _pl
        import urllib.request as _ul, json as _jm, ssl as _sl
        import html as _html_mod
        from fastapi.responses import Response

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)

        # ── Fetch supply data ──────────────────────────────────────────────
        item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
        if not item_row:
            raise HTTPException(status_code=404, detail="Поставка не найдена")

        item = dict(item_row)
        supply_id_str = str(supply_id)

        # ── Legal entity ───────────────────────────────────────────────────
        entities = repository.list_supply_legal_entities(user_id=owner_id)
        supplier_short = str(item.get("supplier_name") or "")
        le = next((e for e in entities if e.get("short_name") == supplier_short), None) or (entities[0] if entities else {})
        org_full = le.get("full_name") or supplier_short
        org_req  = le.get("requisites") or ""
        org_line = ", ".join(filter(None, [org_full, org_req]))

        # ── Dates ──────────────────────────────────────────────────────────
        from datetime import datetime as _dtt
        now = _dtt.now()
        date_disp = now.strftime("%d.%m.%Y")
        raw_sd = str(item.get("supply_date") or "")
        try:
            sd = _dtt.fromisoformat(raw_sd.replace("Z","").split("T")[0]) if raw_sd else now
            supply_date_disp = sd.strftime("%d.%m.%Y")
        except Exception:
            supply_date_disp = date_disp

        driver_name     = str(item.get("driver_name") or "")
        pallets         = int(item.get("pallets_count") or 0)
        VAT_RATE        = 0.22
        wh              = str(item.get("warehouse_name") or "").strip()

        # ── Goods list ─────────────────────────────────────────────────────
        goods_list = repository.get_supply_goods(user_id=owner_id, supply_id=supply_id)
        name_map = repository.get_product_name_by_article(user_id=owner_id)
        for g in goods_list:
            vc = str(g.get("vendor_code") or "")
            g["product_name"] = name_map.get(vc) or vc or ""

        # ── Prices from WB ─────────────────────────────────────────────────
        nm_prices: dict[int, float] = {}
        try:
            src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
            if src and src.get("api_key"):
                api_key = str(src["api_key"])
                ctx = _sl.create_default_context()
                offset = 0
                while True:
                    url = f"https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter?limit=1000&offset={offset}"
                    req = _ul.Request(url, method="GET", headers={"Authorization": api_key, "User-Agent": "Mozilla/5.0"})
                    with _ul.urlopen(req, context=ctx, timeout=15) as r:
                        data = _jm.loads(r.read())
                    page = data.get("data", {}).get("listGoods", [])
                    for g in page:
                        nm = int(g.get("nmID") or 0)
                        sizes = g.get("sizes") or []
                        dp = float(sizes[0].get("discountedPrice", 0)) if sizes else 0.0
                        if nm and dp > 0:
                            nm_prices[nm] = dp
                    offset += len(page)
                    if len(page) < 1000:
                        break
        except Exception as ex:
            _log.warning("ttn-pdf prices fetch: %s", ex)

        # ── Number to Russian words ────────────────────────────────────────
        def _rubles_in_words(n: int) -> str:
            ones_m = ["","один","два","три","четыре","пять","шесть","семь","восемь","девять"]
            ones_f = ["","одна","две","три","четыре","пять","шесть","семь","восемь","девять"]
            teens  = ["десять","одиннадцать","двенадцать","тринадцать","четырнадцать",
                      "пятнадцать","шестнадцать","семнадцать","восемнадцать","девятнадцать"]
            tens   = ["","","двадцать","тридцать","сорок","пятьдесят","шестьдесят","семьдесят","восемьдесят","девяносто"]
            hunds  = ["","сто","двести","триста","четыреста","пятьсот","шестьсот","семьсот","восемьсот","девятьсот"]
            def chunk(x, fem):
                r,w = x%100,[]
                h = x//100
                if h: w.append(hunds[h])
                if r>=10 and r<=19: w.append(teens[r-10])
                else:
                    if r//10: w.append(tens[r//10])
                    d = r%10
                    if d: w.append((ones_f if fem else ones_m)[d])
                return w
            if n==0: return "ноль рублей 00 копеек"
            w=[]
            bn = n//1000000000
            mn = (n//1000000)%1000
            th = (n//1000)%1000
            ru = n%1000
            if bn:
                bw=chunk(bn,False)
                w.extend(bw)
                w.append(["миллиардов","миллиард","миллиарда","миллиардов"][1 if bn%10==1 and bn%100!=11 else 2 if bn%10 in(2,3,4) and bn%100 not in range(12,15) else 0 if bn==0 else 3])
            if mn:
                mw=chunk(mn,False)
                w.extend(mw)
                w.append(["миллионов","миллион","миллиона","миллионов"][1 if mn%10==1 and mn%100!=11 else 2 if mn%10 in(2,3,4) and mn%100 not in range(12,15) else 0 if mn==0 else 3])
            if th:
                tw2=chunk(th,True)
                w.extend(tw2)
                w.append(["тысяч","тысяча","тысячи","тысяч"][1 if th%10==1 and th%100!=11 else 2 if th%10 in(2,3,4) and th%100 not in range(12,15) else 0 if th==0 else 3])
            if ru:
                w.extend(chunk(ru,False))
            rub_w = ["рублей","рубль","рубля","рублей"][1 if ru%10==1 and ru%100!=11 else 2 if ru%10 in(2,3,4) and ru%100 not in range(12,15) else 0 if ru==0 else 3]
            w.append(rub_w)
            w.append("00 копеек")
            return " ".join(w)

        def fmt2(x: float) -> str:
            return f"{x:,.2f}".replace(",", " ").replace(".", ",")

        # ── Build per-row data ─────────────────────────────────────────────
        total_excl = total_vat = total_incl = 0.0
        qty_total = sum(int(g.get("quantity") or 0) for g in goods_list) or pallets

        rows_data = []
        for i, g in enumerate(goods_list):
            qty = int(g.get("quantity") or 0)
            nm  = int(g.get("nm_id") or 0)
            price_incl = nm_prices.get(nm)
            price_excl = price_incl / (1 + VAT_RATE) if price_incl else None
            amt_excl   = price_excl * qty if price_excl is not None else None
            vat_amt    = amt_excl * VAT_RATE if amt_excl is not None else None
            amt_incl   = amt_excl + vat_amt if amt_excl is not None else None
            if amt_excl is not None:
                total_excl += amt_excl; total_vat += vat_amt; total_incl += amt_incl
            rows_data.append({
                "num": i + 1,
                "name": g.get("product_name") or g.get("vendor_code") or "Товар",
                "qty": qty,
                "price_excl": fmt2(price_excl) if price_excl is not None else "—",
                "amt_excl":   fmt2(amt_excl)   if amt_excl   is not None else "—",
                "vat_amt":    fmt2(vat_amt)     if vat_amt    is not None else "—",
                "amt_incl":   fmt2(amt_incl)    if amt_incl   is not None else "—",
            })

        t_excl = fmt2(total_excl) if total_excl else "—"
        t_vat  = fmt2(total_vat)  if total_vat  else "—"
        t_incl = fmt2(total_incl) if total_incl else "—"
        amt_words = _rubles_in_words(round(total_incl)) if total_incl else "—"

        # Generate DOCX from same template as "Скачать ТТН" button
        import zipfile as _zf, io as _io, re as _re, html as _html_esc
        tpl_path = STATIC_DIR / "torg12_tpl.docx"
        with open(tpl_path, "rb") as f:
            tpl_bytes = f.read()
        with _zf.ZipFile(_io.BytesIO(tpl_bytes)) as zin:
            all_files = {name: zin.read(name) for name in zin.namelist()}
        doc_xml = all_files["word/document.xml"].decode("utf-8")

        # Row duplication
        row_rx = _re.compile(r'(<w:tr[\s>](?:(?!</w:tr>).)*?\{\{GOODS_NAME\}\}.*?</w:tr>)', _re.DOTALL)
        m = row_rx.search(doc_xml)
        if m and rows_data:
            row_tpl = m.group(1)
            multi = ""
            for rd in rows_data:
                r = row_tpl
                r = r.replace("{{ROW_NUM}}",         str(rd["num"]))
                r = r.replace("{{GOODS_NAME}}",       _html_esc.escape(rd["name"]))
                r = r.replace("{{PRICE}}",            _html_esc.escape(rd["price_excl"]))
                r = r.replace("{{ROW_AMOUNT_EXCL}}",  _html_esc.escape(rd["amt_excl"]))
                r = r.replace("{{ROW_VAT_SUM}}",      _html_esc.escape(rd["vat_amt"]))
                r = r.replace("{{ROW_AMOUNT_INCL}}",  _html_esc.escape(rd["amt_incl"]))
                r = r.replace("{{ROW_QTY}}",          str(rd["qty"]))
                multi += r
            doc_xml = doc_xml.replace(row_tpl, multi, 1)

        for ph, val in [
            ("{{TTN_NUMBER}}",      supply_id_str),
            ("{{ORG_FULL}}",        org_line),
            ("{{SUPPLIER}}",        org_line),
            ("{{PAYER}}",           org_line),
            ("{{ORDER_DATE}}",      supply_id_str),
            ("{{DOC_NUM_VAL}}",     supply_id_str),
            ("{{DOC_DATE_VAL}}",    supply_date_disp),
            ("{{GOODS_NAME}}",      rows_data[0]["name"] if rows_data else "Товар"),
            ("{{ROW_NUM}}",         "1"),
            ("{{PRICE}}",           rows_data[0]["price_excl"] if rows_data else "—"),
            ("{{ROW_AMOUNT_EXCL}}", rows_data[0]["amt_excl"] if rows_data else "—"),
            ("{{ROW_VAT_SUM}}",     rows_data[0]["vat_amt"] if rows_data else "—"),
            ("{{ROW_AMOUNT_INCL}}", rows_data[0]["amt_incl"] if rows_data else "—"),
            ("{{QTY}}",             str(qty_total)),
            ("{{QTY_SHT}}",         f"{qty_total} шт"),
            ("{{TOTAL_EXCL}}",      t_excl),
            ("{{TOTAL_VAT}}",       t_vat),
            ("{{TOTAL_INCL}}",      t_incl),
            ("{{AMOUNT}}",          t_excl),
            ("{{VAT_SUM}}",         t_vat),
            ("{{AMOUNT_WITH_VAT}}", t_incl),
            ("{{AMOUNT_WORDS}}",    amt_words),
            ("{{PAGES_COUNT}}",     "1"),
            ("{{ITEMS_COUNT}}",     str(len(rows_data) or 1)),
            ("{{SUPPLY_ID}}",       supply_id_str),
            ("{{DOC_DATE_FULL}}",   f"«{now.strftime('%d')}» {['января','февраля','марта','апреля','мая','июня','июля','августа','сентября','октября','ноября','декабря'][now.month-1]} {now.year}"),
            ("{{ISSUED_BY}}",       org_line or supplier_short or "—"),
            ("{{SIGN_SUPPLIER}}",   supplier_short),
            ("{{SIGN_DRIVER}}",     driver_name),
        ]:
            doc_xml = doc_xml.replace(ph, val)
        doc_xml = doc_xml.replace("{{ROW_QTY}}", str(qty_total))

        all_files["word/document.xml"] = doc_xml.encode("utf-8")

        # Write DOCX to temp file and convert to PDF via LibreOffice
        import subprocess as _sp, tempfile as _tf, pathlib as _pl
        tmp_dir   = _tf.mkdtemp()
        docx_path = _pl.Path(tmp_dir) / f"ttn_{supply_id}.docx"
        pdf_path  = _pl.Path(tmp_dir) / f"ttn_{supply_id}.pdf"

        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w", _zf.ZIP_DEFLATED) as zout:
            for name, data in all_files.items():
                zout.writestr(name, data)
        docx_path.write_bytes(buf.getvalue())

        # Try soffice / libreoffice
        import os as _os
        # LibreOffice requires writable XDG dirs — force all to tmp_dir
        lo_env = dict(_os.environ)
        lo_env["HOME"]            = tmp_dir
        lo_env["TMPDIR"]          = tmp_dir
        lo_env["XDG_CACHE_HOME"]  = tmp_dir
        lo_env["XDG_CONFIG_HOME"] = tmp_dir
        lo_env["XDG_RUNTIME_DIR"] = tmp_dir
        lo_env["DCONF_PROFILE"]   = "/dev/null"
        lo_env["UserInstallation"] = f"file://{tmp_dir}/lo_profile"

        _lo_binaries = (
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            "/usr/local/bin/soffice",
            "soffice",
            "/usr/bin/libreoffice",
            "libreoffice",
        )
        lo_ok = False
        for binary in _lo_binaries:
            try:
                result = _sp.run(
                    [binary, "--headless", "--norestore",
                     f"-env:UserInstallation=file://{tmp_dir}/lo_profile",
                     "--convert-to", "pdf",
                     "--outdir", tmp_dir, str(docx_path)],
                    capture_output=True, timeout=60, env=lo_env
                )
                _log.info("soffice %s exit=%d stdout=%s stderr=%s",
                          binary, result.returncode,
                          result.stdout.decode()[:200], result.stderr.decode()[:200])
                if result.returncode == 0 and pdf_path.exists():
                    lo_ok = True
                    break
            except FileNotFoundError:
                continue
            except _sp.TimeoutExpired:
                raise HTTPException(status_code=504, detail="Таймаут конвертации PDF")

        if not lo_ok:
            raise HTTPException(status_code=500,
                detail=f"LibreOffice не смог конвертировать DOCX в PDF. Убедитесь что libreoffice-writer установлен")

        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail="PDF не создан")

        pdf_bytes = pdf_path.read_bytes()
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="TTN_{supply_id}.pdf"'},
        )

    @app.get("/api/supplies/{supply_id}/ttn-error-test")
    def ttn_error_test(request: Request, supply_id: int) -> dict[str, object]:
        """Debug: run ttn.pdf and return error detail instead of 500."""
        import subprocess as _sp, tempfile as _tf, zipfile as _zf, io as _io
        import re as _re, pathlib as _pl, traceback as _tb
        from fastapi.responses import FileResponse
        try:
            return get_ttn_pdf(request, supply_id)
        except HTTPException as e:
            return {"ok": False, "status": e.status_code, "detail": e.detail}
        except Exception as ex:
            return {"ok": False, "error": str(ex), "trace": _tb.format_exc()[-2000:]}

    @app.delete("/api/supplies")
    def clear_supplies(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        deleted = repository.clear_supply_items(user_id=owner_id)
        return {"ok": True, "deleted": deleted}

    @app.get("/api/supply-drivers")
    def list_supply_drivers(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_drivers(user_id=owner_id)

    @app.post("/api/supply-drivers")
    def create_supply_driver(request: Request, payload: CreateSupplyDriverRequest) -> dict[str, object]:
        user = _require_user(request)
        # Drivers: accessible to owners AND managers with can_supplies
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.full_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Имя не может быть пустым")
        # Always save under owner's user_id so drivers are shared across team
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        if repository.driver_exists(user_id=owner_id, full_name=name):
            raise HTTPException(status_code=409, detail=f"Водитель «{name}» уже существует")
        return repository.create_supply_driver(user_id=owner_id, full_name=name, documents=payload.documents)

    @app.patch("/api/supply-drivers/{driver_id}")
    def update_supply_driver_endpoint(request: Request, driver_id: int, payload: UpdateSupplyDriverRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.full_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Имя не может быть пустым")
        ok = repository.update_supply_driver(user_id=_supply_owner_id(user), driver_id=driver_id, full_name=name, documents=payload.documents)
        if not ok:
            raise HTTPException(status_code=404, detail="Водитель не найден")
        return {"ok": True}

    @app.delete("/api/supply-drivers/{driver_id}")
    def delete_supply_driver(request: Request, driver_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        ok = repository.delete_supply_driver(user_id=owner_id, driver_id=driver_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Водитель не найден")
        return {"ok": True}

    @app.get("/api/supply-warehouses")
    def list_supply_warehouses(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_warehouses(user_id=owner_id)

    @app.post("/api/supply-warehouses")
    def create_supply_warehouse(request: Request, payload: CreateSupplyWarehouseRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.warehouse_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Название склада не может быть пустым")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.create_supply_warehouse(user_id=owner_id, warehouse_name=name, address=payload.address.strip())

    @app.patch("/api/supply-warehouses/{warehouse_id}")
    def update_supply_warehouse_endpoint(request: Request, warehouse_id: int, payload: UpdateSupplyWarehouseRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.warehouse_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        ok = repository.update_supply_warehouse(user_id=_supply_owner_id(user), warehouse_id=warehouse_id, warehouse_name=name, address=payload.address.strip())
        if not ok:
            raise HTTPException(status_code=404, detail="Склад не найден")
        return {"ok": True}

    @app.delete("/api/supply-warehouses/{warehouse_id}")
    def delete_supply_warehouse(request: Request, warehouse_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_warehouse(user_id=_supply_owner_id(user), warehouse_id=warehouse_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Склад не найден")
        return {"ok": True}

    @app.get("/api/supply-legal-entities")
    def list_supply_legal_entities(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_legal_entities(user_id=owner_id)

    @app.post("/api/supply-legal-entities")
    def create_supply_legal_entity(request: Request, payload: CreateSupplyLegalEntityRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.short_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Короткое наименование не может быть пустым")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.create_supply_legal_entity(user_id=owner_id, short_name=name, full_name=payload.full_name.strip(), requisites=payload.requisites, signatories=payload.signatories)

    @app.patch("/api/supply-legal-entities/{entity_id}")
    def update_supply_legal_entity_endpoint(request: Request, entity_id: int, payload: UpdateSupplyLegalEntityRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.short_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Короткое наименование не может быть пустым")
        ok = repository.update_supply_legal_entity(user_id=_supply_owner_id(user), entity_id=entity_id, short_name=name, full_name=payload.full_name.strip(), requisites=payload.requisites, signatories=payload.signatories)
        if not ok:
            raise HTTPException(status_code=404, detail="Юридическое лицо не найдено")
        return {"ok": True}

    @app.delete("/api/supply-legal-entities/{entity_id}")
    def delete_supply_legal_entity(request: Request, entity_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_legal_entity(user_id=_supply_owner_id(user), entity_id=entity_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Юридическое лицо не найдено")
        return {"ok": True}

    @app.patch("/api/supplies/{supply_id}/manual-fields")
    def update_supply_manual_fields(
        request: Request,
        supply_id: int,
        payload: SupplyManualFieldsRequest,
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        ok = repository.update_supply_manual_fields(
            user_id=owner_id,
            supply_id=supply_id,
            pass_number=payload.pass_number,
            pallets_count=payload.pallets_count,
            driver_name=payload.driver_name,
            notes=payload.notes,
            production=payload.production,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        return {"ok": True}

    @app.get("/api/supplies/sync/status")
    def get_supply_sync_status(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        with supply_sync_lock:
            return dict(supply_sync_state)

    @app.post("/api/supplies/sync")
    def sync_supplies(request: Request, payload: SyncSuppliesRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()

        with supply_sync_lock:
            if supply_sync_state.get("in_progress"):
                # Reset if stuck (thread died without cleanup)
                started = str(supply_sync_state.get("started_at") or "")
                if started:
                    from datetime import datetime as _dt2, timezone as _tz2
                    try:
                        age = (_dt2.now(_tz2.utc) - _dt2.fromisoformat(started.replace("Z", "+00:00"))).seconds
                        if age < 600:  # allow reset after 10 min
                            return {"ok": False, "message": "Синхронизация уже запущена"}
                    except Exception:
                        pass
                supply_sync_state["in_progress"] = False

        sources = repository.list_supply_sources(user_id=owner_id)
        if payload.source_id:
            sources = [s for s in sources if s["id"] == payload.source_id]
        active_sources = [s for s in sources if s.get("is_enabled")]
        if not active_sources:
            return {"ok": True, "synced": 0, "message": "Нет активных источников"}

        import urllib.request as _urllib
        import json as _json_mod
        import ssl as _ssl
        import time as _time
        from datetime import datetime as _dt, timezone as _tz, timedelta as _td

        def _wb_request(method: str, url: str, api_key: str, body: dict | None = None):
            data = _json_mod.dumps(body).encode() if body else None
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json",
                "User-Agent": "FeedPilot/1.0",
            }
            req = _urllib.Request(url, data=data, headers=headers, method=method)
            ctx = _ssl.create_default_context()
            # Retry up to 3 times with backoff: handles 429, 503, network errors
            for attempt in range(3):
                try:
                    with _urllib.urlopen(req, timeout=30, context=ctx) as r:
                        return r.status, _json_mod.loads(r.read() or b"{}")
                except Exception as e:
                    code = getattr(e, "code", None)
                    if code in (429, 503):
                        # Rate limited or service unavailable — wait and retry
                        wait = (attempt + 1) * 2  # 2s, 4s, 6s
                        _log.warning("WB supplies API %d, retry %d in %ds", code, attempt + 1, wait)
                        _time.sleep(wait)
                        continue
                    return (int(code) if code else 0), {}
            return 0, {}

        def _wb_post(url: str, api_key: str, body: dict):
            s, d = _wb_request("POST", url, api_key, body)
            return s, d if isinstance(d, list) else []

        def _run_sync():
            _now = _dt.now(_tz.utc)
            date_from = (_now - _td(days=30)).strftime("%Y-%m-%d")
            date_to = (_now + _td(days=1)).strftime("%Y-%m-%d")
            active_statuses = {1, 2, 3, 4, 5}  # 3 = Отгрузка разрешена
            total_synced = 0
            errors: list[str] = []

            with supply_sync_lock:
                supply_sync_state.update({
                    "in_progress": True, "page": 0, "synced": 0, "total": 0,
                    "errors": [], "message": "Запуск…",
                    "started_at": _dt.now(_tz.utc).isoformat(), "finished_at": None,
                })

            try:
                for src in active_sources:
                    src_full = repository.get_supply_source_with_key(
                        user_id=owner_id, source_id=int(src["id"])
                    )
                    if not src_full:
                        continue
                    api_key = str(src_full.get("api_key") or "")
                    if not api_key:
                        errors.append(f"Источник «{src['name']}»: нет API-ключа")
                        continue
                    source_id = int(src["id"])
                    synced_this_source = 0
                    with supply_sync_lock:
                        supply_sync_state["message"] = f"«{src['name']}»: загрузка списка…"
                        supply_sync_state["page"] = 1
                    try:
                        # WB API ignores dateFrom/dateTo — always returns ALL supplies.
                        # Fetch once (API also ignores pagination) and filter client-side.
                        http_status, items = _wb_post(
                            "https://supplies-api.wildberries.ru/api/v1/supplies",
                            api_key,
                            {"dateFrom": "2020-01-01", "dateTo": "2099-12-31",
                             "status": "ALL", "page": 1, "pageSize": 1000},
                        )
                        if http_status == 401:
                            errors.append(f"«{src['name']}»: неверный API-ключ")
                        elif isinstance(items, list):
                            # Client-side filter: supplyDate >= last 30 days
                            items = [
                                x for x in items
                                if (x.get("supplyDate") or "")[:10] >= date_from
                                and int(x.get("statusID") or 0) in active_statuses
                            ]
                            with supply_sync_lock:
                                supply_sync_state["total"] = len(items)
                                supply_sync_state["message"] = (
                                    f"«{src['name']}»: найдено {len(items)} поставок, загрузка деталей…"
                                )
                            item_errors = 0
                            for item in items:
                                supply_wb_id = int(item.get("supplyID") or 0)
                                if not supply_wb_id:
                                    continue
                                status_id = int(item.get("statusID") or 0)
                                if status_id not in active_statuses:
                                    continue
                                try:
                                    # For active (1,2,4): always fetch details
                                    # For accepted (5): fetch only if not already cached
                                    need_details = status_id in {1, 2, 3, 4}
                                    if not need_details:
                                        existing = repository.get_supply_item_row(
                                            user_id=owner_id, supply_id=supply_wb_id
                                        )
                                        need_details = not (existing and existing.get("warehouse_name"))
                                    if need_details:
                                        det_status, det_data = _wb_request(
                                            "GET",
                                            f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_wb_id}",
                                            api_key,
                                        )
                                        if det_status == 200 and isinstance(det_data, dict):
                                            item.update({k: v for k, v in det_data.items() if v is not None})
                                        _time.sleep(0.15)
                                    item["supplyID"] = supply_wb_id
                                    repository.upsert_supply_item(source_id=source_id, data=item)
                                    synced_this_source += 1
                                    with supply_sync_lock:
                                        supply_sync_state["synced"] = total_synced + synced_this_source
                                except Exception as item_exc:
                                    item_errors += 1
                                    err_msg = f"{type(item_exc).__name__}: {item_exc}"
                                    _log.error("supply upsert error supply_id=%s: %s", supply_wb_id, err_msg, exc_info=True)
                                    if item_errors == 1:
                                        # Show first error in status
                                        with supply_sync_lock:
                                            supply_sync_state["message"] = f"Ошибка поставки {supply_wb_id}: {err_msg}"
                                        errors.append(f"Поставка {supply_wb_id}: {err_msg}")
                            repository.mark_supply_source_synced(source_id=source_id)
                            total_synced += synced_this_source
                            # Restore manually-entered fields (pass, pallets, driver)
                            # that survived a previous clear_supply_items call.
                            try:
                                repository.restore_supply_manual_fields(user_id=owner_id)
                            except Exception:
                                pass
                    except Exception as exc:
                        _log.error("supply sync source %d: %s", source_id, exc, exc_info=True)
                        err_msg = f"{type(exc).__name__}: {exc}"
                        errors.append(f"«{src['name']}»: {err_msg}")
                        with supply_sync_lock:
                            supply_sync_state["message"] = f"Ошибка: {err_msg}"
            finally:
                with supply_sync_lock:
                    supply_sync_state.update({
                        "in_progress": False,
                        "synced": total_synced,
                        "errors": errors,
                        "message": f"Готово. Загружено {total_synced} поставок." + (
                            f" Ошибки: {'; '.join(errors)}" if errors else ""
                        ),
                        "finished_at": _dt.now(_tz.utc).isoformat(),
                    })

        t = threading.Thread(target=_run_sync, daemon=True)
        t.start()
        return {"ok": True, "started": True, "message": "Синхронизация запущена"}

    # ── End supply module endpoints ───────────────────────────────────────────

    @app.on_event("shutdown")
    def stop_auto_sync_worker() -> None:
        auto_sync_stop_event.set()
        stock_scheduler.stop()
        worker = auto_sync_worker.get("thread")
        if isinstance(worker, threading.Thread) and worker.is_alive():
            worker.join(timeout=1.5)

    @app.post("/api/admin/reviews-clear")
    def admin_clear_reviews(request: Request, payload: ClearReviewsRequest) -> dict[str, object]:
        actor = _require_admin(request)
        if payload.user_id is None:
            target_user_id = _tenant_owner_id(actor) if not _is_super_admin(actor) else int(actor["id"])
        else:
            target_user_id = int(payload.user_id)
            _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        deleted = repository.clear_reviews(user_id=target_user_id)
        return {"ok": True, "deleted": deleted, "user_id": target_user_id}

    @app.post("/api/admin/conversations-clear")
    def admin_clear_conversations_v2(request: Request, payload: ClearConversationsRequest) -> dict[str, object]:
        actor = _require_admin(request)
        if payload.user_id is None:
            target_user_id = _tenant_owner_id(actor) if not _is_super_admin(actor) else int(actor["id"])
        else:
            target_user_id = int(payload.user_id)
            _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        deleted = repository.clear_conversations(user_id=target_user_id, kind=payload.kind, source=payload.source)
        return {"ok": True, "deleted": deleted, "user_id": target_user_id}

    @app.exception_handler(HTTPException)
    def http_exception_handler(_: Request, exc: HTTPException) -> JSONResponse:
        return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

    return app


app = create_app()

def _render_template(name: str, context: dict[str, str] | None = None) -> str:
    template_path = TEMPLATES_DIR / name
    html = template_path.read_text(encoding="utf-8")
    for key, value in (context or {}).items():
        html = html.replace(f"{{{{{key}}}}}", value)
    return html


def build_landing_html() -> str:
    return _render_template("landing.html")


def build_login_html(error: str | None = None) -> str:
    error_html = f"<p class='error'>{escape(error)}</p>" if error else ""
    return _render_template("login.html", {"ERROR_HTML": error_html})


def build_register_html(error: str | None = None) -> str:
    error_html = f"<p class='error'>{escape(error)}</p>" if error else ""
    return _render_template("register.html", {"ERROR_HTML": error_html})


def build_app_html(user: dict[str, object], repository=None) -> str:
    safe_email = escape(str(user["email"]))
    role = str(user.get("role") or ROLE_USER)
    is_super_admin = bool(user.get("is_super_admin"))
    user_id = int(user.get("id") or 0)
    owner_user_id = int(user.get("owner_user_id") or user_id or 0)
    is_tenant_owner = (
        role in ROLE_CAN_ACCESS_SETTINGS
        and user_id > 0
        and owner_user_id == user_id
    )
    role_labels = {
        ROLE_ADMIN: "администратор",
        ROLE_USER: "пользователь",
        ROLE_FEEDBACK_MANAGER: "менеджер обратной связи",
    }
    safe_role = escape(role_labels.get(role, role))
    can_view_analytics = role in ROLE_CAN_ACCESS_ANALYTICS
    can_view_settings = role in ROLE_CAN_ACCESS_SETTINGS
    can_view_supplies = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or bool(user.get("can_supplies"))
    )
    if role in ROLE_CAN_ACCESS_SETTINGS:
        can_view_feedback = True
    elif repository is not None:
        # Manager: feedback sections visible only if they have at least one permission
        _perms = repository.list_manager_permissions(manager_user_id=user_id)
        can_view_feedback = any(
            bool(p.get("can_reviews")) or bool(p.get("can_questions")) or bool(p.get("can_chats"))
            for p in _perms
        )
    else:
        can_view_feedback = True  # safe fallback
    admin_link = '<a class="navbtn nav-admin" href="/admin"><span class="nav-item-icon">○</span> Админ-панель</a>' if role == ROLE_ADMIN else ""
    nav_analytics = (
        '<a id="nav-analytics" class="nav-item" href="#" onclick="showSection(\'analytics\')"><span class="nav-item-icon">∑</span> Аналитика</a>'
        if can_view_analytics
        else ""
    )
    nav_settings = (
        '<a id="nav-settings" class="navbtn" href="#" onclick="showSection(\'settings\')">Настройки</a>'
        if can_view_settings
        else ""
    )
    nav_supplies_wb = (
        '<a id="nav-supplies-wb" class="nav-item" href="#" onclick="showSection(\'supplies-wb\')"><span class="nav-item-icon">▦</span> WB</a>'
        if can_view_supplies else ""
    )
    nav_supplies_settings = (
        '<a id="nav-supplies-settings" class="nav-item" href="#" onclick="showSection(\'supplies-settings\')"><span class="nav-item-icon">≡</span> Настройки</a>'
        if (can_view_settings or can_view_supplies) else ""
    )
    return _render_template(
        "app.html",
        {
            "SAFE_EMAIL": safe_email,
            "SAFE_ROLE": safe_role,
            "ADMIN_LINK": admin_link,
            "NAV_ANALYTICS": nav_analytics,
            "NAV_SETTINGS": nav_settings,
            "NAV_SETTINGS_SUB": (
                '<a id="nav-settings" class="nav-item" href="#" onclick="showSection(\'settings\')"><span class="nav-item-icon">≡</span> Настройки</a>'
                if can_view_settings else ""
            ),
            "NAV_SUPPLIES_WB": nav_supplies_wb,
            "NAV_SUPPLIES_SETTINGS": nav_supplies_settings,
            "CAN_VIEW_ANALYTICS": "true" if can_view_analytics else "false",
            "CAN_VIEW_SETTINGS": "true" if can_view_settings else "false",
            "CAN_VIEW_SUPPLIES": "true" if can_view_supplies else "false",
            "CAN_VIEW_FEEDBACK": "true" if can_view_feedback else "false",
            "IS_ADMIN": "true" if role == ROLE_ADMIN else "false",
            "IS_SUPER_ADMIN": "true" if is_super_admin else "false",
            "IS_TENANT_OWNER": "true" if is_tenant_owner else "false",
        },
    )


def build_admin_html(user: dict[str, object]) -> str:
    safe_email = escape(str(user["email"]))
    return _render_template("admin.html", {"SAFE_EMAIL": safe_email})
